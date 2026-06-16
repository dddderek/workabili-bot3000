import os
import threading
from dataclasses import dataclass
from datetime import datetime

import yaml
from openpyxl import load_workbook, Workbook

from PySide6.QtCore import Qt, QObject, Signal, Slot, QPoint, QRect, QEasingCurve, QPropertyAnimation, QParallelAnimationGroup, Property, QSize, QByteArray, QUrl, QSettings, QEvent, QTimer
from PySide6.QtGui import QFont, QIcon, QPalette, QColor, QFontInfo, QFontMetrics, QPixmap, QPainter, QImage, QDesktopServices, QTextCursor
from PySide6.QtWidgets import (
    QApplication,
    QMainWindow,
    QWidget,
    QLabel,
    QLineEdit,
    QPushButton,
    QCheckBox,
    QFileDialog,
    QPlainTextEdit,
    QTextEdit,
    QMessageBox,
    QDialog,
    QHBoxLayout,
    QVBoxLayout,
    QGridLayout,
    QGroupBox,
    QFrame,
    QToolButton,
    QSizePolicy,
    QStyle,
    QGraphicsOpacityEffect,
    QGraphicsDropShadowEffect,
    QTabWidget,
    QTableWidget,
    QTableWidgetItem,
    QHeaderView,
    QAbstractItemView,
    QComboBox,
)

from PySide6.QtSvg import QSvgRenderer
from PySide6.QtMultimedia import QSoundEffect

from app.runner import run as run_runner, run_serve as run_serve_runner


CONFIG_PATH = os.path.join("config", "config.yaml")

# Splash video (WebM VP9 with alpha)
SPLASH_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "assets", "Shrimpzilla_Intro.webm")
)

# Splash image (transparent PNG) — shown after the video ends
SPLASH_PNG_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "assets", "splash.png")
)

# Splash sound (WAV recommended) — place at: Workabili-Bot3000/assets/splash.wav
SPLASH_SOUND_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "assets", "splash.wav")
)

# Title image (transparent PNG)
TITLE_IMAGE_PATH = os.path.normpath(
    os.path.join(os.path.dirname(__file__), "..", "assets", "workabili-bot3000-title.png")
)


def load_config():
    with open(CONFIG_PATH, "r", encoding="utf-8") as f:
        return yaml.safe_load(f)


def apply_modern_palette(app: QApplication):
    """
    Light-but-modern palette that looks clean with Fusion style.
    """
    app.setStyle("Fusion")
    pal = QPalette()

    # Base surfaces
    pal.setColor(QPalette.Window, QColor("#0f172a"))          # slate-900
    pal.setColor(QPalette.Base, QColor("#0b1224"))            # deeper base
    pal.setColor(QPalette.AlternateBase, QColor("#111b34"))
    pal.setColor(QPalette.Button, QColor("#1e293b"))          # slate-800
    pal.setColor(QPalette.ToolTipBase, QColor("#0b1224"))
    pal.setColor(QPalette.ToolTipText, QColor("#e2e8f0"))

    # Text
    pal.setColor(QPalette.WindowText, QColor("#e2e8f0"))
    pal.setColor(QPalette.Text, QColor("#e2e8f0"))
    pal.setColor(QPalette.ButtonText, QColor("#e2e8f0"))
    pal.setColor(QPalette.PlaceholderText, QColor("#94a3b8"))  # slate-400

    # Highlights
    pal.setColor(QPalette.Highlight, QColor("#38bdf8"))        # sky-400
    pal.setColor(QPalette.HighlightedText, QColor("#0b1224"))

    app.setPalette(pal)


class _SplashClickFilter(QObject):
    """Global click catcher: any mouse click anywhere dismisses the splash."""
    def __init__(self, app: QApplication, splash: QWidget):
        super().__init__()
        self._app = app
        self._splash = splash

    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            try:
                if self._splash and self._splash.isVisible():
                    self._splash.close()
            except Exception:
                pass
            try:
                self._app.removeEventFilter(self)
            except Exception:
                pass
        return False


class PngSplashScreen(QWidget):
    """
    Transparent top-level splash displaying the robot PNG.
    Shown behind the video splash; click to dismiss after the video ends.
    """
    def __init__(self, png_path: str):
        super().__init__(None, Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground, True)

        self._pix = QPixmap(png_path) if png_path else QPixmap()
        self._max_scale = 1.03
        self._drift_px = 15.0
        self._progress = 0.0
        self._anim_started = False
        self._fade_anim = None

        if self._pix.isNull():
            self.resize(1, 1)
        else:
            w = int(self._pix.width() * self._max_scale + 0.999)
            h = int(self._pix.height() * self._max_scale + 0.999)
            self.setFixedSize(max(1, w), max(1, h))

    def showEvent(self, event):
        super().showEvent(event)
        screen = QApplication.screenAt(self.pos()) or QApplication.primaryScreen()
        if screen:
            g = screen.availableGeometry()
            self.move(
                g.left() + (g.width() - self.width()) // 2,
                g.top() + (g.height() - self.height()) // 2,
            )
        if self._pix.isNull() or self._anim_started:
            return
        self._anim_started = True
        self.setWindowOpacity(1.0)
        self._fade_anim = QPropertyAnimation(self, b"progress", self)
        self._fade_anim.setDuration(2000)
        self._fade_anim.setStartValue(0.0)
        self._fade_anim.setEndValue(1.0)
        self._fade_anim.setEasingCurve(QEasingCurve.InOutQuad)
        self._fade_anim.start()

    def paintEvent(self, event):
        if self._pix.isNull():
            return
        t = max(0.0, min(1.0, float(self._progress or 0.0)))
        motion = QEasingCurve(QEasingCurve.OutCubic).valueForProgress(t)
        dy = self._drift_px * (1.0 - motion)
        scale = self._max_scale - (self._max_scale - 1.0) * motion
        painter = QPainter(self)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
        painter.setOpacity(t)
        cx, cy = self.width() / 2.0, self.height() / 2.0
        painter.translate(cx, cy + dy)
        painter.scale(scale, scale)
        painter.translate(-self._pix.width() / 2.0, -self._pix.height() / 2.0)
        painter.drawPixmap(0, 0, self._pix)
        painter.end()

    def mousePressEvent(self, event):
        try:
            if self._fade_anim:
                self._fade_anim.stop()
        except Exception:
            pass
        self.close()
        return super().mousePressEvent(event)

    def _get_progress(self) -> float:
        return float(self._progress or 0.0)

    def _set_progress(self, v: float) -> None:
        self._progress = float(v or 0.0)
        self.update()

    progress = Property(float, _get_progress, _set_progress)


class SplashScreen(QWidget):
    """
    Transparent top-level splash that plays a VP9+alpha WebM video.

    OpenCV cannot extract the VP9 alpha plane on Windows; instead we pipe
    raw yuva420p frames from FFmpeg (bundled via imageio-ffmpeg), convert
    Y/U/V/A planes to BGRA via OpenCV, and paint each frame onto a fully
    transparent frameless Qt window.

    Dismisses on:
      - video reaching its last frame (auto-close)
      - any mouse click (immediate close)
    """

    video_finished = Signal()       # emitted when the last frame plays naturally
    fade_progress = Signal(float)   # 0.0→1.0 during last 10 frames (drives app fade-in)

    # Target display size (visual pixels at 100% DPI; divided by devicePixelRatio at runtime)
    _DISPLAY_W = 1497
    _DISPLAY_H = 840

    def __init__(self, video_path: str, sound_path: str = "", end_sound_path: str = ""):
        super().__init__(None, Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint | Qt.Tool)
        self.setAttribute(Qt.WA_TranslucentBackground, True)

        self._proc = None
        self._valid = False
        self._src_w = 0
        self._src_h = 0
        self._display_w = 1
        self._display_h = 1
        self._frame_bytes = 0
        self._video_duration_ms = 0
        self._total_frames = 0
        self._frame_count = 0
        self._end_sound_path = (end_sound_path or "").strip()
        self._current_pix = QPixmap()
        self._frame_timer = QTimer(self)
        self._frame_timer.timeout.connect(self._next_frame)
        self._frame_interval_ms = 20  # default; overridden from probe

        if video_path and os.path.exists(video_path):
            self._init_ffmpeg(video_path)

        if self._valid:
            self.setFixedSize(self._display_w, self._display_h)
        else:
            self.resize(1, 1)

        # Sound: prefer explicit WAV override, fall back to audio extracted from video
        self._sound = None
        self._temp_audio_path = getattr(self, "_temp_audio_path", None)
        effective_sound = (sound_path or "").strip() or (self._temp_audio_path or "")
        if effective_sound and os.path.exists(effective_sound):
            try:
                s = QSoundEffect(self)
                s.setSource(QUrl.fromLocalFile(effective_sound))
                s.setLoopCount(1)
                s.setVolume(0.9)
                self._sound = s
            except Exception:
                pass

    def _init_ffmpeg(self, video_path: str):
        try:
            import imageio_ffmpeg
            import subprocess
            import tempfile

            ffmpeg_exe = imageio_ffmpeg.get_ffmpeg_exe()

            # Probe dimensions, fps, and whether an audio stream exists
            probe_proc = subprocess.run(
                [ffmpeg_exe, "-i", video_path],
                capture_output=True, text=True,
            )
            W, H, fps, duration_s = self._parse_ffmpeg_probe(probe_proc.stderr)
            if not (W and H):
                return
            self._video_duration_ms = int(duration_s * 1000)
            self._total_frames = max(1, int(duration_s * fps))

            # Extract audio to a temp WAV so QSoundEffect can play it
            if "Audio:" in probe_proc.stderr:
                try:
                    tmp = tempfile.NamedTemporaryFile(suffix=".wav", delete=False)
                    tmp.close()
                    subprocess.run(
                        [ffmpeg_exe, "-y", "-i", video_path, "-vn",
                         "-af", "volume=2.0",
                         "-acodec", "pcm_s16le", "-ar", "48000", "-ac", "2", tmp.name],
                        stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL,
                    )
                    self._temp_audio_path = tmp.name
                except Exception:
                    self._temp_audio_path = None
            else:
                self._temp_audio_path = None

            self._src_w = W
            self._src_h = H
            # Divide by device pixel ratio so the visual size matches _DISPLAY_W/H
            # regardless of Windows display scaling (e.g. 150% DPI)
            screen = QApplication.primaryScreen()
            dpr = screen.devicePixelRatio() if screen else 1.0
            self._display_w = max(1, int(self._DISPLAY_W / dpr))
            self._display_h = max(1, int(self._DISPLAY_H / dpr))
            self._frame_interval_ms = max(1, int(1000.0 / fps))

            # yuva420p layout per frame:
            #   Y plane:  W * H
            #   U plane: (W//2) * (H//2)
            #   V plane: (W//2) * (H//2)
            #   A plane:  W * H
            self._frame_bytes = W * H + (W // 2) * (H // 2) * 2 + W * H

            decode_cmd = [
                ffmpeg_exe,
                "-vcodec", "libvpx-vp9",
                "-i", video_path,
                "-pix_fmt", "yuva420p",
                "-f", "rawvideo",
                "pipe:1",
            ]
            self._proc = subprocess.Popen(
                decode_cmd,
                stdout=subprocess.PIPE,
                stderr=subprocess.DEVNULL,
            )
            self._valid = True
            self._cv2 = __import__("cv2")
            self._np = __import__("numpy")

        except Exception:
            self._valid = False

    @staticmethod
    def _parse_ffmpeg_probe(stderr_text: str):
        """Extract width, height, fps, duration_seconds from `ffmpeg -i` stderr output."""
        import re
        W = H = 0
        fps = 50.0
        duration_s = 0.0
        m = re.search(r"(\d+)x(\d+).*?(\d+(?:\.\d+)?) fps", stderr_text)
        if m:
            W, H = int(m.group(1)), int(m.group(2))
            fps = float(m.group(3))
        d = re.search(r"Duration:\s*(\d+):(\d+):(\d+(?:\.\d+)?)", stderr_text)
        if d:
            duration_s = int(d.group(1)) * 3600 + int(d.group(2)) * 60 + float(d.group(3))
        return W, H, fps, duration_s

    def showEvent(self, event):
        super().showEvent(event)
        if not self._valid:
            return

        screen = QApplication.screenAt(self.pos()) or QApplication.primaryScreen()
        if screen:
            g = screen.availableGeometry()
            self.move(
                g.left() + (g.width() - self.width()) // 2 + 5,
                g.top() + (g.height() - self.height()) // 2 - 13,
            )

        self._next_frame()  # prime first frame immediately (no blank flash)

        try:
            if self._sound:
                self._sound.play()
        except Exception:
            pass

        self._frame_timer.start(self._frame_interval_ms)

        # Fire end sound 500ms before the video finishes (cancellable on early dismiss)
        if self._end_sound_path and os.path.exists(self._end_sound_path) and self._video_duration_ms > 500:
            fire_at_ms = max(0, self._video_duration_ms - 500)
            def _play_end_sound():
                try:
                    s = QSoundEffect(self)
                    s.setSource(QUrl.fromLocalFile(self._end_sound_path))
                    s.setLoopCount(1)
                    s.setVolume(0.63)
                    s.play()
                    self._end_sound_effect = s
                except Exception:
                    pass
            self._end_sound_timer = QTimer(self)
            self._end_sound_timer.setSingleShot(True)
            self._end_sound_timer.timeout.connect(_play_end_sound)
            self._end_sound_timer.start(fire_at_ms)

    def _next_frame(self):
        if not self._valid or self._proc is None:
            return

        raw = self._proc.stdout.read(self._frame_bytes)
        if len(raw) < self._frame_bytes:
            # Stream ended naturally — ensure full crossfade then close
            self._frame_timer.stop()
            self.setWindowOpacity(0.0)
            self.fade_progress.emit(1.0)
            self.video_finished.emit()
            self.close()
            return

        self._frame_count += 1
        frames_remaining = max(0, self._total_frames - self._frame_count)

        if frames_remaining < 10:
            # Fade out: video 1→0, app 0→1
            t = min(1.0, 1.0 - frames_remaining / 10.0)
            self.setWindowOpacity(1.0 - t)
            self.fade_progress.emit(t)
        elif self._frame_count <= 15:
            # Fade in: video 0→1 over first 15 frames
            self.setWindowOpacity(self._frame_count / 15.0)

        np = self._np
        cv2 = self._cv2
        W, H = self._src_w, self._src_h

        # Unpack yuva420p planes
        y_end = W * H
        uv_size = (W // 2) * (H // 2)
        y = np.frombuffer(raw[:y_end], dtype=np.uint8).reshape(H, W)
        u = np.frombuffer(raw[y_end: y_end + uv_size], dtype=np.uint8).reshape(H // 2, W // 2)
        v = np.frombuffer(raw[y_end + uv_size: y_end + 2 * uv_size], dtype=np.uint8).reshape(H // 2, W // 2)
        a = np.frombuffer(raw[y_end + 2 * uv_size:], dtype=np.uint8).reshape(H, W)

        # Upsample chroma planes to full resolution
        u_full = cv2.resize(u, (W, H), interpolation=cv2.INTER_LINEAR)
        v_full = cv2.resize(v, (W, H), interpolation=cv2.INTER_LINEAR)

        # YUV → BGR
        yuv = np.stack([y, u_full, v_full], axis=2)
        bgr = cv2.cvtColor(yuv, cv2.COLOR_YUV2BGR)

        # Merge alpha channel
        bgra = np.concatenate([bgr, a[:, :, np.newaxis]], axis=2)

        # Scale to display size
        bgra = cv2.resize(bgra, (self._display_w, self._display_h), interpolation=cv2.INTER_AREA)

        # BGRA → RGBA (Qt Format_RGBA8888)
        rgba = cv2.cvtColor(bgra, cv2.COLOR_BGRA2RGBA)

        dh, dw = rgba.shape[:2]
        qi = QImage(rgba.tobytes(), dw, dh, dw * 4, QImage.Format_RGBA8888)
        self._current_pix = QPixmap.fromImage(qi)
        self.update()

    def paintEvent(self, event):
        if self._current_pix.isNull():
            return
        painter = QPainter(self)
        painter.setRenderHint(QPainter.SmoothPixmapTransform, True)
        painter.drawPixmap(0, 0, self._current_pix)
        painter.end()

    def mousePressEvent(self, event):
        self._dismiss()
        self.close()
        return super().mousePressEvent(event)

    def closeEvent(self, event):
        self._dismiss()
        super().closeEvent(event)

    def _dismiss(self):
        try:
            self.fade_progress.emit(1.0)
        except Exception:
            pass
        try:
            self._frame_timer.stop()
        except Exception:
            pass
        try:
            if getattr(self, "_end_sound_timer", None):
                self._end_sound_timer.stop()
        except Exception:
            pass
        try:
            if self._sound:
                self._sound.stop()
        except Exception:
            pass
        try:
            if self._proc:
                self._proc.stdout.close()
                self._proc.terminate()
                self._proc = None
        except Exception:
            pass
        try:
            if self._temp_audio_path and os.path.exists(self._temp_audio_path):
                os.unlink(self._temp_audio_path)
                self._temp_audio_path = None
        except Exception:
            pass


class UiBus(QObject):
    """
    Thread-safe UI message bus: worker thread emits signals, UI updates on main thread.
    """
    user_msg = Signal(str)
    dev_msg = Signal(str)
    finished_ok = Signal(dict)
    finished_error = Signal(str)
    state_running = Signal(bool)

    # Busy indicator (dots) control
    busy_start = Signal(str)
    busy_stop = Signal()


@dataclass
class RunArgs:
    excel_path: str
    username: str
    password: str
    headless: bool
    mode: str = "enroll_transfer"


class MainWindow(QMainWindow):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Workabili-Bot3000 🤖")

        # Two-stage UX: compact by default; expands when Dev details are shown
        self._collapsed_h = 620
        self._expanded_h = 980

        self.setMinimumSize(980, self._collapsed_h)

        self._progress_i = 0
        self._progress_n = 0
        self._progress_name = ""
        self._run_start_dt = None

        try:
            self.cfg = load_config()
        except Exception as e:
            self._modal_error("Config Error", f"Could not read {CONFIG_PATH}\n\n{e}")
            raise

        self.bus = UiBus()
        self.bus.user_msg.connect(self._append_status)
        self.bus.dev_msg.connect(self._append_dev)
        self.bus.finished_ok.connect(self._on_finished_ok)
        self.bus.finished_error.connect(self._on_finished_error)
        self.bus.state_running.connect(self._set_running_state)

        # Busy indicator (dots)
        self.bus.busy_start.connect(self._busy_start)
        self.bus.busy_stop.connect(self._busy_stop)

        self._stop_event = threading.Event()
        self._worker_thread = None

        # Per-user UI preferences (safe, tiny, non-invasive)
        self._settings = QSettings("Workabili", "Workabili-Bot3000")

        # Store the real (non-elided) Excel path separately from any UI labels
        self._excel_path = ""
        self._loaded_students_count = 0

        self._build_ui()
        self._seed_defaults()
        self._refresh_last_results_button()

        # Start compact: Dev Console hidden and window height reduced
        if hasattr(self, "dev_card") and self.dev_card:
            self.dev_card.setVisible(False)
        if hasattr(self, "_details_btn") and self._details_btn:
            self._details_btn.setChecked(False)
            self._details_btn.setText("Details")
        self.resize(self.width(), self._collapsed_h)

        self._append_status("Ready. 👇 Pick your Excel file, then choose Enroll / Transfer or Serve.\n")


        # Global click detection (disabled buttons don't receive events)
        self._hint_toast = None
        self._hint_fx = None
        self._hint_anim = None
        app = QApplication.instance()
        if app:
            app.installEventFilter(self)

        # Busy indicator state (spinner during quiet headless phases)
        self._busy_timer = QTimer(self)
        self._busy_timer.setInterval(120)  # smooth spinner without feeling frantic
        self._busy_timer.timeout.connect(self._busy_tick)
        self._busy_active = False
        self._busy_prefix = ""
        self._spinner_frames = ["⠋","⠙","⠹","⠸","⠼","⠴","⠦","⠧","⠇","⠏"]
        self._spinner_i = 0

    # -------------------------
    # UI Layout
    # -------------------------
    def _build_ui(self):
        root = QWidget()
        self.setCentralWidget(root)

        outer = QVBoxLayout(root)
        outer.setContentsMargins(18, 16, 18, 16)
        outer.setSpacing(14)

        # Header
        header_row = QHBoxLayout()
        outer.addLayout(header_row)

        # Title image (transparent PNG)
        title = QLabel()
        title.setStyleSheet("QLabel { background: transparent; }")

        title_pix = QPixmap(os.path.normpath(
            os.path.join(os.path.dirname(__file__), "..", "assets", "workabili-bot3000-title.png")
        ))

        if not title_pix.isNull():
            TARGET_W = 300  # desired title width in pixels

            scaled = title_pix.scaled(
                TARGET_W,
                title_pix.height(),
                Qt.KeepAspectRatio,
                Qt.SmoothTransformation,
            )

            title.setPixmap(scaled)
            title.setFixedSize(scaled.size())
        else:
            # Fallback: if PNG missing, show text title
            title.setText("Workabili-Bot3000")
            for fam in ["Bahnschrift", "Segoe UI Variable Display", "Segoe UI"]:
                f = QFont(fam, 22, QFont.Bold)
                if QFontInfo(f).family() == fam:
                    title.setFont(f)
                    break
            else:
                title.setFont(QFont("Segoe UI", 22, QFont.Bold))

        header_row.addWidget(title)
        version = str(self.cfg.get("app", {}).get("version", "0.1.0") or "0.1.0").strip()
        version_lbl = QLabel(f"v{version}")
        version_lbl.setStyleSheet("QLabel { color: #94a3b8; background: transparent; }")
        version_lbl.setFont(QFont("Segoe UI", 8, QFont.Medium))
        version_lbl.setAlignment(Qt.AlignLeft | Qt.AlignBottom)
        header_row.addWidget(version_lbl)

        badge = QLabel("🤖  Modern WAI student uploader")
        badge.setStyleSheet(
            "QLabel { padding: 6px 10px; border-radius: 10px; background: #1e293b; color: #e2e8f0; }"
        )
        badge.setFont(QFont("Segoe UI", 10, QFont.Medium))
        header_row.addWidget(badge)
        header_row.addStretch(1)

        # Full-width input file bar (prevents chopped-off paths)
        file_bar = QGroupBox()
        file_bar.setStyleSheet(
            "QGroupBox { border:1px solid #24324a; border-radius:16px; background:#0f172a; }"
        )
        file_lay = QHBoxLayout(file_bar)
        file_lay.setContentsMargins(14, 10, 14, 10)
        file_lay.setSpacing(10)

        file_lbl = QLabel("Input file:")
        file_lbl.setFont(QFont("Segoe UI", 11, QFont.Bold))
        file_lay.addWidget(file_lbl)

        self.top_excel_path_lbl = QLabel("No file selected")
        self.top_excel_path_lbl.setStyleSheet("QLabel { color:#e2e8f0; }")
        self.top_excel_path_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        self.top_excel_path_lbl.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Preferred)
        file_lay.addWidget(self.top_excel_path_lbl, stretch=1)

        # Boss-tier: one-click copy full path (even if label is elided)
        self.btn_copy_path = QToolButton()
        self.btn_copy_path.setText("📋 Copy")
        self.btn_copy_path.setEnabled(False)
        self.btn_copy_path.setVisible(False)
        self.btn_copy_path.setStyleSheet(
            "QToolButton { padding:4px 10px; border-radius:10px; background:#1e293b; color:#e2e8f0; }"
            "QToolButton:hover { background:#111b34; }"
            "QToolButton:disabled { color:#64748b; background:#0b1224; }"
        )
        self.btn_copy_path.clicked.connect(
            lambda: (
                QApplication.clipboard().setText(self._excel_path or ""),
                self.bus.user_msg.emit("📋 Path copied to clipboard.")
            )
        )
        file_lay.addWidget(self.btn_copy_path)

        self.btn_open_folder = QToolButton()
        self.btn_open_folder.setText("📂 Folder")
        self.btn_open_folder.setEnabled(False)
        self.btn_open_folder.setVisible(False)
        self.btn_open_folder.setStyleSheet(
            "QToolButton { padding:4px 10px; border-radius:10px; background:#1e293b; color:#e2e8f0; }"
            "QToolButton:hover { background:#111b34; }"
            "QToolButton:disabled { color:#64748b; background:#0b1224; }"
        )
        self.btn_open_folder.clicked.connect(
            lambda: (
                QDesktopServices.openUrl(QUrl.fromLocalFile(os.path.dirname(self._excel_path or ""))),
                self.bus.user_msg.emit("📂 Opened input file folder.")
            )
        )
        file_lay.addWidget(self.btn_open_folder)

        outer.addWidget(file_bar)


        # Main two-column area
        main_row = QHBoxLayout()

        main_row.setSpacing(14)
        outer.addLayout(main_row, stretch=1)

        # Left column: Inputs
        left_col = QVBoxLayout()
        left_col.setSpacing(14)
        main_row.addLayout(left_col, stretch=0)

        # Right column: Status / Console
        right_col = QVBoxLayout()
        right_col.setSpacing(12)
        main_row.addLayout(right_col, stretch=1)

        # --- Card: Inputs
        inputs_card = self._card("Run Setup")
        inputs_card.setMaximumWidth(520)
        left_col.addWidget(inputs_card)

        g = QGridLayout()
        g.setHorizontalSpacing(10)
        g.setVerticalSpacing(10)
        inputs_card.layout().addLayout(g)

        self.excel_path_lbl = QLabel("No file selected")
        self.excel_path_lbl.setStyleSheet(
            "QLabel { border:1px solid #24324a; border-radius:12px; padding:8px 10px; background:#0b1224; color:#e2e8f0; }"
        )
        self.excel_path_lbl.setTextInteractionFlags(Qt.TextSelectableByMouse)
        # Don't let a long filename force the whole left column wider
        self.excel_path_lbl.setSizePolicy(QSizePolicy.Ignored, QSizePolicy.Preferred)
        self.excel_path_lbl.setAlignment(Qt.AlignLeft | Qt.AlignVCenter)

        self.btn_browse = QPushButton("Browse…")
        self.btn_browse.clicked.connect(self._browse_excel)

        input_excel_lbl = QLabel("Input Excel")
        g.addWidget(input_excel_lbl, 0, 0, alignment=Qt.AlignVCenter)
        g.addWidget(self.excel_path_lbl, 0, 1, 1, 2)

        # Browse button sits BELOW the filename label, right-aligned
        g.addWidget(QWidget(), 1, 1)  # spacer
        g.addWidget(self.btn_browse, 1, 2, alignment=Qt.AlignRight)


        self.user_edit = QLineEdit()
        self.user_edit.setPlaceholderText("you@yourorg.org")

        # Persist username preference (QSettings overrides config.yaml on next launch)
        self.user_edit.textChanged.connect(
            lambda t: (
                self._settings.setValue("ui/username", t.strip())
                if t.strip()
                else self._settings.remove("ui/username")
            )
        )

        self.pass_edit = QLineEdit()
        self.pass_edit.textChanged.connect(self._on_password_changed)
        self._require_password_change = False   
        self.pass_edit.setEchoMode(QLineEdit.Password)
        self.pass_edit.setPlaceholderText("Password")

        g.addWidget(QLabel("Username"), 2, 0)
        g.addWidget(self.user_edit, 2, 1, 1, 2)

        g.addWidget(QLabel("Password"), 3, 0)
        g.addWidget(self.pass_edit, 3, 1, 1, 2)

        self.headless_chk = QCheckBox("Hide browser")
        self.headless_chk.setChecked(bool(self.cfg.get("run", {}).get("headless", False)))

        # Persist headless preference (QSettings overrides config.yaml on next launch)
        self.headless_chk.toggled.connect(
            lambda v: self._settings.setValue("ui/headless", bool(v))
        )

        g.addWidget(self.headless_chk, 4, 1, 1, 2)

        # --- Card: Controls
        controls_card = self._card("Controls")
        controls_card.setMaximumWidth(520)
        left_col.addWidget(controls_card)

        primary_row = QHBoxLayout()
        primary_row.setSpacing(10)
        controls_card.layout().addLayout(primary_row)

        secondary_row = QHBoxLayout()
        secondary_row.setSpacing(10)
        controls_card.layout().addLayout(secondary_row)

        BTN_H = 42
        BTN_RADIUS = 12

        BTN_FONT = QFont("Segoe UI Variable Text", 11)
        if QFontInfo(BTN_FONT).family() != "Segoe UI Variable Text":
            BTN_FONT = QFont("Segoe UI", 11)
        BTN_FONT.setWeight(QFont.Medium)

        base_btn_css = f"""
        QPushButton {{
            border-radius: {BTN_RADIUS}px;
            padding: 8px 14px;
            border: 1px solid #334155;
            background: #0b1224;
            color: #e2e8f0;
        }}
        QPushButton:hover {{
            background: #111b34;
            border: 1px solid #475569;
        }}
        QPushButton:pressed {{
            background: #0a1020;
            border: 1px solid #64748b;
        }}
        QPushButton:disabled {{
            background: #0b1224;
            color: #64748b;
            border: 1px solid #24324a;
        }}
        """

        primary_btn_css = f"""
        QPushButton {{
            border-radius: {BTN_RADIUS}px;
            padding: 8px 14px;
            border: 1px solid rgba(148, 163, 184, 90);   /* soft slate outline */
            background: rgba(56, 189, 248, 60);         /* very muted blue */
            color: #ffffff;                             /* white text */
        }}
        QPushButton:hover {{
            background: rgba(56, 189, 248, 90);
            border: 1px solid rgba(148, 163, 184, 130);
        }}
        QPushButton:pressed {{
            background: rgba(56, 189, 248, 120);
            border: 1px solid rgba(148, 163, 184, 160);
        }}
        QPushButton:disabled {{
            background: rgba(15, 23, 42, 140);
            color: rgba(148, 163, 184, 140);
            border: 1px solid rgba(36, 50, 74, 160);
        }}
        """

        danger_btn_css = base_btn_css + """
        QPushButton:hover { border: 1px solid #ef4444; }
        QPushButton:pressed { border: 1px solid #ef4444; }
        """

        # --- Enroll / Transfer (Primary) ---
        self.btn_run = QPushButton("Enroll / Transfer")
        self.btn_run.setFont(BTN_FONT)
        self.btn_run.setStyleSheet(primary_btn_css)
        self.btn_run.setFixedHeight(BTN_H)
        self.btn_run.clicked.connect(self._on_run_clicked)
        self.btn_run.setToolTip("Enroll new students and request or complete transfers.")

        # Use built-in Qt icon (crisp, not pixel glyph)
        self.btn_run.setIcon(self._icon_from_svg(self._svg_with_color(self.LUCIDE_PLAY, "#ffffff"), 18))
        self.btn_run.setIconSize(QSize(18, 18))

        # --- Serve (Primary) ---
        self.btn_serve = QPushButton("Serve")
        self.btn_serve.setFont(BTN_FONT)
        self.btn_serve.setStyleSheet(primary_btn_css)
        self.btn_serve.setFixedHeight(BTN_H)
        self.btn_serve.clicked.connect(self._on_serve_clicked)
        self.btn_serve.setToolTip("Dry-run the Array of Services serve workflow.")
        self.btn_serve.setIcon(self._icon_from_svg(self._svg_with_color(self.LUCIDE_PLAY, "#ffffff"), 18))
        self.btn_serve.setIconSize(QSize(18, 18))

        # --- Stop (Muted Amber) ---
        self.btn_stop = QPushButton("Stop")
        self.btn_stop.setFont(BTN_FONT)
        self.btn_stop.setStyleSheet(f"""
        QPushButton {{
            border-radius: {BTN_RADIUS}px;
            padding: 8px 14px;
            border: 1px solid rgba(148, 163, 184, 90);    /* soft slate outline */
            background: rgba(251, 191, 36, 60);          /* muted amber */
            color: #ffffff;
        }}
        QPushButton:hover {{
            background: rgba(251, 191, 36, 90);
            border: 1px solid rgba(148, 163, 184, 130);
        }}
        QPushButton:pressed {{
            background: rgba(251, 191, 36, 120);
            border: 1px solid rgba(148, 163, 184, 160);
        }}
        QPushButton:disabled {{
            background: rgba(15, 23, 42, 140);
            color: rgba(148, 163, 184, 140);
            border: 1px solid rgba(36, 50, 74, 160);
        }}
        """)
        self.btn_stop.setFixedHeight(BTN_H)
        self.btn_stop.clicked.connect(self._on_stop_clicked)
        self.btn_stop.setEnabled(False)
        self.btn_stop.setToolTip("This button will become clickable once the app is running.")

        self.btn_stop.setIcon(self._icon_from_svg(self._svg_with_color(self.LUCIDE_STOP, "#ffffff"), 18))
        self.btn_stop.setIconSize(QSize(18, 18))

        # --- Quit (Secondary/Danger hover) ---
        self.btn_quit = QPushButton("Quit")
        self.btn_quit.setFont(BTN_FONT)
        self.btn_quit.setStyleSheet(danger_btn_css)
        self.btn_quit.setFixedHeight(BTN_H)
        self.btn_quit.setToolTip("Quit this application entirely.")
        self.btn_quit.clicked.connect(
            lambda: (
                self.close()
                if QMessageBox.question(
                    self,
                    "Quit Workabili-Bot3000",
                    "Are you sure you want to quit?",
                    QMessageBox.Yes | QMessageBox.No,
                    QMessageBox.No,
                )
                == QMessageBox.Yes
                else None
            )
        )

        self.btn_quit.setIcon(self._icon_from_svg(self._svg_with_color(self.LUCIDE_X, "#e2e8f0"), 18))
        self.btn_quit.setIconSize(QSize(18, 18))

        # Equal widths
        self.btn_run.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_serve.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_stop.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)
        self.btn_quit.setSizePolicy(QSizePolicy.Expanding, QSizePolicy.Fixed)

        primary_row.addWidget(self.btn_run, stretch=1)
        primary_row.addWidget(self.btn_serve, stretch=1)
        secondary_row.addWidget(self.btn_stop, stretch=1)
        secondary_row.addWidget(self.btn_quit, stretch=1)

        left_col.addStretch(1)

        # --- Status monitor (big, clean)
        status_card = self._card("Status Monitor")
        right_col.addWidget(status_card, stretch=2)

        self.status_box = QTextEdit()
        self.status_box.setReadOnly(True)
        self.status_box.setLineWrapMode(QTextEdit.WidgetWidth)
        self.status_box.setStyleSheet(
            "QTextEdit { border:1px solid #24324a; border-radius:12px; padding:10px; background:#0b1224; }"
        )
        self.status_box.setFont(QFont("Segoe UI Emoji", 12))
        status_card.layout().addWidget(self.status_box)

        # --- Dev console (hidden by default; shown via subtle Details toggle)
        self.dev_card = self._card("Developer Console", collapsible=False)
        right_col.addWidget(self.dev_card, stretch=2)

        self.dev_box = QPlainTextEdit()
        self.dev_box.setReadOnly(True)
        self.dev_box.setLineWrapMode(QPlainTextEdit.NoWrap)
        self.dev_box.setStyleSheet(
            "QPlainTextEdit { border:1px solid #24324a; border-radius:12px; padding:10px; background:#070d1b; }"
        )
        self.dev_box.setFont(QFont("Cascadia Mono", 10))
        self.dev_card.layout().addWidget(self.dev_box)

        # Footer: subtle Details toggle (right-aligned)
        footer_row = QHBoxLayout()
        footer_row.addStretch(1)

        # Show Last Results (reopen dialog if user closed it; persists across app restarts)
        self.btn_show_last_results = QToolButton()
        self.btn_show_last_results.setText("Show Last Results")
        self.btn_show_last_results.setEnabled(False)
        self.btn_show_last_results.clicked.connect(self._on_show_last_results_clicked)
        self.btn_show_last_results.setStyleSheet(
            "QToolButton { padding:4px 10px; border-radius:10px; background:#0b1224; color:#94a3b8; border:1px solid #24324a; }"
            "QToolButton:hover { background:#111b34; color:#e2e8f0; border:1px solid #334155; }"
            "QToolButton:disabled { color:#64748b; background:#070d1b; border:1px solid #24324a; }"
        )
        footer_row.addWidget(self.btn_show_last_results)

        self._details_btn = QToolButton()
        self._details_btn.setText("Details")
        self._details_btn.setCheckable(True)
        self._details_btn.setChecked(False)
        self._details_btn.clicked.connect(self._toggle_dev_console)
        self._details_btn.setStyleSheet(
            "QToolButton { padding:4px 10px; border-radius:10px; background:#0b1224; color:#94a3b8; border:1px solid #24324a; }"
            "QToolButton:hover { background:#111b34; color:#e2e8f0; border:1px solid #334155; }"
        )

        footer_row.addWidget(self._details_btn)
        outer.addLayout(footer_row)

    def _card(self, title: str, collapsible: bool = False) -> QGroupBox:
        box = QGroupBox()
        box.setStyleSheet(
            "QGroupBox { border:1px solid #24324a; border-radius:16px; background:#0f172a; }"
        )
        lay = QVBoxLayout(box)
        lay.setContentsMargins(14, 12, 14, 14)
        lay.setSpacing(10)

        header = QHBoxLayout()
        header.setSpacing(10)
        lay.addLayout(header)

        lbl = QLabel(title)
        lbl.setFont(QFont("Segoe UI", 12, QFont.Bold))
        header.addWidget(lbl)

        if collapsible:
            header.addStretch(1)
            self._toggle_btn = QToolButton()
            self._toggle_btn.setText("Hide")
            self._toggle_btn.setCheckable(True)
            self._toggle_btn.setChecked(True)
            self._toggle_btn.clicked.connect(self._toggle_dev_console)
            self._toggle_btn.setStyleSheet(
                "QToolButton { padding:4px 10px; border-radius:10px; background:#1e293b; }"
            )
            header.addWidget(self._toggle_btn)
        else:
            header.addStretch(1)

        return box

    def _toggle_dev_console(self):
        show = bool(getattr(self, "_details_btn", None) and self._details_btn.isChecked())

        # Update button label
        if getattr(self, "_details_btn", None):
            self._details_btn.setText("Hide Details" if show else "Details")

        # Show/hide the entire card so it takes zero space when hidden
        if getattr(self, "dev_card", None):
            self.dev_card.setVisible(show)

        # Resize window height between compact/expanded modes
        target_h = self._expanded_h if show else self._collapsed_h
        self.setMinimumSize(self.minimumWidth(), target_h)
        self.resize(self.width(), target_h)

    def _seed_defaults(self):
        self._set_excel_path(self.cfg.get("input", {}).get("default_excel_path", "") or "")

        # Username precedence: QSettings > config.yaml
        saved_user = (self._settings.value("ui/username", "", type=str) or "").strip()
        cfg_user = (self.cfg.get("credentials", {}).get("username", "") or "").strip()
        self.user_edit.setText(saved_user or cfg_user)

        # Headless precedence: QSettings > config.yaml
        saved_headless = self._settings.value("ui/headless", None)
        if saved_headless is None:
            self.headless_chk.setChecked(bool(self.cfg.get("run", {}).get("headless", False)))
        else:
            self.headless_chk.setChecked(bool(saved_headless))

    def _set_excel_path(self, path: str):
        self._excel_path = (path or "").strip()
        self._update_excel_path_labels()

    def _elide_middle(self, label: QLabel, text: str) -> str:
        if not text:
            return "No file selected"
        # Small padding so elide calculation matches visual reality
        w = max(40, label.width() - 12)
        fm = QFontMetrics(label.font())
        return fm.elidedText(text, Qt.ElideMiddle, w)

    def _update_excel_path_labels(self):
        full = self._excel_path or ""

        if full:
            self.excel_path_lbl.setToolTip(full)
            self.excel_path_lbl.unsetCursor()
        else:
            self.excel_path_lbl.setToolTip("Click to select an input Excel file.")
            self.excel_path_lbl.setCursor(Qt.PointingHandCursor)
        self.top_excel_path_lbl.setToolTip(full)

        # Boss-tier copy affordance
        if hasattr(self, "btn_copy_path"):
            has_file = bool(full)
            self.btn_copy_path.setVisible(has_file)
            self.btn_copy_path.setEnabled(has_file)
            self.btn_copy_path.setToolTip(full)

        if hasattr(self, "btn_open_folder"):
            folder = os.path.dirname(full) if full else ""
            has_folder = bool(folder) and os.path.exists(folder)
            self.btn_open_folder.setVisible(has_folder)
            self.btn_open_folder.setEnabled(has_folder)
            self.btn_open_folder.setToolTip(folder)

        # Run Setup: show filename only (no elide). If too long, it will clip on the right.
        self.excel_path_lbl.setText(os.path.basename(full) if full else "No file selected")

        # Top bar: keep elide-middle for full path display
        self.top_excel_path_lbl.setText(self._elide_middle(self.top_excel_path_lbl, full))

    def resizeEvent(self, event):
        super().resizeEvent(event)
        self._update_excel_path_labels()

    # -------------------------
    # Disabled-button hint toast
    # -------------------------
    def eventFilter(self, obj, event):
        if event.type() == QEvent.MouseButtonPress:
            if (
                obj is getattr(self, "excel_path_lbl", None)
                and not (self._excel_path or "").strip()
                and getattr(self, "btn_browse", None)
                and self.btn_browse.isEnabled()
            ):
                self._browse_excel()
                return True

            try:
                gp = event.globalPosition().toPoint()  # Qt6
            except Exception:
                try:
                    gp = event.globalPos()  # fallback
                except Exception:
                    gp = None

            if gp is not None:
                # If user clicks a disabled button area, show a toast near the click.
                # (Disabled widgets won't receive their own mouse events.)
                for btn in [getattr(self, "btn_run", None), getattr(self, "btn_serve", None), getattr(self, "btn_stop", None)]:
                    if not btn:
                        continue
                    if btn.isEnabled():
                        continue

                    # Compute the button rect in GLOBAL coordinates
                    tl = btn.mapToGlobal(QPoint(0, 0))
                    rect = btn.rect()
                    global_rect = rect.translated(tl)

                    if global_rect.contains(gp):
                        # Use the button tooltip if present; otherwise provide a sensible default.
                        msg = (btn.toolTip() or "").strip()
                        if not msg:
                            msg = "Not available right now."

                        self._show_hint_toast(msg, gp)
                        return True  # eat the click so it doesn't do anything weird

        return super().eventFilter(obj, event)

    def _show_hint_toast(self, text: str, global_click_pos: QPoint):
        """
        Small tooltip-like toast near the click:
        fade in → hold ~2s → fade out.
        """
        # Kill existing toast if present
        if getattr(self, "_hint_toast", None):
            try:
                self._hint_toast.close()
                self._hint_toast.deleteLater()
            except Exception:
                pass
            self._hint_toast = None
            # (windowOpacity animation; no opacity graphics effect object)
            self._hint_anim = None

        toast = QWidget(None, Qt.Popup | Qt.FramelessWindowHint | Qt.WindowStaysOnTopHint)
        toast.setAttribute(Qt.WA_TranslucentBackground, True)

        lay = QVBoxLayout(toast)
        lay.setContentsMargins(0, 0, 0, 0)
        lay.setSpacing(0)

        lbl = QLabel(text)
        lbl.setWordWrap(True)
        lbl.setStyleSheet("""
            QLabel {
                background-color: rgba(235, 240, 255, 232);   /* ~91% opaque, softer butter-blue */
                color: #0b1224;
                border: 1px solid rgba(56, 189, 248, 160);
                border-radius: 14px;
                padding: 10px 12px;
                font-size: 13px;
                font-weight: 600;
            }
        """)
        lay.addWidget(lbl)

        toast.adjustSize()

        # Position: slightly above the click, centered horizontally
        w = toast.width()
        h = toast.height()
        x = int(global_click_pos.x() - (w / 2))
        y = int(global_click_pos.y() - h - 14)

        # Clamp to screen so it never renders off-screen
        screen = QApplication.screenAt(global_click_pos) or QApplication.primaryScreen()
        if screen:
            g = screen.availableGeometry()
            x = max(g.left() + 8, min(x, g.right() - w - 8))
            y = max(g.top() + 8, min(y, g.bottom() - h - 8))

        toast.move(x, y)

        # Opacity (for fade in/out)
        toast.setWindowOpacity(0.0)

        # Drop shadow: on Windows, layered+translucent popups can spam
        # "UpdateLayeredWindowIndirect failed ..." during opacity animation.
        # Keep the shadow on non-Windows platforms where it behaves.
        if os.name != "nt":
            shadow = QGraphicsDropShadowEffect(toast)
            shadow.setBlurRadius(32)
            shadow.setOffset(0, 12)
            shadow.setColor(QColor(2, 6, 23, 180))  # deep slate, premium shadow
            toast.setGraphicsEffect(shadow)

        toast.show()

        # Fade in
        anim_in = QPropertyAnimation(toast, b"windowOpacity", toast)
        anim_in.setDuration(140)
        anim_in.setEasingCurve(QEasingCurve.OutCubic)
        anim_in.setStartValue(0.0)
        anim_in.setEndValue(1.0)

        # Fade out (after hold)
        anim_out = QPropertyAnimation(toast, b"windowOpacity", toast)
        anim_out.setDuration(180)
        anim_out.setEasingCurve(QEasingCurve.InCubic)
        anim_out.setStartValue(1.0)
        anim_out.setEndValue(0.0)

        def start_fade_out():
            # Hold ~2 seconds then fade out
            QTimer.singleShot(2000, anim_out.start)

        def cleanup():
            try:
                toast.close()
                toast.deleteLater()
            except Exception:
                pass
            self._hint_toast = None
            # (windowOpacity animation; no opacity graphics effect object)
            self._hint_anim = None

        anim_in.finished.connect(start_fade_out)
        anim_out.finished.connect(cleanup)

        # Keep references so Qt doesn't GC mid-animation
        self._hint_toast = toast
        self._hint_anim = (anim_in, anim_out)

        anim_in.start()

    # -------------------------
    # UX helpers
    # -------------------------
    def _ts(self):
        return datetime.now().strftime("%H:%M:%S")

    def _hm_ampm(self) -> str:
        # Windows-safe: %I is 01-12, we strip leading zero.
        return datetime.now().strftime("%I:%M %p").lstrip("0")

    # -------------------------
    # Busy indicator (spinner)
    # -------------------------
    def _busy_start(self, prefix: str = "⏳ Working") -> None:
        """
        Starts a subtle spinner line while the runner is quiet.
        UI-thread only (called via UiBus signals).
        """
        self._busy_prefix = (prefix or "⏳ Working").strip()
        self._spinner_i = 0
        self._busy_active = True

        # Start on a fresh line and keep spinner on the same line
        self.status_box.append("")
        frame = self._spinner_frames[self._spinner_i % len(self._spinner_frames)]
        self.status_box.insertPlainText(f"{self._busy_prefix} {frame}")

        self._busy_timer.start()
        self.status_box.verticalScrollBar().setValue(self.status_box.verticalScrollBar().maximum())

    def _busy_stop(self) -> None:
        """
        Stops the spinner and finalizes the line cleanly.
        UI-thread only (called via UiBus signals).
        """
        if not getattr(self, "_busy_active", False):
            return

        self._busy_active = False
        try:
            self._busy_timer.stop()
        except Exception:
            pass

        # Remove " <spinner>" (2 chars: space + frame), then newline
        cursor = self.status_box.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.movePosition(QTextCursor.Left, QTextCursor.KeepAnchor, 2)
        cursor.removeSelectedText()
        self.status_box.setTextCursor(cursor)

        self.status_box.insertPlainText("\n")
        self.status_box.verticalScrollBar().setValue(self.status_box.verticalScrollBar().maximum())

    def _busy_tick(self) -> None:
        if not getattr(self, "_busy_active", False):
            return

        self._spinner_i = (self._spinner_i + 1) % len(self._spinner_frames)
        frame = self._spinner_frames[self._spinner_i]

        # Replace the last glyph in-place (avoid adding new lines / spam)
        cursor = self.status_box.textCursor()
        cursor.movePosition(QTextCursor.End)
        cursor.movePosition(QTextCursor.Left, QTextCursor.KeepAnchor, 1)
        cursor.removeSelectedText()
        cursor.insertText(frame)
        self.status_box.setTextCursor(cursor)

        self.status_box.verticalScrollBar().setValue(self.status_box.verticalScrollBar().maximum())

    def _count_students_in_excel(self, path: str) -> int:
        # Best-effort count (excluding header row, ignoring blank rows)
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active
            n = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None and str(v).strip() != "" for v in row):
                    n += 1
            wb.close()
            return n
        except Exception:
            return 0

    @Slot(str)
    def _append_status(self, msg: str):
        # Status Monitor is user-facing: no timestamps, just clean narrative.
        m = msg or ""

        # If caller provides HTML (e.g., to highlight student name), render it.
        if "<span" in m or "<b>" in m or "</b>" in m:
            self.status_box.append(m)
        else:
            # Plain text (QTextEdit treats append() as rich text; escape minimal)
            self.status_box.append(m.replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"))

        self.status_box.verticalScrollBar().setValue(self.status_box.verticalScrollBar().maximum())

    @Slot(str)
    def _append_dev(self, msg: str):
        self.dev_box.appendPlainText(f"{self._ts()}  {msg}")
        self.dev_box.verticalScrollBar().setValue(self.dev_box.verticalScrollBar().maximum())

    def _modal_info(self, title: str, body: str):
        QMessageBox.information(self, title, body)

    def _modal_error(self, title: str, body: str):
        QMessageBox.critical(self, title, body)

    def _modal_bad_creds(self):
        box = QMessageBox(self)
        box.setIcon(QMessageBox.Critical)
        box.setWindowTitle("🚫 Login Failed — Workabili-Bot3000")

        box.setText("🚨 LOGIN FAILED 🚨")
        box.setInformativeText(
            "The username or password is incorrect.\n\n"
            "This usually means:\n"
            "• A typo\n"
            "• Caps Lock is on\n"
            "• Wrong account\n\n"
            "Please re-enter your info and try again."
        )

        box.setStandardButtons(QMessageBox.Ok)
        box.setDefaultButton(QMessageBox.Ok)

        # HARD STOP behavior
        box.setWindowModality(Qt.ApplicationModal)
        box.setWindowFlag(Qt.WindowStaysOnTopHint, True)

        # MAKE IT OBNOXIOUS (on purpose 😄)
        box.setStyleSheet("""
            QLabel {
                font-size: 16px;
                font-weight: 600;
            }
            QLabel:first-child {
                font-size: 20px;
                font-weight: 800;
                color: #b00020;
            }
            QPushButton {
                min-width: 140px;
                min-height: 40px;

                font-family: "Segoe UI Variable Text", "Segoe UI";
                font-size: 15px;
                font-weight: 600;

                border-radius: 12px;
                padding: 8px 16px;

                border: 1px solid rgba(148, 163, 184, 90);
                background: rgba(56, 189, 248, 60);
                color: #ffffff;
            }
            QPushButton:hover {
                background: rgba(56, 189, 248, 90);
                border: 1px solid rgba(148, 163, 184, 130);
            }
            QPushButton:pressed {
                background: rgba(56, 189, 248, 120);
                border: 1px solid rgba(148, 163, 184, 160);
            }
            QPushButton:disabled {
                background: rgba(15, 23, 42, 140);
                color: rgba(148, 163, 184, 140);
                border: 1px solid rgba(36, 50, 74, 160);
            }
        """)

        box.exec()

    def _shake_widget(self, widget, distance=12, duration_ms=200):
        """
        macOS-style horizontal shake animation for a widget.
        """
        anim = QPropertyAnimation(widget, b"pos", self)
        anim.setDuration(duration_ms)

        start_pos = widget.pos()

        anim.setKeyValueAt(0.00, start_pos)
        anim.setKeyValueAt(0.15, start_pos + QPoint(-distance, 0))
        anim.setKeyValueAt(0.30, start_pos + QPoint(distance, 0))
        anim.setKeyValueAt(0.45, start_pos + QPoint(-distance, 0))
        anim.setKeyValueAt(0.60, start_pos + QPoint(distance, 0))
        anim.setKeyValueAt(0.75, start_pos + QPoint(-distance // 2, 0))
        anim.setKeyValueAt(1.00, start_pos)

        anim.setEasingCurve(QEasingCurve.OutCubic)

        # Keep a reference so Qt doesn’t garbage-collect mid-animation
        self._shake_anim = anim
        anim.start()

    def _validate_inputs(self, mode: str = "enroll_transfer") -> RunArgs | None:
        excel_path = (self._excel_path or "").strip()
        username = self.user_edit.text().strip()
        password = self.pass_edit.text()

        if not excel_path:
            self._modal_error("Missing Excel File", "Please select an input Excel (.xlsx) file.")
            return None
        if not os.path.exists(excel_path):
            self._modal_error("File Not Found", f"Excel file does not exist:\n{excel_path}")
            return None
        if not username:
            self._modal_error("Missing Username", "Please enter a username.")
            return None
        if not password:
            self._modal_error("Missing Password", "Please enter a password.")

            # UX love: guide attention + gentle mac-style shake
            self.pass_edit.setFocus()
            self._shake_widget(self.pass_edit)

            return None

        return RunArgs(
            excel_path=excel_path,
            username=username,
            password=password,
            headless=self.headless_chk.isChecked(),
            mode=mode,
        )

    # -------------------------
    # Actions
    # -------------------------
    def _browse_excel(self):
        path, _ = QFileDialog.getOpenFileName(
            self,
            "Select input Excel file",
            "",
            "Excel files (*.xlsx);;All files (*.*)",
        )
        if not path:
            return

        # Scan workbook and count student rows (excluding header, ignoring blank rows)
        try:
            wb = load_workbook(path, read_only=True, data_only=True)
            ws = wb.active

            n = 0
            for row in ws.iter_rows(min_row=2, values_only=True):
                if any(v is not None and str(v).strip() != "" for v in row):
                    n += 1

            wb.close()
        except Exception as e:
            self._modal_error("Excel Read Error", f"Could not read this Excel file:\n\n{path}\n\n{e}")
            return

        filename = os.path.basename(path)

        dlg = QDialog(self)
        dlg.setWindowTitle("Confirm Input File")
        dlg.setWindowModality(Qt.ApplicationModal)
        dlg.setWindowFlag(Qt.WindowStaysOnTopHint, True)

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(12)

        noun = "student" if n == 1 else "students"
        t1 = QLabel(f"Loaded {n} {noun} to process.")
        t1.setStyleSheet("QLabel { font-size: 16px; font-weight: 700; color: #e2e8f0; }")
        lay.addWidget(t1)

        t2 = QLabel(f"File: {filename}\n\nClick OK to proceed, or Cancel to choose a different file.")
        t2.setStyleSheet("QLabel { font-size: 13px; color: #cbd5e1; }")
        t2.setWordWrap(True)
        lay.addWidget(t2)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)

        btn_cancel = QPushButton("Cancel")
        btn_ok = QPushButton("OK")

        # OK = primary (muted glass blue). Cancel = darker, clearly secondary.
        ok_btn_css = """
        QPushButton {
            min-width: 96px;
            min-height: 22px;
            font-family: "Segoe UI Variable Text", "Segoe UI";
            font-size: 13px;
            font-weight: 600;
            border-radius: 12px;
            padding: 4px 12px;
            border: 1px solid rgba(148, 163, 184, 90);
            background: rgba(56, 189, 248, 60);
            color: #ffffff;
        }
        QPushButton:hover {
            background: rgba(56, 189, 248, 90);
            border: 1px solid rgba(148, 163, 184, 130);
        }
        QPushButton:pressed {
            background: rgba(56, 189, 248, 120);
            border: 1px solid rgba(148, 163, 184, 160);
        }
        """

        cancel_btn_css = """
        QPushButton {
            min-width: 96px;
            min-height: 22px;
            font-family: "Segoe UI Variable Text", "Segoe UI";
            font-size: 13px;
            font-weight: 600;
            border-radius: 12px;
            padding: 4px 12px;
            border: 1px solid rgba(148, 163, 184, 90);
            background: rgba(11, 18, 36, 220);   /* darker, low-attention */
            color: rgba(226, 232, 240, 220);
        }
        QPushButton:hover {
            background: rgba(17, 27, 52, 220);
            border: 1px solid rgba(148, 163, 184, 130);
        }
        QPushButton:pressed {
            background: rgba(10, 16, 32, 230);
            border: 1px solid rgba(148, 163, 184, 160);
        }
        """

        btn_cancel.setStyleSheet(cancel_btn_css)
        btn_ok.setStyleSheet(ok_btn_css)

        # FORCE Mac convention order: Cancel (left) then OK (right)
        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        btn_cancel.clicked.connect(dlg.reject)
        btn_ok.clicked.connect(dlg.accept)
        btn_ok.setDefault(True)

        if dlg.exec() != QDialog.Accepted:
            return

        self._loaded_students_count = n
        self._set_excel_path(path)
    def _show_run_preflight(self) -> bool:
        """
        Returns True if user confirms proceeding with the run.
        Respects 'Do not show again' via QSettings.
        """
        show = self._settings.value("ui/show_run_preflight", True, type=bool)
        if not show:
            return True

        n = int(getattr(self, "_loaded_students_count", 0))
        if n <= 0:
            return True

        dlg = QDialog(self)
        dlg.setWindowTitle("Confirm Run")
        dlg.setWindowModality(Qt.ApplicationModal)
        dlg.setWindowFlag(Qt.WindowStaysOnTopHint, True)

        lay = QVBoxLayout(dlg)
        lay.setContentsMargins(16, 14, 16, 14)
        lay.setSpacing(12)

        noun = "student" if n == 1 else "students"
        hdr = QLabel(f"About to process {n} {noun}.")
        hdr.setStyleSheet("QLabel { font-size:16px; font-weight:700; color:#e2e8f0; }")
        lay.addWidget(hdr)

        body = QLabel(
            "Click OK to proceed.\n\n"
            "If you need to stop mid-process for any reason, click the Stop button (to the right of the Run button). "
            "The run will finish the current step and halt safely."
        )
        body.setWordWrap(True)
        body.setStyleSheet("QLabel { font-size:13px; color:#cbd5e1; }")
        lay.addWidget(body)

        chk = QCheckBox("Do not show this next time")
        chk.setStyleSheet("QCheckBox { color:#e2e8f0; }")
        lay.addWidget(chk)

        btn_row = QHBoxLayout()
        btn_row.addStretch(1)

        btn_cancel = QPushButton("Cancel")
        btn_ok = QPushButton("OK")

        ok_btn_css = """
        QPushButton {
            min-width: 96px;
            min-height: 22px;
            font-family: "Segoe UI Variable Text", "Segoe UI";
            font-size: 13px;
            font-weight: 600;
            border-radius: 12px;
            padding: 4px 12px;
            border: 1px solid rgba(148, 163, 184, 90);
            background: rgba(56, 189, 248, 60);
            color: #ffffff;
        }
        QPushButton:hover {
            background: rgba(56, 189, 248, 90);
            border: 1px solid rgba(148, 163, 184, 130);
        }
        QPushButton:pressed {
            background: rgba(56, 189, 248, 120);
            border: 1px solid rgba(148, 163, 184, 160);
        }
        """

        cancel_btn_css = """
        QPushButton {
            min-width: 96px;
            min-height: 22px;
            font-family: "Segoe UI Variable Text", "Segoe UI";
            font-size: 13px;
            font-weight: 600;
            border-radius: 12px;
            padding: 4px 12px;
            border: 1px solid rgba(148, 163, 184, 90);
            background: rgba(11, 18, 36, 220);
            color: rgba(226, 232, 240, 220);
        }
        QPushButton:hover {
            background: rgba(17, 27, 52, 220);
            border: 1px solid rgba(148, 163, 184, 130);
        }
        QPushButton:pressed {
            background: rgba(10, 16, 32, 230);
            border: 1px solid rgba(148, 163, 184, 160);
        }
        """

        btn_ok.setStyleSheet(ok_btn_css)
        btn_cancel.setStyleSheet(cancel_btn_css)

        btn_row.addWidget(btn_cancel)
        btn_row.addWidget(btn_ok)
        lay.addLayout(btn_row)

        btn_cancel.clicked.connect(dlg.reject)
        btn_ok.clicked.connect(dlg.accept)
        btn_ok.setDefault(True)

        result = dlg.exec()

        if chk.isChecked():
            self._settings.setValue("ui/show_run_preflight", False)

        return result == QDialog.Accepted

    def _on_run_clicked(self):
        args = self._validate_inputs(mode="enroll_transfer")
        if not args:
            return

        if not self._show_run_preflight():
            return

        self._start_workflow(args, "Enroll / Transfer")

    def _on_serve_clicked(self):
        args = self._validate_inputs(mode="serve")
        if not args:
            return

        self._start_workflow(args, "Serve")

    def _start_workflow(self, args: RunArgs, workflow_label: str):
        self._stop_event.clear()
        self.bus.state_running.emit(True)

        # User-facing run header
        self._run_start_dt = datetime.now()
        self.bus.user_msg.emit(f"▶️ {workflow_label} started at {self._hm_ampm()}")

        # Ensure we have a student count (Browse sets this; fallback counts silently)
        if getattr(self, "_loaded_students_count", 0) <= 0 and args.excel_path:
            self._loaded_students_count = self._count_students_in_excel(args.excel_path)

        noun = "student" if int(getattr(self, '_loaded_students_count', 0)) == 1 else "students"
        self.bus.user_msg.emit(f"📄 Loaded {int(getattr(self, '_loaded_students_count', 0))} {noun} from {os.path.basename(args.excel_path)}")
        self.bus.user_msg.emit("────────────────────────")

        # Headless login can feel like a freeze — show a subtle heartbeat
        self._busy_start("⏳ Logging in")

        self.bus.dev_msg.emit(f"Mode: {args.mode}")
        self.bus.dev_msg.emit(f"Excel: {args.excel_path}")
        self.bus.dev_msg.emit(f"Headless: {args.headless}")

        # Run runner in background thread (keeps UI buttery)
        self._worker_thread = threading.Thread(target=self._run_worker, args=(args,), daemon=True)
        self._worker_thread.start()

    def _on_stop_clicked(self):
        self.bus.user_msg.emit("🛑 Stop requested — finishing current step and halting safely")
        self._stop_event.set()

    def _run_worker(self, args: RunArgs):
        """
        Worker thread. Emits to UI via bus signals.
        """
        def dev_log(msg: str):
            self.bus.dev_msg.emit(msg)

        def user_log(msg: str):
            # Runner may emit either:
            #   - plain human text
            #   - a state code (e.g., "CREATED")
            #   - a templated state payload (e.g., "CREATED||display_name=John Doe")
            mapping = self.cfg.get("user_messages", {}) or {}

            m = msg.strip()

            # Suppress redundant "Starting ..." line from runner (GUI already prints run header).
            if m.startswith("🚀 Starting") and "Workabili-Bot3000" in m:
                return

            # Suppress runner's end-of-run line (GUI will render the final summary cleanly).
            if m == "✅ Run complete. Check the ledger for details.":
                return

            # Logged in message arrives after a quiet headless phase — stop the login heartbeat,
            # then start a short “opening records” heartbeat until first student processing appears.
            if m.startswith("📚 Logged in. Opening Student Records"):
                self.bus.busy_stop.emit()
                self.bus.user_msg.emit(m)
                self.bus.busy_start.emit("⏳ Opening Student Records")
                return

            # Creation can take 15–20s headless — show the crawling spinner until outcome arrives.
            if m.startswith("📝 Not found in WAI. Entering now"):
                self.bus.user_msg.emit(m)
                self.bus.busy_start.emit("🐍 Creating student")
                return

            # Convert progress line into immediate, user-friendly "Processing" line.
            # Example: "🔎 1 of 2: Checking Alexander Cano…"
            if m.startswith("🔎 ") and ": Checking " in m:
                # First progress message = we're alive and moving; stop heartbeat.
                self.bus.busy_stop.emit()
                try:
                    prefix, rest = m.split(": Checking ", 1)
                    nums = prefix.replace("🔎", "").strip()  # "1 of 2"
                    i_str, n_str = [s.strip() for s in nums.split("of", 1)]
                    name = rest.strip().rstrip("…").strip()

                    # Store in case useful later (optional)
                    self._progress_i = int(i_str)
                    self._progress_n = int(n_str)
                    self._progress_name = name

                    self.bus.user_msg.emit("")
                    name_html = f'<span style="color:#facc15; font-weight:700;">{name}</span>'
                    self.bus.user_msg.emit(f"👤 {self._progress_i} of {self._progress_n}: Processing {name_html}…")
                except Exception:
                    # If parsing fails, show the original line rather than drop it.
                    self.bus.user_msg.emit(m)
                return

            state = m
            ctx = {}

            if "||" in m:
                parts = m.split("||")
                state = parts[0].strip()
                for p in parts[1:]:
                    if "=" in p:
                        k, v = p.split("=", 1)
                        ctx[k.strip()] = v.strip()

            # Outcome line: show ONLY the mapped outcome (Processing line already emitted earlier).
            if state in mapping:
                if state in {
                    "CREATED",
                    "ALREADY_OWNED",
                    "TRANSFER_REQUESTED",
                    "TRANSFER_PENDING",
                    "TRANSFERRED_PRIOR_YEAR",
                    "SERVE_DRY_RUN_FILLED",
                    "SERVE_SAVED",
                    "SERVE_SKIPPED_NOT_OWNED",
                    "SERVE_NOT_FOUND",
                    "SKIPPED_RESUME",
                    "SKIPPED_MISSING_SSID",
                    "NEEDS_INPUT_DATA",
                    "ERROR",
                    "SERVE_ERROR",
                }:
                    self.bus.busy_stop.emit()
                self.bus.user_msg.emit(self._user_message_for_state(state, **ctx))
            else:
                self.bus.user_msg.emit(msg)


        try:
            # NOTE: Stop button requires a tiny runner.py addition (cooperative cancel).
            # For now we pass stop_event; your runner will need to check it between steps.
            runner_fn = run_serve_runner if args.mode == "serve" else run_runner
            runner_kwargs = {}
            if args.mode == "serve":
                runner_kwargs["save_after_fill"] = True

            summary = runner_fn(
                excel_path=args.excel_path,
                config_path=CONFIG_PATH,
                username=args.username,
                password=args.password,
                log_fn=dev_log,                 # dev console
                override_headless=args.headless,
                stop_event=self._stop_event,    # <-- requires runner to accept + check
                user_log_fn=user_log,           # <-- requires runner to emit user-friendly messages
                **runner_kwargs,
            )
            self.bus.finished_ok.emit(summary or {})

        except Exception as e:
            self.bus.finished_error.emit(str(e))

        finally:
            self.bus.state_running.emit(False)

    @Slot(bool)
    def _set_running_state(self, running: bool):
        self.btn_run.setEnabled(not running)
        if hasattr(self, "btn_serve"):
            self.btn_serve.setEnabled(not running)
        self.btn_browse.setEnabled(not running)
        self.excel_path_lbl.setEnabled(not running)
        self.top_excel_path_lbl.setEnabled(not running)
        if hasattr(self, "btn_copy_path"):
            self.btn_copy_path.setEnabled((not running) and bool(self._excel_path or ""))

        if hasattr(self, "btn_open_folder"):
            folder = os.path.dirname(self._excel_path or "")
            self.btn_open_folder.setEnabled((not running) and bool(folder) and os.path.exists(folder))

        self.user_edit.setEnabled(not running)

        self.pass_edit.setEnabled(not running)
        self.headless_chk.setEnabled(not running)
        self.btn_stop.setEnabled(running)
        self.btn_stop.setToolTip(
            "Click to gracefully stop the current run."
            if running
            else "This button will become clickable once the app is running."
        )


    @Slot(dict)
    def _on_finished_ok(self, summary: dict):
        self._busy_stop()
        s = summary or {}
        n_errors = int(s.get("errors", 0))
        n_stopped = int(s.get("stopped", 0))

        self.bus.user_msg.emit("")
        self.bus.user_msg.emit("────────────────────────")

        if n_stopped > 0:
            self.bus.user_msg.emit("🛑 Run stopped by user")
        elif n_errors > 0:
            self.bus.user_msg.emit("⚠️ Run finished with issues — see details")
        else:
            mins = None
        try:
            if getattr(self, "_run_start_dt", None):
                secs = (datetime.now() - self._run_start_dt).total_seconds()
                mins = int((secs + 59) // 60)  # round up to whole minutes
        except Exception:
            mins = None

        if mins is not None and mins > 0:
            unit = "min" if mins == 1 else "mins"
            self.bus.user_msg.emit(f"✅ Run completed in {mins} {unit}\n\n")
        else:
            self.bus.user_msg.emit("✅ Run completed\n\n")


        self._persist_last_results(s)
        self._modal_success(s)


    @Slot(str)
    def _on_finished_error(self, err: str):
        self._busy_stop()
        self.bus.user_msg.emit("❌ Run failed.")

        # SUPER OBVIOUS BAD-CREDENTIALS FLOW
        if "LOGIN_BAD_CREDS" in err:
            self._modal_bad_creds()

            # Boss move sequence
            self.pass_edit.clear()
            self.pass_edit.setFocus()
            self._shake_widget(self.pass_edit)

            # Disable Run until a new password is typed
            self._require_password_change = True
            self.btn_run.setEnabled(False)
            self.btn_run.setToolTip("Please re-enter your password and then this button will become clickable.")
            if hasattr(self, "btn_serve"):
                self.btn_serve.setEnabled(False)
                self.btn_serve.setToolTip("Please re-enter your password and then this button will become clickable.")

            return

        if "SERVE_LAYOUT_CHANGED" in err:
            self._modal_error(
                "WAI Page Layout Changed",
                "The WAI Array of Services page no longer matches the expected row or column labels.\n\n"
                "The Serve workflow has been halted so it does not write to the wrong boxes.\n\n"
                f"Details:\n{err}",
            )
            return

        self._modal_error("Workabili-Bot3000", f"Run failed:\n\n{err}")

    def _on_password_changed(self):
        if getattr(self, "_require_password_change", False):
            if self.pass_edit.text().strip():
                self._require_password_change = False

                # Re-enable Run only if we're not running
                if not self.btn_stop.isEnabled():
                    self.btn_run.setEnabled(True)
                    self.btn_run.setToolTip("Enroll new students and request or complete transfers.")
                    if hasattr(self, "btn_serve"):
                        self.btn_serve.setEnabled(True)
                        self.btn_serve.setToolTip("Dry-run the Array of Services serve workflow.")

    def _svg_with_color(self, svg: str, color_hex: str) -> str:
        # Lucide uses stroke="currentColor"; we replace with an explicit stroke color
        return svg.replace('stroke="currentColor"', f'stroke="{color_hex}"')

    def _icon_from_svg(self, svg_text: str, size: int = 20) -> QIcon:
        renderer = QSvgRenderer(QByteArray(svg_text.encode("utf-8")))
        pixmap = QPixmap(size, size)
        pixmap.fill(Qt.transparent)
        painter = QPainter(pixmap)
        renderer.render(painter)
        painter.end()
        return QIcon(pixmap)

    LUCIDE_PLAY = """
    <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24"
    viewBox="0 0 24 24" fill="none" stroke="currentColor"
    stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
    <polygon points="6 3 20 12 6 21 6 3"></polygon>
    </svg>
    """

    LUCIDE_STOP = """
    <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24"
    viewBox="0 0 24 24" fill="none" stroke="currentColor"
    stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
    <rect x="6" y="6" width="12" height="12" rx="2"></rect>
    </svg>
    """

    LUCIDE_X = """
    <svg xmlns="http://www.w3.org/2000/svg" width="24" height="24"
    viewBox="0 0 24 24" fill="none" stroke="currentColor"
    stroke-width="2" stroke-linecap="round" stroke-linejoin="round">
    <line x1="18" y1="6" x2="6" y2="18"></line>
    <line x1="6" y1="6" x2="18" y2="18"></line>
    </svg>
    """
    def _modal_success(self, summary: dict):
        # Safe reads
        n_total = int(summary.get("input_rows", 0))
        n_created = int(summary.get("created", 0))
        n_tr_req = int(summary.get("transfer_requested", 0))
        n_tr_pend = int(summary.get("transfer_pending", 0))
        n_tr_prior = int(summary.get("transferred_prior_year", 0))
        n_owned = int(summary.get("already_owned", 0))
        n_skipped = int(summary.get("skipped", 0))
        n_errors = int(summary.get("errors", 0))
        n_stopped = int(summary.get("stopped", 0))
        n_serve_filled = int(summary.get("serve_dry_run_filled", 0))
        n_serve_saved = int(summary.get("serve_saved", 0))
        n_serve_not_owned = int(summary.get("serve_skipped_not_owned", 0))
        n_serve_not_found = int(summary.get("serve_not_found", 0))

        # Headline + vibe
        headline = "✅ SUCCESS!"
        vibe = "That was a FULL SEND. You just saved hours of manual work. 💪🤖"

        # If stopped or errors, adjust tone slightly
        if n_stopped > 0:
            headline = "🛑 RUN STOPPED"
            vibe = "Run stopped by user. No worries — everything logged cleanly."
        elif n_errors > 0:
            headline = "⚠️ RUN FINISHED WITH ISSUES"
            vibe = "Most work completed, but some records need attention (see Developer Console / ledger)."

        box = QMessageBox(self)
        box.setWindowTitle("Workabili-Bot3000")
        box.setIcon(QMessageBox.Information)
        box.setStandardButtons(QMessageBox.Ok)
        box.setDefaultButton(QMessageBox.Ok)
        box.setWindowModality(Qt.ApplicationModal)
        box.setWindowFlag(Qt.WindowStaysOnTopHint, True)

        # Big, clean, pro
        box.setTextFormat(Qt.RichText)
        box.setText(f"""
            <div style="font-size:22px; font-weight:900; margin-bottom:8px;">{headline}</div>
            <div style="font-size:14px; color:#cbd5e1; margin-bottom:14px;">{vibe}</div>
        """)

        # Summary table vibes
        box.setInformativeText(f"""
            <div style="font-size:14px; line-height:1.7;">
            <b>Students in input file:</b> {n_total}<br>
            <b>✅ Created:</b> {n_created}<br>
            <b>📨 Transfer requested:</b> {n_tr_req}<br>
            <b>⏳ Transfer pending:</b> {n_tr_pend}<br>
            <b>📨 Prior-year transferred:</b> {n_tr_prior}<br>
            <b>🏫 Already with us:</b> {n_owned}<br>
            <b>⏭️ Already processed:</b> {n_skipped}<br>
            <b>❌ Errors:</b> {n_errors}
            </div>
        """)

        if n_serve_filled or n_serve_saved or n_serve_not_owned or n_serve_not_found:
            box.setInformativeText(f"""
                <div style="font-size:14px; line-height:1.7;">
                <b>Students in input file:</b> {n_total}<br>
                <b>✅ Serve saved:</b> {n_serve_saved}<br>
                <b>⏭️ Not with us:</b> {n_serve_not_owned}<br>
                <b>⚠️ Not found:</b> {n_serve_not_found}<br>
                <b>⏭️ Invalid/skipped input:</b> {n_skipped}<br>
                <b>❌ Errors:</b> {n_errors}
                </div>
            """)

        # Make it feel premium
        box.setStyleSheet("""
            QMessageBox {
                background: #0b1224;
            }
            QLabel {
                color: #e2e8f0;
            }
            QPushButton {
                min-width: 140px;
                min-height: 40px;

                font-family: "Segoe UI Variable Text", "Segoe UI";
                font-size: 15px;
                font-weight: 600;

                border-radius: 12px;
                padding: 8px 16px;

                border: 1px solid rgba(148, 163, 184, 90);
                background: rgba(56, 189, 248, 60);
                color: #ffffff;
            }

            QPushButton:hover {
                background: rgba(56, 189, 248, 90);
                border: 1px solid rgba(148, 163, 184, 130);
            }
            QPushButton:pressed {
                background: rgba(56, 189, 248, 120);
                border: 1px solid rgba(148, 163, 184, 160);
            }
            QPushButton:disabled {
                background: rgba(15, 23, 42, 140);
                color: rgba(148, 163, 184, 140);
                border: 1px solid rgba(36, 50, 74, 160);
            }
        """)

        box.exec()

        # After user acknowledges success, show a clean grouped results view (end-user butter)
        try:
            self._show_run_results_dialog(summary or {})
        except Exception as e:
            # Never block the user if results UI fails for any reason
            self.bus.dev_msg.emit(f"Run Results UI failed: {e}")

    # -------------------------
    # Run Results (end-user view)
    # -------------------------
    def _read_ledger_rows(self, ledger_path: str) -> list[dict]:
        """
        Reads the ledger XLSX into a list of dict rows.
        Expected headers include: run_id, timestamp, ssid, student_name, action, details, screenshot, trace
        """
        out: list[dict] = []
        if not ledger_path or not os.path.exists(ledger_path):
            return out

        try:
            wb = load_workbook(ledger_path, read_only=True, data_only=True)
            ws = wb.active

            rows = list(ws.iter_rows(values_only=True))
            if not rows:
                wb.close()
                return out

            headers = [str(h).strip() if h is not None else "" for h in rows[0]]
            idx = {h: i for i, h in enumerate(headers) if h}

            def cell(r, name: str) -> str:
                i = idx.get(name)
                if i is None:
                    return ""
                v = r[i] if i < len(r) else ""
                return str(v).strip() if v is not None else ""

            for r in rows[1:]:
                ssid = cell(r, "ssid")
                action = cell(r, "action")
                if not ssid and not action:
                    continue
                out.append({
                    "run_id": cell(r, "run_id"),
                    "timestamp": cell(r, "timestamp"),
                    "ssid": ssid,
                    "student_name": cell(r, "student_name"),
                    "action": action,
                    "details": cell(r, "details"),
                    "screenshot": cell(r, "screenshot"),
                    "trace": cell(r, "trace"),
                })

            wb.close()
            return out
        except Exception:
            return out

    def _parse_iso_ts(self, s: str) -> datetime | None:
        try:
            if not s:
                return None
            # ledger uses datetime.utcnow().isoformat()
            return datetime.fromisoformat(s)
        except Exception:
            return None

    def _latest_by_ssid(self, rows: list[dict]) -> list[dict]:
        """
        Dedupes by SSID keeping the most recent row by timestamp (ISO).
        """
        best: dict[str, dict] = {}
        for r in rows:
            ssid = (r.get("ssid") or "").strip()
            if not ssid:
                continue
            ts = self._parse_iso_ts((r.get("timestamp") or "").strip())
            if ssid not in best:
                best[ssid] = r
                continue
            ts_best = self._parse_iso_ts((best[ssid].get("timestamp") or "").strip())
            if ts_best is None and ts is not None:
                best[ssid] = r
            elif ts is not None and ts_best is not None and ts > ts_best:
                best[ssid] = r
            elif ts_best is None and ts is None:
                # fallback: keep the later one by string compare
                if (r.get("timestamp") or "") > (best[ssid].get("timestamp") or ""):
                    best[ssid] = r
        # Sort descending by timestamp for display (latest first)
        def key(r):
            ts = self._parse_iso_ts((r.get("timestamp") or "").strip())
            return ts or datetime.min
        return sorted(best.values(), key=key, reverse=True)

    def _show_run_results_dialog(self, summary: dict, force: bool = False) -> None:
        """
        Opens the Run Results dialog after the success modal closes.
        Uses ledger as source of truth.
        """
        try:
            run_id = str(summary.get("run_id", "") or "").strip()
            ledger_path = str(summary.get("ledger_path", "") or "").strip()
            if not run_id or not ledger_path:
                return

            # Guard: prevent accidental double-show (e.g., merged code calling it twice)
            if (not force) and getattr(self, "_results_dialog_last_run_id", None) == run_id:
                return
            self._results_dialog_last_run_id = run_id

            all_rows = self._read_ledger_rows(ledger_path)

            # This run rows
            run_rows = [r for r in all_rows if (r.get("run_id") or "").strip() == run_id]

            # Transfers: this run
            transfers_run = [r for r in run_rows if (r.get("action") in ("TRANSFER_REQUESTED", "TRANSFER_PENDING", "TRANSFERRED_PRIOR_YEAR"))]

            # Transfers: all open. Dedup across all ledger rows first so later
            # resolved states remove older transfer requests from the open list.
            latest_all = self._latest_by_ssid(all_rows)
            transfers_open = [
                r for r in latest_all
                if (r.get("action") in ("TRANSFER_REQUESTED", "TRANSFER_PENDING"))
            ]

            dlg = RunResultsDialog(
                parent=self,
                summary=summary,
                run_rows=run_rows,
                all_rows=all_rows,
                transfers_open_rows=transfers_open,
            )

            # Default tab selection:
            # - If transfers exist this run: Transfers + scope "This Run"
            # - Else if any open transfers exist: Transfers + scope "All Open Transfers"
            # - Else prefer Created if any, otherwise All
            if len(transfers_run) > 0:
                dlg.set_transfers_scope("This Run")
                dlg.select_tab("Transfers")
            elif len(transfers_open) > 0:
                dlg.select_tab("Transfers")
            else:
                created_run = [r for r in run_rows if (r.get("action") == "CREATED")]
                if len(created_run) > 0:
                    dlg.select_tab("Created")
                else:
                    dlg.select_tab("All")

            dlg.exec()
        except Exception:
            return

    def _persist_last_results(self, summary: dict) -> None:
        """
        Persist last-run summary so users can reopen Results even after closing the dialog
        (and even after app restart).
        """
        try:
            if not isinstance(summary, dict):
                return
            s = dict(summary)
            if not s.get("run_id") or not s.get("ledger_path"):
                return
            self._settings.setValue("ui/last_results_yaml", yaml.safe_dump(s))
            self._refresh_last_results_button()
        except Exception:
            pass

    def _load_last_results(self) -> dict | None:
        try:
            raw = self._settings.value("ui/last_results_yaml", "", type=str) or ""
            raw = raw.strip()
            if not raw:
                return None
            data = yaml.safe_load(raw)
            if isinstance(data, dict) and data.get("run_id") and data.get("ledger_path"):
                return data
            return None
        except Exception:
            return None

    def _refresh_last_results_button(self) -> None:
        try:
            has_last = bool(self._load_last_results())
            if hasattr(self, "btn_show_last_results"):
                self.btn_show_last_results.setEnabled(has_last)
                self.btn_show_last_results.setToolTip(
                    "Show the results from the most recent run."
                    if has_last
                    else "No previous run results found yet."
                )
        except Exception:
            pass

    def _on_show_last_results_clicked(self):
        s = self._load_last_results()
        if not s:
            self._modal_info("No Results", "No previous run results were found.")
            return
        self._show_run_results_dialog(s, force=True)

    def _user_message_for_state(self, state: str, **ctx) -> str:
        """
        Translate internal runner state into a clear, user-facing message.
        Supports lightweight templating via config.yaml (e.g., {display_name}).
        Falls back gracefully if unmapped.
        """
        mapping = self.cfg.get("user_messages", {}) or {}

        msg = mapping.get(state)
        if msg:
            # Safe formatting: if template references a missing key, keep it as-is.
            class SafeDict(dict):
                def __missing__(self, key):
                    return "{" + key + "}"
            return msg.format_map(SafeDict(ctx))

        # Fallback (never show raw state alone)
        return f"Processed ({state.replace('_', ' ').title()})."


class RunResultsDialog(QDialog):
    """
    End-user friendly run summary + grouped lists.
    Uses the ledger as truth.
    """
    ACTION_LABELS = {
        "CREATED": "✅ Created",
        "ALREADY_OWNED": "🏫 Already with us",
        "TRANSFER_REQUESTED": "📨 Transfer requested",
        "TRANSFER_PENDING": "⏳ Transfer pending",
        "TRANSFERRED_PRIOR_YEAR": "📨 Prior-year transferred",
        "SERVE_DRY_RUN_FILLED": "✅ Serve dry-run filled",
        "SERVE_SAVED": "✅ Serve saved",
        "SERVE_SKIPPED_NOT_OWNED": "⏭️ Serve skipped - not with us",
        "SERVE_NOT_FOUND": "⚠️ Serve skipped - not found",
        "SERVE_ERROR": "❌ Serve error",
        "SKIPPED_RESUME": "⏭️ Already processed",
        "SKIPPED_MISSING_SSID": "⚠️ Missing SSID — skipped",
        "NEEDS_INPUT_DATA": "⚠️ Missing/invalid required data — skipped",
        "ERROR": "❌ Error",
    }

    def __init__(self, parent, summary: dict, run_rows: list[dict], all_rows: list[dict], transfers_open_rows: list[dict]):
        super().__init__(parent)
        self.setWindowTitle("Run Results — Workabili-Bot3000")
        self.setWindowModality(Qt.ApplicationModal)
        self.setWindowFlag(Qt.WindowStaysOnTopHint, True)
        self.resize(980, 720)

        self.summary = summary or {}
        self.run_rows = run_rows or []
        self.all_rows = all_rows or []
        self.transfers_open_rows = transfers_open_rows or []

        outer = QVBoxLayout(self)
        outer.setContentsMargins(16, 14, 16, 14)
        outer.setSpacing(12)

        # Header strip (counts)
        counts = {
            "created": int(self.summary.get("created", 0)),
            "transfer_requested": int(self.summary.get("transfer_requested", 0)),
            "transfer_pending": int(self.summary.get("transfer_pending", 0)),
            "transferred_prior_year": int(self.summary.get("transferred_prior_year", 0)),
            "already_owned": int(self.summary.get("already_owned", 0)),
            "skipped": int(self.summary.get("skipped", 0)),
            "errors": int(self.summary.get("errors", 0)),
        }

        hdr = QLabel("Run Results")
        hdr.setStyleSheet("QLabel { font-size: 18px; font-weight: 800; color: #e2e8f0; }")
        outer.addWidget(hdr)

        strip = QLabel(
            f"<b>✅ Created:</b> {counts['created']} &nbsp;&nbsp; "
            f"<b>📨+⏳ Transfers:</b> {counts['transfer_requested'] + counts['transfer_pending'] + counts['transferred_prior_year']} &nbsp;&nbsp; "
            f"<b>🏫 Already with us:</b> {counts['already_owned']} &nbsp;&nbsp; "
            f"<b>⏭️ Skipped:</b> {counts['skipped']} &nbsp;&nbsp; "
            f"<b>❌ Errors:</b> {counts['errors']}"
        )
        strip.setTextFormat(Qt.RichText)
        strip.setStyleSheet("QLabel { font-size: 13px; color: #cbd5e1; }")
        outer.addWidget(strip)

        # Tabs
        self.tabs = QTabWidget()
        self.tabs.setStyleSheet("""
            QTabWidget::pane { border: 1px solid #24324a; border-radius: 12px; background: #0b1224; }
            QTabBar::tab {
                padding: 8px 12px;
                border-top-left-radius: 10px;
                border-top-right-radius: 10px;
                background: #0f172a;
                color: #94a3b8;
                border: 1px solid #24324a;
                margin-right: 6px;
            }
            QTabBar::tab:selected {
                background: #111b34;
                color: #e2e8f0;
                border: 1px solid #334155;
            }
        """)
        outer.addWidget(self.tabs, stretch=1)

        # Build tabs
        self._tab_map = {}

        self._add_table_tab("Created", self._filter_run(action="CREATED"))
        self._add_table_tab("Served", self._filter_run(actions={"SERVE_DRY_RUN_FILLED", "SERVE_SAVED", "SERVE_SKIPPED_NOT_OWNED", "SERVE_NOT_FOUND"}))
        self._add_transfers_tab()
        self._add_table_tab("Already With Us", self._filter_run(action="ALREADY_OWNED"))
        self._add_table_tab("Skipped", self._filter_run(actions={"SKIPPED_RESUME", "SKIPPED_MISSING_SSID", "NEEDS_INPUT_DATA"}))
        self._add_table_tab("Errors", self._filter_run(actions={"ERROR", "SERVE_ERROR"}))
        self._add_table_tab("All", self._filter_run(actions=None))

        # Footer buttons
        btn_row = QHBoxLayout()
        btn_row.addStretch(1)

        self.btn_copy = QPushButton("📋 Copy List")
        self.btn_copy.setToolTip("Copy the currently visible table to clipboard.")
        self.btn_copy.clicked.connect(self._copy_current_tab)

        self.btn_close = QPushButton("Close")
        self.btn_close.clicked.connect(self.accept)

        # Copy List = dark / low-attention (like the old Close)
        self.btn_copy.setStyleSheet("""
            QPushButton {
                min-width: 120px;
                min-height: 34px;
                font-family: "Segoe UI Variable Text", "Segoe UI";
                font-size: 13px;
                font-weight: 600;
                border-radius: 12px;
                padding: 6px 12px;
                border: 1px solid rgba(148, 163, 184, 90);
                background: rgba(11, 18, 36, 220);
                color: rgba(226, 232, 240, 220);
            }
            QPushButton:hover {
                background: rgba(17, 27, 52, 220);
                border: 1px solid rgba(148, 163, 184, 130);
            }
            QPushButton:pressed {
                background: rgba(10, 16, 32, 230);
                border: 1px solid rgba(148, 163, 184, 160);
            }
        """)

        # Close = primary (like the old Copy)
        self.btn_close.setStyleSheet("""
            QPushButton {
                min-width: 120px;
                min-height: 34px;
                font-family: "Segoe UI Variable Text", "Segoe UI";
                font-size: 13px;
                font-weight: 600;
                border-radius: 12px;
                padding: 6px 12px;
                border: 1px solid rgba(148, 163, 184, 90);
                background: rgba(56, 189, 248, 60);
                color: #ffffff;
            }
            QPushButton:hover {
                background: rgba(56, 189, 248, 90);
                border: 1px solid rgba(148, 163, 184, 130);
            }
            QPushButton:pressed {
                background: rgba(56, 189, 248, 120);
                border: 1px solid rgba(148, 163, 184, 160);
            }
        """)

        btn_row.addWidget(self.btn_copy)
        btn_row.addWidget(self.btn_close)
        outer.addLayout(btn_row)

        # Default transfers scope
        self.set_transfers_scope("This Run")

    # -------------------------
    # Public controls
    # -------------------------
    def select_tab(self, name: str) -> None:
        idx = self._tab_map.get(name)
        if idx is not None:
            self.tabs.setCurrentIndex(idx)

    def set_transfers_scope(self, scope_label: str) -> None:
        if hasattr(self, "transfer_scope") and self.transfer_scope:
            i = self.transfer_scope.findText(scope_label)
            if i >= 0:
                self.transfer_scope.setCurrentIndex(i)
                self._refresh_transfers()

    # -------------------------
    # Filtering + tab creation
    # -------------------------
    def _filter_run(self, action: str | None = None, actions: set[str] | None = None) -> list[dict]:
        if action is not None:
            return [r for r in self.run_rows if (r.get("action") == action)]
        if actions is not None:
            return [r for r in self.run_rows if (r.get("action") in actions)]
        return list(self.run_rows)

    def _ordinal(self, n: int) -> str:
        # 1 -> 1st, 2 -> 2nd, 3 -> 3rd, 4 -> 4th ...
        if 10 <= (n % 100) <= 20:
            suf = "th"
        else:
            suf = {1: "st", 2: "nd", 3: "rd"}.get(n % 10, "th")
        return f"{n}{suf}"

    def _format_ts(self, iso_ts: str) -> str:
        """
        Converts ledger ISO timestamp into human-friendly string like:
        "Mon, Feb 7th 8:37 PM"
        """
        s = (iso_ts or "").strip()
        if not s:
            return ""
        try:
            dt = datetime.fromisoformat(s)  # ledger writes utcnow().isoformat() (naive)
            dow = dt.strftime("%a")
            mon = dt.strftime("%b")
            day = self._ordinal(dt.day)
            tm = dt.strftime("%I:%M %p").lstrip("0")
            return f"{dow}, {mon} {day} {tm}"
        except Exception:
            return s

    def _make_table(self) -> QTableWidget:
        t = QTableWidget()
        t.setColumnCount(5)
        t.setHorizontalHeaderLabels(["Student", "SSID", "Category", "Timestamp", "Details"])
        t.verticalHeader().setVisible(False)
        t.setEditTriggers(QAbstractItemView.NoEditTriggers)
        t.setSelectionBehavior(QAbstractItemView.SelectRows)
        t.setSelectionMode(QAbstractItemView.SingleSelection)
        t.setAlternatingRowColors(True)
        # Make Student column narrower; hard-lock it so Qt can't auto-grow it later.
        t.horizontalHeader().setStretchLastSection(False)

        t.horizontalHeader().setSectionResizeMode(0, QHeaderView.Fixed)
        t.setColumnWidth(0, 150)  # tighter (roughly half of typical stretched width)

        t.horizontalHeader().setSectionResizeMode(1, QHeaderView.ResizeToContents)
        t.horizontalHeader().setSectionResizeMode(2, QHeaderView.ResizeToContents)
        t.horizontalHeader().setSectionResizeMode(3, QHeaderView.ResizeToContents)
        t.horizontalHeader().setSectionResizeMode(4, QHeaderView.Stretch)
        t.setStyleSheet("""
            QTableWidget {
                border: 1px solid #24324a;
                border-radius: 12px;
                background: #0b1224;
                color: #e2e8f0;
                gridline-color: #24324a;
            }
            QHeaderView::section {
                background: #0f172a;
                color: #cbd5e1;
                border: 1px solid #24324a;
                padding: 6px 8px;
                font-weight: 700;
            }
        """)
        return t

    def _populate_table(self, table: QTableWidget, rows: list[dict]) -> None:
        table.setRowCount(0)
        for r in rows:
            student = (r.get("student_name") or "").strip()
            ssid = (r.get("ssid") or "").strip()
            action = (r.get("action") or "").strip()
            ts = self._format_ts((r.get("timestamp") or "").strip())
            details = (r.get("details") or "").strip()

            cat = self.ACTION_LABELS.get(action, action)

            i = table.rowCount()
            table.insertRow(i)

            table.setItem(i, 0, QTableWidgetItem(student))
            table.setItem(i, 1, QTableWidgetItem(ssid))
            table.setItem(i, 2, QTableWidgetItem(cat))
            table.setItem(i, 3, QTableWidgetItem(ts))
            table.setItem(i, 4, QTableWidgetItem(details))

    def _add_table_tab(self, title: str, rows: list[dict]) -> None:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)

        # Search box
        search = QLineEdit()
        search.setPlaceholderText("Search name/SSID/details…")
        search.setStyleSheet("""
            QLineEdit {
                border: 1px solid #24324a;
                border-radius: 12px;
                padding: 8px 10px;
                background: #070d1b;
                color: #e2e8f0;
            }
        """)
        lay.addWidget(search)

        table = self._make_table()
        lay.addWidget(table, stretch=1)

        # Store
        w._rows_all = list(rows)
        w._table = table
        w._search = search

        def apply_filter():
            q = (search.text() or "").strip().lower()
            if not q:
                self._populate_table(table, w._rows_all)
                return
            filtered = []
            for r in w._rows_all:
                hay = " ".join([
                    (r.get("student_name") or ""),
                    (r.get("ssid") or ""),
                    (r.get("action") or ""),
                    (r.get("details") or ""),
                    (r.get("timestamp") or ""),
                ]).lower()
                if q in hay:
                    filtered.append(r)
            self._populate_table(table, filtered)

        search.textChanged.connect(apply_filter)

        # Initial populate
        self._populate_table(table, w._rows_all)

        idx = self.tabs.addTab(w, title)
        self._tab_map[title] = idx

    def _add_transfers_tab(self) -> None:
        w = QWidget()
        lay = QVBoxLayout(w)
        lay.setContentsMargins(12, 12, 12, 12)
        lay.setSpacing(10)

        # Scope row: This Run vs All Open Transfers
        scope_row = QHBoxLayout()
        scope_lbl = QLabel("Scope:")
        scope_lbl.setStyleSheet("QLabel { color: #cbd5e1; font-weight: 700; }")
        scope_row.addWidget(scope_lbl)

        self.transfer_scope = QComboBox()
        self.transfer_scope.addItems(["This Run", "All Open Transfers"])
        self.transfer_scope.setStyleSheet("""
            QComboBox {
                border: 1px solid #24324a;
                border-top-left-radius: 10px;
                border-bottom-left-radius: 10px;
                border-top-right-radius: 0px;      /* right side square (avoid arrow clipping battles) */
                border-bottom-right-radius: 0px;
                padding: 6px 10px;
                background: #070d1b;
                color: #e2e8f0;
            }
        """)
        self.transfer_scope.currentTextChanged.connect(self._refresh_transfers)
        scope_row.addWidget(self.transfer_scope)

        scope_row.addStretch(1)

        self.btn_copy_followup = QPushButton("📋 Copy Follow-up List")
        self.btn_copy_followup.setToolTip("Copy Name — SSID — Transfer Status for the Transfers tab.")
        self.btn_copy_followup.setStyleSheet("""
            QPushButton {
                min-width: 180px;
                min-height: 32px;
                font-family: "Segoe UI Variable Text", "Segoe UI";
                font-size: 13px;
                font-weight: 700;
                border-radius: 12px;
                padding: 6px 12px;
                border: 1px solid rgba(148, 163, 184, 90);
                background: rgba(251, 191, 36, 70);
                color: #ffffff;
            }
            QPushButton:hover { background: rgba(251, 191, 36, 100); }
            QPushButton:pressed { background: rgba(251, 191, 36, 130); }
        """)
        self.btn_copy_followup.clicked.connect(self._copy_transfers_followup)
        scope_row.addWidget(self.btn_copy_followup)

        lay.addLayout(scope_row)

        # Search box
        self.transfer_search = QLineEdit()
        self.transfer_search.setPlaceholderText("Search transfers…")
        self.transfer_search.setStyleSheet("""
            QLineEdit {
                border: 1px solid #24324a;
                border-radius: 12px;
                padding: 8px 10px;
                background: #070d1b;
                color: #e2e8f0;
            }
        """)
        self.transfer_search.textChanged.connect(self._refresh_transfers)
        lay.addWidget(self.transfer_search)

        self.transfer_table = self._make_table()
        lay.addWidget(self.transfer_table, stretch=1)

        # Empty-state guidance (shown when "This Run" has no transfers)
        self.transfer_empty_lbl = QLabel(
            'No transfers on this run.  You can switch the dropdown above to "All Open Transfers" '
            "to see pending transfers from prior runs."
        )
        self.transfer_empty_lbl.setWordWrap(True)
        self.transfer_empty_lbl.setAlignment(Qt.AlignCenter)
        self.transfer_empty_lbl.setStyleSheet(
            "QLabel { color:#94a3b8; font-size:13px; padding:12px; }"
        )
        self.transfer_empty_lbl.setVisible(False)
        lay.addWidget(self.transfer_empty_lbl)

        idx = self.tabs.addTab(w, "Transfers")
        self._tab_map["Transfers"] = idx

        self._refresh_transfers()

    def _refresh_transfers(self) -> None:
        scope = self.transfer_scope.currentText() if hasattr(self, "transfer_scope") else "This Run"
        q = (self.transfer_search.text() or "").strip().lower() if hasattr(self, "transfer_search") else ""

        if scope == "All Open Transfers":
            base = list(self.transfers_open_rows)
        else:
            base = [r for r in self.run_rows if (r.get("action") in ("TRANSFER_REQUESTED", "TRANSFER_PENDING", "TRANSFERRED_PRIOR_YEAR"))]

        # Add a human status label into details/Category is already there
        filtered = []
        for r in base:
            # For display consistency, keep action as-is; table renders category label
            hay = " ".join([
                (r.get("student_name") or ""),
                (r.get("ssid") or ""),
                (r.get("action") or ""),
                (r.get("details") or ""),
                (r.get("timestamp") or ""),
            ]).lower()
            if (not q) or (q in hay):
                filtered.append(r)

        self._populate_table(self.transfer_table, filtered)

        # Show guidance when Transfers scope is "This Run" and there are zero transfers
        if hasattr(self, "transfer_empty_lbl") and self.transfer_empty_lbl:
            show_empty = (scope == "This Run") and (len(filtered) == 0)
            self.transfer_empty_lbl.setVisible(show_empty)
            # Hide the empty table when showing the message (cleaner UX)
            self.transfer_table.setVisible(not show_empty)

    # -------------------------
    # Copy helpers
    # -------------------------
    def _rows_from_table(self, table: QTableWidget) -> list[list[str]]:
        rows = []
        for r in range(table.rowCount()):
            row = []
            for c in range(table.columnCount()):
                item = table.item(r, c)
                row.append(item.text() if item else "")
            rows.append(row)
        return rows

    def _copy_current_tab(self) -> None:
        w = self.tabs.currentWidget()
        if not w:
            return

        # Transfers tab uses transfer_table
        if hasattr(self, "transfer_table") and w is self.tabs.widget(self._tab_map.get("Transfers", -1)):
            table = self.transfer_table
        else:
            table = getattr(w, "_table", None)

        if not table:
            return

        rows = self._rows_from_table(table)
        if not rows:
            QApplication.clipboard().setText("")
            return

        header = "\t".join(["Student", "SSID", "Category", "Timestamp", "Details"])
        body = "\n".join("\t".join(r) for r in rows)
        QApplication.clipboard().setText(header + "\n" + body)

    def _copy_transfers_followup(self) -> None:
        table = getattr(self, "transfer_table", None)
        if not table:
            return

        lines = []
        for r in range(table.rowCount()):
            name = table.item(r, 0).text() if table.item(r, 0) else ""
            ssid = table.item(r, 1).text() if table.item(r, 1) else ""
            cat = table.item(r, 2).text() if table.item(r, 2) else ""
            # cat is like "📨 Transfer requested" or "⏳ Transfer pending"
            lines.append(f"{name} — {ssid} — {cat}")

        QApplication.clipboard().setText("\n".join(lines))

def main():
    app = QApplication([])
    apply_modern_palette(app)

    # Main window — opacity set to 0 BEFORE show so the frame never flashes visible
    w = MainWindow()
    w.setWindowOpacity(0.0)
    w.show()

    # PNG splash (robot) — starts invisible behind video; crossfade reveals it with the app
    png_splash = PngSplashScreen(SPLASH_PNG_PATH)
    if not png_splash._pix.isNull():
        png_splash.setWindowOpacity(0.0)
        png_splash.show()
        png_filter = _SplashClickFilter(app, png_splash)
        app.installEventFilter(png_filter)
        app._workabili_png_splash = png_splash
        app._workabili_png_filter = png_filter

    # Video splash — plays on top; crossfades into app+PNG during last 10 frames
    splash = SplashScreen(SPLASH_PATH, "", end_sound_path=SPLASH_SOUND_PATH)
    if getattr(splash, "_valid", False):
        def _on_fade(t: float):
            w.setWindowOpacity(t)
            if not png_splash._pix.isNull():
                png_splash.setWindowOpacity(t)
        splash.fade_progress.connect(_on_fade)

        # Safety net: if anything goes wrong, ensure app is fully visible
        def _ensure_visible():
            w.setWindowOpacity(1.0)
            if not png_splash._pix.isNull():
                png_splash.setWindowOpacity(1.0)
        splash.video_finished.connect(_ensure_visible)

        splash.show()
        splash.raise_()
        splash.activateWindow()
        app._workabili_splash = splash

    app.exec()


if __name__ == "__main__":
    main()
