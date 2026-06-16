import os
import re
import difflib
import csv
import uuid
import yaml
import time
import unicodedata
from dataclasses import dataclass
from datetime import date, datetime
from getpass import getpass
from typing import Any, Dict, List, Optional, Set, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font

from playwright.sync_api import sync_playwright, expect
from playwright.sync_api import TimeoutError as PWTimeoutError


# =========================
# Ledger
# =========================

LEDGER_HEADERS = [
    "run_id", "timestamp", "ssid", "student_name",
    "action", "details", "screenshot", "trace"
]

def suggest_closest_match(target: str, candidates: list[str], cutoff: float = 0.6) -> str | None:
    """
    Returns the closest fuzzy match from candidates, or None.
    Uses case-insensitive normalized comparison.
    """
    if not target or not candidates:
        return None

    target_norm = norm(target).casefold()
    cand_map = {norm(c).casefold(): c for c in candidates}

    matches = difflib.get_close_matches(
        target_norm,
        list(cand_map.keys()),
        n=1,
        cutoff=cutoff,
    )
    if matches:
        return cand_map[matches[0]]
    return None

def preflight_validate_ui_mappings(page, cfg: Dict[str, Any], log_fn=None, run_id=None):
    """
    Validates that config-mapped UI values actually exist in the WAI dropdowns.
    This prevents typo-induced Playwright flake later.
    """
    def _l(msg):
        if log_fn and run_id:
            _log(log_fn, f"[{run_id}] {msg}")
        else:
            _log(log_fn, msg)

    _l("Preflight: reading UI options for Race dropdown...")
    ui_races = get_combobox_options(page, re.compile(r"Race", re.I))
    ui_races_norm = {norm(x).casefold() for x in ui_races}

    # cfg["race_mapping"] maps raw input -> UI value
    bad = []
    for raw, ui_val in (cfg.get("race_mapping") or {}).items():
        if norm(ui_val).casefold() not in ui_races_norm:
            bad.append((raw, ui_val))

    if bad:
        # show first few to keep message readable
        sample = bad[:10]
        raise RuntimeError(
            "CONFIG_MAPPING_ERROR: Some race_mapping UI values do not exist in WAI Race dropdown.\n"
            f"Examples (raw -> ui_value): {sample}\n"
            f"UI Race options: {ui_races}"
        )

    _l("Preflight: race_mapping UI values all match WAI dropdown options. ✅")

def get_combobox_options(page, label_regex, timeout_ms: int = 8000) -> List[str]:
    combo = page.get_by_label(label_regex)
    expect(combo).to_be_visible(timeout=timeout_ms)
    expect(combo).to_be_enabled(timeout=timeout_ms)

    page.keyboard.press("Escape")
    page.wait_for_timeout(100)
    combo.click(timeout=3000)

    listbox = page.get_by_role("listbox")
    expect(listbox).to_be_visible(timeout=timeout_ms)

    opts = [s.strip() for s in listbox.get_by_role("option").all_inner_texts()]
    opts = [s for s in opts if s]

    page.keyboard.press("Escape")
    try:
        expect(listbox).to_be_hidden(timeout=2000)
    except Exception:
        pass

    return opts

def wait_for_toast_text(page, pattern, appear_timeout_ms: int = 8000, disappear_timeout_ms: int = 8000):
    """
    Waits for a transient toast/snackbar containing `pattern` to appear (and optionally disappear).
    `pattern` can be a compiled regex or string.
    """
    toast = page.get_by_text(pattern).first
    toast.wait_for(state="visible", timeout=appear_timeout_ms)

    # Optional: wait for it to go away (useful to ensure save is truly finished)
    try:
        toast.wait_for(state="hidden", timeout=disappear_timeout_ms)
    except PWTimeoutError:
        pass


def click_save_robust(
    page,
    timeout_ms: int = 15000,
    log_fn=None,
    run_id=None,
    toast_regex: re.Pattern | None = None,
    toast_appear_timeout_ms: int = 8000,
    toast_disappear_timeout_ms: int = 8000,
):
    def _l(msg):
        if log_fn and run_id:
            _log(log_fn, f"[{run_id}] {msg}")
        elif log_fn:
            _log(log_fn, msg)

    # Prefer a visible Save button (and usually the bottom one is the "final" save)
    save_btn = page.locator("button:has-text('Save'):visible").last

    # 1) Ensure no combobox/listbox is still open (ESC is cheap + safe)
    page.keyboard.press("Escape")
    page.wait_for_timeout(150)

    # 2) Wait for button to be visible + enabled
    expect(save_btn).to_be_visible(timeout=timeout_ms)
    expect(save_btn).to_be_enabled(timeout=timeout_ms)

    # 3) Scroll into view (some layouts require this)
    try:
        save_btn.scroll_into_view_if_needed(timeout=3000)
    except Exception:
        pass

    # 4) Click with retries
    last_err = None
    for attempt in range(1, 4):
        try:
            _l(f"Clicking Save (attempt {attempt}/3)...")
            save_btn.click(timeout=5000)

            # 5) Optional: wait for success toast text (this is the "truth signal")
            if toast_regex is not None:
                toast = page.get_by_text(toast_regex).first
                _l(f"Waiting for toast: {toast_regex.pattern!r}")
                toast.wait_for(state="visible", timeout=toast_appear_timeout_ms)

                # Best-effort: wait for it to go away (ensures UI finished its save cycle)
                try:
                    toast.wait_for(state="hidden", timeout=toast_disappear_timeout_ms)
                except PWTimeoutError:
                    pass

            return

        except Exception as e:
            last_err = e
            _l(f"Save click attempt {attempt} failed: {e}")

            # Close any portal/overlay and try again
            page.keyboard.press("Escape")
            page.wait_for_timeout(250)

            # Re-assert enabled/visible (can re-render)
            try:
                expect(save_btn).to_be_visible(timeout=3000)
                expect(save_btn).to_be_enabled(timeout=3000)
            except Exception:
                pass

    raise RuntimeError(f"Failed to click Save after retries: {last_err}")


def load_ledger_state_xlsx(ledger_path: str) -> Tuple[Set[str], Dict[str, str], Set[str]]:
    """
    Returns:
      completed: SSIDs that are terminal and should be skipped on resume
      last_action_by_ssid: last recorded action for each SSID (for SKIPPED_RESUME details)
      transfer_requested_ssids: legacy return value; transfer actions are no longer
        treated as terminal because released students must be rechecked live.
    """
    completed: Set[str] = set()
    last_action_by_ssid: Dict[str, str] = {}
    transfer_requested_ssids: Set[str] = set()

    if not os.path.exists(ledger_path):
        return completed, last_action_by_ssid, transfer_requested_ssids

    wb = load_workbook(ledger_path, read_only=True, data_only=True)
    ws = wb.active

    headers = [
        str(c.value).strip() if c.value is not None else ""
        for c in next(ws.iter_rows(min_row=1, max_row=1))
    ]
    idx = {h: i for i, h in enumerate(headers)}

    def get_cell(row, col_name: str) -> str:
        i = idx.get(col_name)
        if i is None:
            return ""
        v = row[i].value
        return str(v).strip() if v is not None else ""

    for r in ws.iter_rows(min_row=2):
        ssid = get_cell(r, "ssid")
        action = get_cell(r, "action")
        if not ssid:
            continue

        if action:
            last_action_by_ssid[ssid] = action

        # Terminal outcomes. Transfer requests/pending are intentionally not terminal:
        # a later run may find that the releasing org has approved the transfer.
        if action in {
            "CREATED",
            "ALREADY_OWNED",
            "SKIPPED_MISSING_SSID",
            "TRANSFERRED_PRIOR_YEAR",
        }:
            completed.add(ssid)


    return completed, last_action_by_ssid, transfer_requested_ssids


def append_ledger_xlsx(path: str, row: Dict[str, str]) -> None:
    ensure_dir(os.path.dirname(path))

    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Ledger"
        ws.append(LEDGER_HEADERS)

        # Freeze header row
        ws.freeze_panes = "A2"

        # Bold header row
        for cell in ws[1]:
            cell.font = Font(bold=True)

        default_widths = {
            "run_id": 10,
            "timestamp": 22,
            "ssid": 12,
            "student_name": 24,
            "action": 28,
            "details": 60,
            "screenshot": 50,
            "trace": 45,
        }
        for i, h in enumerate(LEDGER_HEADERS, start=1):
            ws.column_dimensions[get_column_letter(i)].width = default_widths.get(h, 20)

    ws.append([row.get(h, "") for h in LEDGER_HEADERS])
    wb.save(path)


def select_ethnicity_radio(page, choice_text: str):
    block = page.locator("div").filter(
        has_text=re.compile(r"Ethnicity\s*-\s*Is this student Hispanic or Latino", re.I)
    ).first
    block.get_by_text(choice_text, exact=True).click()


def save_transfer_pending_screenshot(page, run_id: str, ssid: str) -> str:
    ensure_dir("output/screenshots")
    path = os.path.join("output", "screenshots", f"transfer_pending_{run_id}_{ssid}.png")
    page.screenshot(path=path, full_page=False)
    return path


def ensure_dir(path: str) -> None:
    os.makedirs(path, exist_ok=True)


def append_text_log(path: str, msg: str) -> None:
    ensure_dir(os.path.dirname(path))
    with open(path, "a", encoding="utf-8") as f:
        f.write(msg.rstrip() + "\n")


def log_input_audit(path: str, run_id: str, msg: str, log_fn=None) -> None:
    stamped = f"[{run_id}] {msg}"
    _log(log_fn, stamped)
    append_text_log(path, f"{datetime.utcnow().isoformat()} {stamped}")


def make_input_audit_logger(path: str, run_id: str, log_fn=None):
    def _audit(msg: str) -> None:
        log_input_audit(path, run_id, msg, log_fn=log_fn)

    return _audit


def get_found_row_owning_org(page, expected_ssid: str, timeout_ms: int = 5000) -> str:
    row = page.locator("tr").filter(has_text=re.compile(re.escape(expected_ssid))).first
    row.wait_for(state="visible", timeout=timeout_ms)

    owning_td = row.locator("td").nth(4)
    owning_td.wait_for(state="visible", timeout=timeout_ms)

    # Wait briefly for text to populate (handles late render)
    for _ in range(max(1, timeout_ms // 100)):
        txt = " ".join((owning_td.inner_text() or "").split())
        if txt:
            return txt
        page.wait_for_timeout(100)

    return " ".join((owning_td.inner_text() or "").split())



def is_already_owned_by_us(found_owning_org: str, cfg: Dict[str, Any]) -> bool:
    found = norm(found_owning_org).casefold()
    if not found:
        return False

    workability_cfg = cfg.get("workability", {}) or {}
    targets = workability_cfg.get("owning_org_names")

    if not targets:
        targets = [workability_cfg.get("owning_org_name", "")]

    return any(found == norm(target).casefold() for target in targets if norm(target))


def _log(log_fn, msg: str) -> None:
    if log_fn:
        log_fn(msg)
    else:
        print(msg)


# =========================
# Config + Mapping
# =========================

def norm(s: Any) -> str:
    return " ".join(str(s or "").strip().split())


EXCEL_ROW_NUM_KEY = "__excel_row_num__"
ADDRESS_COLUMN_KEYS = [
    "street_address",
    "city",
    "state",
    "zip",
    "phone_number",
    "parent_name",
]
ADDRESS_REQUIRED_COLUMN_KEYS = [
    "street_address",
    "city",
    "state",
    "zip",
    "phone_number",
    "parent_name",
]


def school_lookup_key(s: Any) -> str:
    return norm(s).upper()


EXCEL_TEXT_REPLACEMENTS = {
    "\u00a0": (" ", "converted nonbreaking spaces"),
    "\u2007": (" ", "converted nonstandard spaces"),
    "\u202f": (" ", "converted nonstandard spaces"),
    "\u2018": ("'", "normalized smart apostrophes"),
    "\u2019": ("'", "normalized smart apostrophes"),
    "\u201a": ("'", "normalized smart apostrophes"),
    "\u201b": ("'", "normalized smart apostrophes"),
    "\u2032": ("'", "normalized prime/apostrophe-like marks"),
    "\u2035": ("'", "normalized prime/apostrophe-like marks"),
    "\u201c": ('"', "normalized smart quotes"),
    "\u201d": ('"', "normalized smart quotes"),
    "\u201e": ('"', "normalized smart quotes"),
    "\u201f": ('"', "normalized smart quotes"),
    "\u2033": ('"', "normalized quote-like marks"),
}

EXCEL_TEXT_REMOVALS = {
    "\u200b": "removed zero-width characters",
    "\u200c": "removed zero-width characters",
    "\u200d": "removed zero-width characters",
    "\ufeff": "removed byte-order marks",
}


def sanitize_excel_text(value: str) -> Tuple[str, List[str]]:
    cleaned_chars = []
    changes = set()

    for ch in value:
        replacement = EXCEL_TEXT_REPLACEMENTS.get(ch)
        if replacement:
            new_ch, reason = replacement
            cleaned_chars.append(new_ch)
            changes.add(reason)
            continue

        removal_reason = EXCEL_TEXT_REMOVALS.get(ch)
        if removal_reason:
            changes.add(removal_reason)
            continue

        if ch in {"\r", "\n", "\t"}:
            cleaned_chars.append(" ")
            changes.add("converted tabs/newlines to spaces")
            continue

        cleaned_chars.append(ch)

    cleaned = "".join(cleaned_chars)
    collapsed = " ".join(cleaned.strip().split())
    if collapsed != cleaned:
        changes.add("trimmed/collapsed whitespace")

    return collapsed, sorted(changes)


def sanitize_excel_value(value: Any) -> Tuple[Any, List[str]]:
    if isinstance(value, str):
        return sanitize_excel_text(value)
    return value, []


def sanitize_person_name_value(value: Any) -> Tuple[str, List[str]]:
    if value is None:
        return "", []

    changes = set()
    if isinstance(value, str):
        text, text_changes = sanitize_excel_text(value)
        changes.update(text_changes)
    else:
        text = norm(value)

    cleaned_chars = []

    for ch in unicodedata.normalize("NFKD", text):
        if unicodedata.combining(ch):
            changes.add("removed diacritics from name")
            continue

        if ("A" <= ch <= "Z") or ("a" <= ch <= "z") or ch in {"-", "'", " "}:
            cleaned_chars.append(ch)
            continue

        if ch.isspace():
            cleaned_chars.append(" ")
            changes.add("normalized name whitespace")
            continue

        cleaned_chars.append(" ")
        changes.add("removed unsupported name characters")

    cleaned = " ".join("".join(cleaned_chars).strip().split())
    tokens = [token.strip("-'") for token in cleaned.split(" ")]
    stripped = " ".join(token for token in tokens if token)
    if stripped != cleaned:
        changes.add("trimmed unsupported name punctuation")
    cleaned = stripped

    if cleaned != text and not changes:
        changes.add("normalized name")

    return cleaned, sorted(changes)


def _excel_log_value(value: Any, max_len: int = 90) -> str:
    text = "" if value is None else str(value)
    escaped = text.encode("unicode_escape", errors="backslashreplace").decode("ascii")
    if len(escaped) > max_len:
        return escaped[: max_len - 3] + "..."
    return escaped


def log_excel_sanitization(
    log_fn,
    excel_row_num: int,
    excel_col_num: int,
    header: str,
    original: Any,
    cleaned: Any,
    changes: List[str],
) -> None:
    if not changes:
        return

    col = get_column_letter(excel_col_num)
    header_hint = f" ({header})" if header else ""
    _log(
        log_fn,
        "Excel sanitized "
        f"row {excel_row_num}, column {col}{header_hint}: "
        f"{'; '.join(changes)}; "
        f"{_excel_log_value(original)!r} -> {_excel_log_value(cleaned)!r}",
    )


def load_config(config_path: str) -> Dict[str, Any]:
    with open(config_path, "r", encoding="utf-8") as f:
        cfg = yaml.safe_load(f)

    sm = cfg.get("school_mapping", {}) or {}
    normalized = {}
    original_keys = {}
    collisions = {}

    for raw_key, val in sm.items():
        nk = school_lookup_key(raw_key)
        if nk in normalized and normalized[nk] != val:
            collisions.setdefault(nk, []).append(raw_key)
        normalized[nk] = val
        original_keys[nk] = norm(raw_key)

    cfg["school_mapping"] = normalized
    cfg["_school_mapping_original_keys"] = original_keys

    if collisions:
        print("WARNING: school_mapping key collisions after normalization:")
        for nk, raw_keys in collisions.items():
            print(f"  {nk!r} <= {raw_keys}")

    return cfg


def normalized_integral_text(raw: Any) -> str:
    if raw is None:
        return ""
    if isinstance(raw, bool):
        return norm(raw)
    if isinstance(raw, int):
        return str(raw)
    if isinstance(raw, float) and raw.is_integer():
        return str(int(raw))
    return norm(raw)


def map_grade(raw_grade: Any, cfg: Dict[str, Any]) -> Optional[str]:
    return cfg["grade_mapping"].get(normalized_integral_text(raw_grade))


def map_disability(raw_dis: Any, cfg: Dict[str, Any]) -> Optional[str]:
    return cfg["disability_mapping"].get(norm(raw_dis))


def map_race(raw_race: Any, cfg: Dict[str, Any], validation_log_fn=None, excel_row_num=None) -> str:
    raw = norm(raw_race)
    mapped = cfg["race_mapping"].get(raw)
    if mapped:
        return mapped

    msg = f"Race value {raw!r} is not in race_mapping; falling back to 'Other'."
    if excel_row_num:
        msg = f"Excel row {excel_row_num}: {msg}"
    if validation_log_fn:
        _log(validation_log_fn, msg)
    return "Other"


def resolve_school(aeries_site: Any, cfg: Dict[str, Any], validation_log_fn=None, excel_row_num=None) -> str:
    raw = norm(aeries_site)
    key = school_lookup_key(raw)
    entry = cfg["school_mapping"].get(key)
    if not entry:
        raise KeyError(f"No school mapping for Aeries site raw={aeries_site!r} normalized={key!r}")
    if not entry.get("active", True):
        raise ValueError(f"Mapped WAI school is inactive/closed: {entry.get('wai_school')!r}")

    original_key = (cfg.get("_school_mapping_original_keys") or {}).get(key, key)
    if raw and raw != original_key:
        msg = f"School name {raw!r} matched config key {original_key!r} using case-insensitive lookup."
        if excel_row_num:
            msg = f"Excel row {excel_row_num}: {msg}"
        if validation_log_fn:
            _log(validation_log_fn, msg)

    return entry["wai_school"]


def parse_dob(raw: Any) -> date:
    if isinstance(raw, datetime):
        return raw.date()
    if isinstance(raw, date):
        return raw

    s = norm(raw)
    if not s:
        raise ValueError("Birthdate is required")

    for fmt in ("%m/%d/%Y", "%m/%d/%y", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass

    raise ValueError(f"Invalid DOB format/date {s!r} (expected M/D/YYYY)")


def age_on(dob: date, today: date) -> int:
    return today.year - dob.year - ((today.month, today.day) < (dob.month, dob.day))


def normalize_dob_mmddyyyy(raw: Any) -> str:
    dob = parse_dob(raw)
    return f"{dob.month:02d}/{dob.day:02d}/{dob.year:04d}"


def contains_alpha(s: str) -> bool:
    return any(ch.isalpha() for ch in s)


@dataclass
class Student:
    first_name: str
    last_name: str
    ssid: str
    dob: str
    gender_ui: str
    grade_ui: str
    disability_ui: str
    ethnicity_ui: str
    race_ui: str
    aeries_school: str
    wai_school: str
    street_address: str = ""
    city: str = ""
    state: str = ""
    zip_code: str = ""
    phone_number: str = ""
    parent_name: str = ""


@dataclass
class PreparedInputRow:
    excel_row_num: int
    student: Student
    display_name: str


@dataclass
class InvalidInputRow:
    excel_row_num: int
    display_name: str
    ssid: str
    action: str
    details: str


def row_display_name(row: Dict[str, Any], cfg: Dict[str, Any]) -> str:
    cols = cfg["columns"]
    fn = norm(row.get(cols["first_name"]))
    ln = norm(row.get(cols["last_name"]))
    return f"{fn} {ln}".strip()


def validate_and_prepare(
    row: Dict[str, Any],
    cfg: Dict[str, Any],
    validation_log_fn=None,
    today: date | None = None,
    require_address_fields: bool = False,
) -> Student:
    cols = cfg["columns"]
    today = today or date.today()
    excel_row_num = row.get(EXCEL_ROW_NUM_KEY)
    errors: List[str] = []

    def add_error(msg: str) -> None:
        errors.append(msg)

    first_name_raw = row.get(cols["first_name"])
    first_name, first_name_changes = sanitize_person_name_value(first_name_raw)
    if first_name_changes and validation_log_fn:
        _log(
            validation_log_fn,
            f"Excel row {excel_row_num}: First Name sanitized for WAI name field: "
            f"{_excel_log_value(first_name_raw)!r} -> {_excel_log_value(first_name)!r} "
            f"({'; '.join(first_name_changes)}).",
        )
    if not first_name:
        if norm(first_name_raw):
            add_error("First Name must contain at least one alphabetic character")
        else:
            add_error("First Name is required")
    elif not contains_alpha(first_name):
        add_error("First Name must contain at least one alphabetic character")

    last_name_raw = row.get(cols["last_name"])
    last_name, last_name_changes = sanitize_person_name_value(last_name_raw)
    if last_name_changes and validation_log_fn:
        _log(
            validation_log_fn,
            f"Excel row {excel_row_num}: Last Name sanitized for WAI name field: "
            f"{_excel_log_value(last_name_raw)!r} -> {_excel_log_value(last_name)!r} "
            f"({'; '.join(last_name_changes)}).",
        )
    if not last_name:
        if norm(last_name_raw):
            add_error("Last Name must contain at least one alphabetic character")
        else:
            add_error("Last Name is required")
    elif not contains_alpha(last_name):
        add_error("Last Name must contain at least one alphabetic character")

    ssid = normalized_integral_text(row.get(cols["ssid"]))
    if not ssid:
        add_error("Missing SSID")
    elif not re.fullmatch(r"\d{10}", ssid):
        add_error(f"State Student ID must be exactly 10 digits; got {ssid!r}")

    validation_cfg = cfg.get("validation", {}) or {}

    gender_input = norm(row.get(cols["gender"]))
    gender_raw = gender_input.upper()
    allowed_gender = [str(x).upper() for x in validation_cfg.get("allowed_gender", [])]
    if not gender_raw:
        add_error("Gender is required")
    elif gender_raw not in allowed_gender:
        add_error(f"Invalid gender {gender_input!r} (allowed {allowed_gender})")
    elif gender_input and gender_input != gender_raw and validation_log_fn:
        _log(validation_log_fn, f"Excel row {excel_row_num}: Gender {gender_input!r} normalized to {gender_raw!r}.")
    gender_ui = cfg["gender_mapping"].get(gender_raw, "")

    dob_ui = ""
    try:
        dob = parse_dob(row.get(cols["dob"]))
        min_age_years = int(validation_cfg.get("min_age_years", 13))
        student_age = age_on(dob, today)
        if student_age < min_age_years:
            add_error(f"Birthdate indicates age {student_age}; student must be at least {min_age_years}")
        dob_ui = f"{dob.month:02d}/{dob.day:02d}/{dob.year:04d}"
    except ValueError as e:
        add_error(str(e))

    grade_ui = map_grade(row.get(cols["grade"]), cfg)
    grade_raw = normalized_integral_text(row.get(cols["grade"]))
    allowed_grades = [str(x) for x in validation_cfg.get("allowed_grades", ["7", "8", "9", "10", "11", "12"])]
    if not grade_raw:
        add_error("Grade is required")
    elif grade_raw not in allowed_grades:
        add_error(f"Invalid grade {grade_raw!r} (allowed {allowed_grades})")
    elif not grade_ui:
        add_error(f"Invalid grade {grade_raw!r} (no mapping)")

    disability_ui = map_disability(row.get(cols["disability"]), cfg)
    disability_raw = norm(row.get(cols["disability"]))
    if not disability_raw:
        add_error("Description_CSE_DI is required")
    elif not disability_ui:
        add_error(f"Invalid disability {disability_raw!r} (not in disability_mapping)")

    eth_input = norm(row.get(cols["ethcd"]))
    eth_raw = eth_input.upper()
    allowed_ethcd = [str(x).upper() for x in validation_cfg.get("allowed_ethcd", ["Y", "N"])]
    eth_map = {"Y": "Yes", "N": "No"}
    ethnicity_ui = eth_map.get(eth_raw)
    if not eth_raw:
        add_error("EthCd is required")
    elif eth_raw not in allowed_ethcd:
        add_error(f"Invalid ethnicity code {eth_input!r} (allowed {allowed_ethcd})")
    elif eth_input and eth_input != eth_raw and validation_log_fn:
        _log(validation_log_fn, f"Excel row {excel_row_num}: EthCd {eth_input!r} normalized to {eth_raw!r}.")

    race_raw = norm(row.get(cols["race1"]))
    race_ui = ""
    if not race_raw:
        add_error("Description_STU_RC1 is required")
    else:
        race_ui = map_race(
            race_raw,
            cfg,
            validation_log_fn=validation_log_fn,
            excel_row_num=excel_row_num,
        )

    aeries_school = norm(row.get(cols["aeries_school"]))
    wai_school = ""
    if not aeries_school:
        add_error("School name is required")
    else:
        try:
            wai_school = resolve_school(
                aeries_school,
                cfg,
                validation_log_fn=validation_log_fn,
                excel_row_num=excel_row_num,
            )
        except (KeyError, ValueError) as e:
            add_error(str(e))

    street_address = norm(row.get(cols.get("street_address", "")))
    city = norm(row.get(cols.get("city", "")))
    state = norm(row.get(cols.get("state", ""))).upper()
    zip_code = normalized_integral_text(row.get(cols.get("zip", "")))
    phone_number = norm(row.get(cols.get("phone_number", "")))

    parent_name_raw = row.get(cols.get("parent_name", ""))
    parent_name, parent_name_changes = sanitize_person_name_value(parent_name_raw)
    if parent_name_changes and validation_log_fn:
        _log(
            validation_log_fn,
            f"Excel row {excel_row_num}: Parent Name sanitized for WAI contact field: "
            f"{_excel_log_value(parent_name_raw)!r} -> {_excel_log_value(parent_name)!r} "
            f"({'; '.join(parent_name_changes)}).",
        )

    if require_address_fields:
        if not street_address:
            add_error("Street address is required for address patch")
        if not city:
            add_error("City is required for address patch")
        if not state:
            add_error("State is required for address patch")
        elif not re.fullmatch(r"[A-Z]{2}", state):
            add_error(f"State must be a 2-letter code for address patch; got {state!r}")
        if not zip_code:
            add_error("Zip is required for address patch")
        elif not re.fullmatch(r"\d{5}(?:-\d{4})?", zip_code):
            add_error(f"Zip must be 5 digits or ZIP+4 for address patch; got {zip_code!r}")
        if not phone_number:
            add_error("Phone Number is required for address patch")
        if not parent_name:
            if norm(parent_name_raw):
                add_error("Parent Name must contain at least one alphabetic character for address patch")
            else:
                add_error("Parent Name is required for address patch")

    if errors:
        raise ValueError("; ".join(errors))

    return Student(
        first_name=first_name,
        last_name=last_name,
        ssid=ssid,
        dob=dob_ui,
        gender_ui=gender_ui,
        grade_ui=grade_ui,
        disability_ui=disability_ui,
        ethnicity_ui=ethnicity_ui,
        race_ui=race_ui,
        aeries_school=aeries_school,
        wai_school=wai_school,
        street_address=street_address,
        city=city,
        state=state,
        zip_code=zip_code,
        phone_number=phone_number,
        parent_name=parent_name,
    )


def prepare_input_rows(
    rows: List[Dict[str, Any]],
    cfg: Dict[str, Any],
    validation_log_fn=None,
    today: date | None = None,
    require_address_fields: bool = False,
) -> Tuple[List[PreparedInputRow], List[InvalidInputRow]]:
    prepared: List[PreparedInputRow] = []
    invalid: List[InvalidInputRow] = []
    cols = cfg["columns"]
    seen_ssid_rows: Dict[str, int] = {}

    for i, row in enumerate(rows, start=1):
        excel_row_num = int(row.get(EXCEL_ROW_NUM_KEY) or (i + 1))
        display_name = row_display_name(row, cfg)
        ssid_raw = normalized_integral_text(row.get(cols["ssid"]))

        try:
            student = validate_and_prepare(
                row,
                cfg,
                validation_log_fn=validation_log_fn,
                today=today,
                require_address_fields=require_address_fields,
            )
            if student.ssid in seen_ssid_rows:
                first_row = seen_ssid_rows[student.ssid]
                invalid.append(
                    InvalidInputRow(
                        excel_row_num=excel_row_num,
                        display_name=f"{student.first_name} {student.last_name}".strip(),
                        ssid=student.ssid,
                        action="NEEDS_INPUT_DATA",
                        details=(
                            f"Excel row {excel_row_num}: Duplicate State Student ID {student.ssid!r}; "
                            f"already present on Excel row {first_row}"
                        ),
                    )
                )
                continue
            seen_ssid_rows[student.ssid] = excel_row_num
            prepared.append(
                PreparedInputRow(
                    excel_row_num=excel_row_num,
                    student=student,
                    display_name=f"{student.first_name} {student.last_name}".strip(),
                )
            )
        except ValueError as e:
            action = "SKIPPED_MISSING_SSID" if not ssid_raw else "NEEDS_INPUT_DATA"
            invalid.append(
                InvalidInputRow(
                    excel_row_num=excel_row_num,
                    display_name=display_name,
                    ssid=ssid_raw,
                    action=action,
                    details=f"Excel row {excel_row_num}: Input data incomplete or invalid: {e}",
                )
            )

    return prepared, invalid


# =========================
# Excel Reader
# =========================

def expected_input_headers(cfg: Dict[str, Any]) -> List[str]:
    cols = cfg["columns"]
    return [
        cols["first_name"],
        cols["last_name"],
        cols["ssid"],
        cols["dob"],
        cols["gender"],
        cols["grade"],
        cols["disability"],
        cols["ethcd"],
        cols["race1"],
        cols["aeries_school"],
    ]


def expected_address_headers(cfg: Dict[str, Any]) -> List[str]:
    cols = cfg["columns"]
    return [cols[key] for key in ADDRESS_COLUMN_KEYS if cols.get(key)]


def validate_input_headers(
    headers: List[str],
    cfg: Dict[str, Any],
    require_address_columns: bool = False,
) -> None:
    expected = expected_input_headers(cfg)
    expected_address = expected_address_headers(cfg)
    errors = []

    for i, expected_header in enumerate(expected, start=1):
        actual = headers[i - 1] if i <= len(headers) else ""
        if actual != expected_header:
            errors.append(
                f"column {get_column_letter(i)} expected {expected_header!r}, got {actual!r}"
            )

    extra_headers = [h for h in headers[len(expected):] if norm(h)]
    allowed_prefix_len = 0
    for i in range(len(extra_headers) + 1):
        if extra_headers[:i] == expected_address[:i]:
            allowed_prefix_len = i
        else:
            break

    for offset, extra in enumerate(extra_headers[:allowed_prefix_len], start=1):
        expected_extra = expected_address[offset - 1]
        if extra != expected_extra:
            idx = len(expected) + offset
            errors.append(f"column {get_column_letter(idx)} expected {expected_extra!r}, got {extra!r}")

    if extra_headers[allowed_prefix_len:]:
        for offset, extra in enumerate(extra_headers[allowed_prefix_len:], start=allowed_prefix_len + 1):
            idx = len(expected) + offset
            expected_extra = expected_address[offset - 1] if offset - 1 < len(expected_address) else None
            if expected_extra:
                errors.append(f"column {get_column_letter(idx)} expected {expected_extra!r}, got {extra!r}")
            else:
                errors.append(f"unexpected extra header in column {get_column_letter(idx)}: {extra!r}")

    if require_address_columns and extra_headers != expected_address:
        if len(extra_headers) < len(expected_address):
            missing = expected_address[len(extra_headers):]
            errors.append(f"missing required address headers at end of sheet: {missing!r}")

    if errors:
        raise ValueError("INPUT_HEADER_ERROR: " + "; ".join(errors))


def read_excel_rows(
    excel_path: str,
    cfg: Dict[str, Any] | None = None,
    log_fn=None,
    require_address_columns: bool = False,
) -> List[Dict[str, Any]]:
    wb = load_workbook(excel_path, read_only=True, data_only=True)
    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))
    wb.close()
    if not rows:
        return []

    headers = []
    for i, raw_header in enumerate(rows[0], start=1):
        header_text = "" if raw_header is None else str(raw_header)
        clean_header, changes = sanitize_excel_text(header_text)
        log_excel_sanitization(log_fn, 1, i, "header", header_text, clean_header, changes)
        headers.append(clean_header)

    if cfg is not None:
        validate_input_headers(headers, cfg, require_address_columns=require_address_columns)

    out: List[Dict[str, Any]] = []
    name_headers = set()
    if cfg is not None:
        cols = cfg.get("columns", {}) or {}
        name_headers = {cols.get("first_name"), cols.get("last_name")}

    for excel_row_num, r in enumerate(rows[1:], start=2):
        cleaned_row = []
        for i, raw_value in enumerate(r, start=1):
            clean_value, changes = sanitize_excel_value(raw_value)
            header = headers[i - 1] if i <= len(headers) else ""
            if header in name_headers:
                name_value, name_changes = sanitize_person_name_value(clean_value)
                if name_changes:
                    clean_value = name_value
                    changes = sorted(set(changes + name_changes))
            log_excel_sanitization(log_fn, excel_row_num, i, header, raw_value, clean_value, changes)
            cleaned_row.append(clean_value)

        if all(v is None or str(v).strip() == "" for v in cleaned_row):
            continue

        row_dict = {headers[i]: cleaned_row[i] for i in range(min(len(headers), len(cleaned_row)))}
        row_dict[EXCEL_ROW_NUM_KEY] = excel_row_num
        out.append(row_dict)
    return out


def rows_have_address_columns(rows: List[Dict[str, Any]], cfg: Dict[str, Any]) -> bool:
    if not rows:
        return False
    cols = cfg["columns"]
    headers = set(rows[0].keys())
    return all(cols.get(key) in headers for key in ADDRESS_REQUIRED_COLUMN_KEYS)


def rows_have_partial_address_columns(rows: List[Dict[str, Any]], cfg: Dict[str, Any]) -> bool:
    if not rows:
        return False
    cols = cfg["columns"]
    headers = set(rows[0].keys())
    present = [cols.get(key) for key in ADDRESS_REQUIRED_COLUMN_KEYS if cols.get(key) in headers]
    return bool(present) and len(present) < len(ADDRESS_REQUIRED_COLUMN_KEYS)


# =========================
# Playwright Actions
# =========================

ZERO_RESULTS_RE = re.compile(r"The following\s+0\s+records match your search criteria", re.I)


def login(page, username: str, password: str, cfg: Dict[str, Any], log_fn=None) -> None:
    def _local(msg: str):
        if log_fn:
            log_fn(msg)
        else:
            print(msg)

    _local("Navigating to login page...")
    page.goto(cfg["workability"]["login_url"])

    user = page.get_by_placeholder("Username")
    pw = page.get_by_placeholder("Password")
    btn = page.get_by_role("button", name=re.compile(r"Log\s*In", re.I))

    expect(user).to_be_visible(timeout=15000)
    expect(pw).to_be_visible(timeout=15000)

    _local("Filling credentials...")
    user.fill(username)
    pw.fill(password)

    _local("Clicking Log In...")
    btn.click()

    student_data_link = page.get_by_role("link", name=re.compile(r"Student\s*Data", re.I))

    def bad_creds_detected() -> bool:
        try:
            if not user.is_visible() or not pw.is_visible():
                return False
            pw_val = pw.input_value()
            if (pw_val or "").strip() != "":
                return False
            user_val = user.input_value()
            if (user_val or "").strip() == "":
                return True
            return True
        except Exception:
            return False

    page.wait_for_timeout(600)

    if bad_creds_detected():
        raise RuntimeError("LOGIN_BAD_CREDS: username/password rejected (login form remained and password cleared).")

    try:
        student_data_link.wait_for(state="visible", timeout=6000)
        _local("Login appears successful.")
        return
    except PWTimeoutError:
        if bad_creds_detected():
            raise RuntimeError("LOGIN_BAD_CREDS: username/password rejected (silent refresh to blank login form).")
        raise RuntimeError("LOGIN_UNKNOWN: login did not reach post-login state and did not match bad-credentials pattern.")


def goto_student_records(page, cfg=None, log_fn=None, run_id=None):
    def _log_local(msg: str):
        if log_fn:
            log_fn(msg)
        else:
            print(msg)

    _log_local("Navigating to Student Records (hover-only, no parent clicks)...")

    close_transient_overlays(page)

    try:
        if "/student-data/search-student-records" in (page.url or "") and page.get_by_placeholder("SSID").is_visible():
            _log_local("Already on Student Records - using current page.")
            return
    except Exception:
        pass

    submenu_selector = "a[href='/student-data/search-student-records']"
    submenu = page.locator(submenu_selector)

    try:
        if submenu.is_visible():
            _log_local("Fast path: Student Records visible — clicking.")
            submenu.click()
            expect(page.get_by_placeholder("SSID")).to_be_visible(timeout=8000)
            return
    except Exception:
        pass

    parent_anchor = page.locator("a:has-text('Student Data')").first

    hover_attempts = 3
    hover_wait_ms = 350
    for attempt in range(1, hover_attempts + 1):
        try:
            _log_local(f"Hover attempt {attempt}/{hover_attempts} (force=True)...")
            parent_anchor.hover(force=True)
            page.wait_for_timeout(hover_wait_ms)
            if submenu.is_visible():
                _log_local(f"Student Records visible after hover attempt {attempt} — clicking.")
                submenu.click()
                expect(page.get_by_placeholder("SSID")).to_be_visible(timeout=8000)
                return
            _log_local(f"Not visible after hover attempt {attempt}.")
        except Exception as e:
            _log_local(f"Hover attempt {attempt} threw: {e}")

    try:
        _log_local("Dispatching mouseover/mouseenter/mousemove events via JS (no clicks)...")
        page.evaluate(
            """() => {
                const candidates = Array.from(document.querySelectorAll('a,button,div,span'));
                const parent = candidates.find(e => e.textContent && e.textContent.trim().match(/^Student\\s*Data/i));
                if (!parent) return false;
                ['mouseover','mouseenter','mousemove'].forEach(evName => {
                    const ev = new Event(evName, { bubbles: true, cancelable: true });
                    parent.dispatchEvent(ev);
                });
                return true;
            }"""
        )
        page.wait_for_timeout(400)
        if submenu.is_visible():
            _log_local("Student Records visible after dispatch events — clicking.")
            submenu.click()
            expect(page.get_by_placeholder("SSID")).to_be_visible(timeout=8000)
            return
        _log_local("Dispatch events did not reveal Student Records.")
    except Exception as e:
        _log_local(f"Dispatch events attempt failed: {e}")

    try:
        _log_local("Last-resort: removing 'hidden' class from submenu <ul> via JS (no clicking parent).")
        page.evaluate(
            """() => {
                const parent = Array.from(document.querySelectorAll('a')).find(a => a.textContent && a.textContent.trim().startsWith('Student Data'));
                if (!parent) return false;
                let ul = parent.parentElement ? parent.parentElement.querySelector('ul') : null;
                if (!ul) {
                    const sib = parent.nextElementSibling;
                    if (sib && sib.tagName && sib.tagName.toLowerCase() === 'ul') ul = sib;
                }
                if (!ul) return false;
                ul.classList.remove('hidden');
                ul.style.display = 'block';
                ul.style.maxHeight = '500px';
                return true;
            }"""
        )
        page.wait_for_timeout(300)
        if submenu.is_visible():
            _log_local("Student Records visible after DOM tweak — clicking.")
            submenu.click()
            expect(page.get_by_placeholder("SSID")).to_be_visible(timeout=8000)
            return
        _log_local("DOM tweak did not reveal Student Records.")
    except Exception as e:
        _log_local(f"DOM tweak attempt failed: {e}")

    try:
        ensure_dir("output/screenshots")
        debug_path = os.path.join("output", "screenshots", f"student_records_hover_debug_{run_id or 'noid'}.png")
        page.screenshot(path=debug_path, full_page=False)
        _log_local(f"Saved debug screenshot: {debug_path}")
    except Exception as e:
        _log_local(f"Failed to save debug screenshot: {e}")

    raise RuntimeError("Failed to reveal/click Student Records via hover-based strategies. Check screenshot.")


def close_transient_overlays(page, timeout_ms: int = 1000) -> None:
    """Close leftover dropdown portals that can intercept the next click."""
    for _ in range(2):
        try:
            listbox = page.get_by_role("listbox")
            visible = False
            for i in range(min(listbox.count(), 5)):
                try:
                    if listbox.nth(i).is_visible():
                        visible = True
                        break
                except Exception:
                    pass
            if not visible:
                break
            page.keyboard.press("Escape")
            page.wait_for_timeout(100)
        except Exception:
            break

    try:
        expect(page.get_by_role("listbox")).to_be_hidden(timeout=timeout_ms)
    except Exception:
        pass


def wait_for_search_loader_to_settle(page) -> None:
    loader = page.get_by_test_id("loader")
    try:
        loader.wait_for(state="visible", timeout=1500)
    except Exception:
        pass
    try:
        if loader.count() > 0 and loader.is_visible():
            try:
                loader.wait_for(state="hidden", timeout=10_000)
            except Exception:
                loader.wait_for(state="detached", timeout=10_000)
    except Exception:
        pass


def set_search_project_scope(page, cfg: Dict[str, Any]) -> None:
    target = cfg["workability"]["search_project"]
    page.get_by_label("WAI Project").click()
    page.get_by_label(target, exact=True).click()

    # Optional but recommended: wait for project-scope refresh to settle
    wait_for_search_loader_to_settle(page)


def set_search_program_year_all(page) -> None:
    select_combobox_option(page, re.compile(r"Program\s*Year", re.I), "All")
    wait_for_search_loader_to_settle(page)


def click_search_radio(page, label_text: str) -> None:
    radio = (
        page.locator("div")
        .filter(has_text=re.compile(rf"^{re.escape(label_text)}$"))
        .locator("#radio-group-item")
        .first
    )
    expect(radio).to_be_visible(timeout=8000)
    radio.click()


def set_search_scope_filters(page, cfg: Dict[str, Any]) -> None:
    set_search_project_scope(page, cfg)
    set_search_program_year_all(page)
    click_search_radio(page, "All Students")
    click_search_radio(page, "Any Baseline or Follow-Up Record")



def search_by_ssid(page, ssid: str, cfg: Dict[str, Any]) -> None:

    set_search_scope_filters(page, cfg)

    page.get_by_placeholder("SSID").fill(ssid)
    page.get_by_role("button", name="Search").click()

    # --- Wait for async search to settle (race-condition guard) ---
    wait_for_search_loader_to_settle(page)



def determine_search_outcome(page, timeout_ms: int = 3000) -> str:
    loader = page.get_by_test_id("loader")
    edit_btn = page.get_by_role("button", name="Edit")
    zero_msg = page.locator("p").filter(has_text=ZERO_RESULTS_RE).first

    step_ms = 100
    steps = max(1, timeout_ms // step_ms)

    # Give the loader a short chance to appear (handles "appears a beat later")
    try:
        loader.wait_for(state="visible", timeout=min(1500, timeout_ms))
    except Exception:
        pass

    for _ in range(steps):
        # If the loader is visible, the page is still resolving results.
        # Do NOT trust "0 records" or absence of Edit yet.
        try:
            if loader.count() > 0 and loader.is_visible():
                page.wait_for_timeout(step_ms)
                continue
        except Exception:
            # If loader lookup is flaky for any reason, don't fail the run; fall through.
            pass

        # Now we're in a "settled enough" moment — safe to evaluate outcome.
        try:
            if edit_btn.is_visible():
                return "FOUND"
        except Exception:
            pass

        try:
            if zero_msg.is_visible():
                return "NOT_FOUND"
        except Exception:
            pass

        page.wait_for_timeout(step_ms)

    raise RuntimeError(
        "Search outcome unclear: neither Edit nor zero-results message appeared within timeout, "
        "or two Edit buttons appeared - one for Baseline and one for Follow-Up and this flow isn't currently supported."
    )


def open_existing_student_edit(page) -> None:
    page.get_by_role("button", name="Edit").first.click()
    expect(page.locator("body")).to_be_visible()
    try:
        page.wait_for_load_state("domcontentloaded", timeout=10000)
    except Exception:
        pass
    page.get_by_text(re.compile(r"Student\s+Basics|Baseline\s+and\s+Follow", re.I)).first.wait_for(
        state="visible",
        timeout=10000,
    )


def request_transfer_if_possible(page) -> str:
    transfer_btn = page.get_by_label("Request Student Transfer")
    if transfer_btn.is_visible():
        transfer_btn.click()
        dialog = page.get_by_role("alertdialog")
        expect(dialog.get_by_role("heading", name=re.compile(r"Are you sure", re.I))).to_be_visible(timeout=5000)
        dialog.get_by_role("button", name="Yes").click()
        expect(dialog).to_be_hidden(timeout=5000)
        return "TRANSFER_REQUESTED"
    return "TRANSFER_PENDING"


def get_prior_year_transfer_control(page, timeout_ms: int = 8000):
    deadline = time.monotonic() + (timeout_ms / 1000)

    while True:
        candidates = [
            page.get_by_role("button", name=re.compile(r"^Transfer$", re.I)).first,
            page.get_by_role("link", name=re.compile(r"^Transfer$", re.I)).first,
            page.get_by_text("Transfer", exact=True).first,
        ]

        for candidate in candidates:
            try:
                if candidate.count() > 0 and candidate.is_visible():
                    return candidate
            except Exception:
                pass

        if time.monotonic() >= deadline:
            return None

        page.wait_for_timeout(250)


def strip_school_parenthetical(school_name: str) -> str:
    return norm(re.sub(r"\s*\([^)]*\)\s*$", "", norm(school_name)))


def prior_year_school_candidates(wai_school: str) -> List[str]:
    candidates = []
    trimmed = strip_school_parenthetical(wai_school)
    for candidate in [trimmed, norm(wai_school)]:
        if candidate and candidate not in candidates:
            candidates.append(candidate)
    return candidates


def select_combobox_option_from_candidates(
    page,
    label_regex,
    option_candidates: List[str],
    field_name: str,
    timeout_ms: int = 8000,
    close_with_escape: bool = True,
) -> str:
    errors = []
    for option_text in option_candidates:
        try:
            select_combobox_option(
                page,
                label_regex,
                option_text,
                timeout_ms=timeout_ms,
                close_with_escape=close_with_escape,
            )
            return option_text
        except Exception as e:
            errors.append(f"{option_text!r}: {e}")

    raise RuntimeError(
        f"Could not select {field_name}. Tried: {option_candidates}. "
        f"Errors: {' | '.join(errors)}"
    )


def transfer_prior_year_student(
    page,
    cfg: Dict[str, Any],
    wai_school: str,
    save: bool = True,
    log_fn=None,
    run_id=None,
) -> Dict[str, str]:
    transfer_control = get_prior_year_transfer_control(page)
    if transfer_control is None:
        return {}

    transfer_to_project = norm(cfg.get("workability", {}).get("transfer_to_project"))
    if not transfer_to_project:
        raise KeyError("config.yaml workability.transfer_to_project is required for prior-year transfers.")

    def _l(msg: str):
        if log_fn and run_id:
            _log(log_fn, f"[{run_id}] {msg}")
        elif log_fn:
            _log(log_fn, msg)

    _l("Prior-year Transfer link found. Opening transfer modal...")
    transfer_control.click()

    expect(page.get_by_text(re.compile(r"^Transfer Student$", re.I))).to_be_visible(timeout=10000)

    _l(f"Selecting prior-year transfer project: {transfer_to_project}")
    select_combobox_option(
        page,
        re.compile(r"Transfer\s+to", re.I),
        transfer_to_project,
        timeout_ms=10000,
        close_with_escape=False,
    )

    school_candidates = prior_year_school_candidates(wai_school)
    _l(f"Selecting prior-year transfer school from candidates: {school_candidates}")
    selected_school = select_combobox_option_from_candidates(
        page,
        re.compile(r"School\s+of\s+Attendance", re.I),
        school_candidates,
        "prior-year transfer School of Attendance",
        timeout_ms=10000,
        close_with_escape=False,
    )

    if save:
        _l("Saving prior-year transfer...")
        click_save_robust(page, timeout_ms=15000, log_fn=log_fn, run_id=run_id)
        try:
            expect(page.get_by_text(re.compile(r"^Transfer Student$", re.I))).to_be_hidden(timeout=10000)
        except Exception:
            pass
    else:
        _l("Prior-year transfer modal filled. Save was intentionally skipped.")

    return {
        "transfer_to_project": transfer_to_project,
        "school": selected_school,
    }


def select_combobox_option(
    page,
    label_regex,
    option_text: str,
    timeout_ms: int = 8000,
    close_with_escape: bool = True,
):
    """
    Select an option from a Radix/shadcn combobox (robust + deterministic).
    - Verifies the option exists in the UI list before clicking (typo-proof).
    - Re-acquires the listbox each retry to avoid stale/incorrect portals.
    - Closes the portal at the end to prevent eating the next click (Save).
    """

    def get_active_listbox():
        # Radix portals: usually there’s exactly one visible listbox when open.
        lb = page.get_by_role("listbox")
        expect(lb).to_be_visible(timeout=timeout_ms)
        return lb

    combo = page.get_by_label(label_regex)
    expect(combo).to_be_visible(timeout=timeout_ms)
    expect(combo).to_be_enabled(timeout=timeout_ms)
    scroll_timeout_ms = min(timeout_ms, 3000)

    # Always clear any lingering portal first. Skip this inside modals because
    # Escape can close the dialog when no dropdown is open.
    if close_with_escape:
        close_transient_overlays(page)

    # Open dropdown (retry in case a prior overlay blocks click)
    last_err = None
    for attempt in range(1, 4):
        try:
            combo.click(timeout=3000)
            listbox = get_active_listbox()
            break
        except Exception as e:
            last_err = e
            if close_with_escape:
                page.keyboard.press("Escape")
                page.wait_for_timeout(150)
    else:
        if close_with_escape:
            close_transient_overlays(page)
        raise RuntimeError(f"Failed to open combobox after retries: {last_err}")

    # Pull option texts from the UI (source of truth)
    try:
        available = [s.strip() for s in listbox.get_by_role("option").all_inner_texts()]
        available = [s for s in available if s]
    except Exception:
        available = []

    target_norm = norm(option_text).casefold()
    available_norm = {norm(a).casefold() for a in available}

    # Deterministic mismatch guardrail
    if available and target_norm not in available_norm:
        if close_with_escape:
            close_transient_overlays(page)

        suggestion = suggest_closest_match(option_text, available)

        hint = f"\nDid you mean: {suggestion!r} ?" if suggestion else ""

        raise RuntimeError(
            f"UI option not found for field "
            f"{label_regex.pattern if hasattr(label_regex,'pattern') else label_regex!r}: "
            f"{option_text!r}\n"
            f"Available options ({len(available)}): {available}"
            f"{hint}"
        )
    
    # Find the option (case-insensitive exact match on the visible label)
    option_re = re.compile(rf"^{re.escape(option_text)}$", re.I)

    last_err = None
    for attempt in range(1, 4):
        try:
            # Re-acquire listbox each attempt (Radix can re-render portals)
            listbox = get_active_listbox()
            option = listbox.get_by_role("option", name=option_re).first

            expect(option).to_be_visible(timeout=timeout_ms)
            option.scroll_into_view_if_needed(timeout=scroll_timeout_ms)

            # Normal click first
            option.click(timeout=3000)
            break

        except Exception as e:
            last_err = e

            # Close & reopen to reset portal state
            if close_with_escape:
                close_transient_overlays(page)

                try:
                    combo.click(timeout=3000)
                    listbox = get_active_listbox()
                except Exception:
                    pass

            # Last attempt: sometimes Radix options get finicky; allow force click
            if attempt == 3:
                try:
                    option = page.get_by_role("option", name=option_re).first
                    option.scroll_into_view_if_needed(timeout=scroll_timeout_ms)
                    option.click(timeout=3000, force=True)
                    break
                except Exception as e2:
                    last_err = e2
    else:
        if close_with_escape:
            close_transient_overlays(page)
        raise RuntimeError(f"Failed to click combobox option '{option_text}' after retries: {last_err}")

    # Ensure portal closes so it doesn't eat next click (Save). Skip the extra
    # Escape inside modals because the selected option normally closes the
    # listbox, and Escape could close the dialog.
    if close_with_escape:
        close_transient_overlays(page)
    try:
        expect(page.get_by_role("listbox")).to_be_hidden(timeout=2000)
    except Exception:
        pass

    # Tiny settle (animations)
    page.wait_for_timeout(100)


def select_combobox_locator_option(
    page,
    combo,
    option_text: str,
    field_name: str,
    timeout_ms: int = 8000,
) -> None:
    close_transient_overlays(page)
    expect(combo).to_be_visible(timeout=timeout_ms)
    expect(combo).to_be_enabled(timeout=timeout_ms)

    try:
        current = norm(combo.inner_text(timeout=1000))
        if current.casefold() == norm(option_text).casefold():
            return
    except Exception:
        pass

    combo.click(timeout=3000)
    listbox = page.get_by_role("listbox")
    expect(listbox).to_be_visible(timeout=timeout_ms)

    option_re = re.compile(rf"^{re.escape(option_text)}$", re.I)
    option = listbox.get_by_role("option", name=option_re).first
    expect(option).to_be_visible(timeout=timeout_ms)
    option.scroll_into_view_if_needed(timeout=min(timeout_ms, 3000))
    option.click(timeout=3000)

    close_transient_overlays(page)
    try:
        expect(page.get_by_role("listbox")).to_be_hidden(timeout=2000)
    except Exception:
        pass

    try:
        current = norm(combo.inner_text(timeout=1000))
        if current.casefold() != norm(option_text).casefold():
            raise RuntimeError(f"{field_name} still shows {current!r} after selecting {option_text!r}")
    except Exception:
        pass


def select_radio_by_label_text(page, label_text: str) -> None:
    try:
        page.get_by_role("radio", name=re.compile(rf"^{re.escape(label_text)}$", re.I)).check(timeout=100)
        return
    except Exception:
        pass

    container = page.locator("div").filter(has_text=re.compile(rf"^{re.escape(label_text)}$", re.I)).first
    for sel in ["#radio-group-item", "[role='radio']", "button[role='radio']"]:
        try:
            container.locator(sel).first.click(timeout=3000)
            return
        except Exception:
            pass

    loose = page.locator(f"text={label_text}").first
    for sel in ["xpath=ancestor::*[1]//*[@role='radio']", "xpath=ancestor::*[2]//*[@role='radio']"]:
        try:
            loose.locator(sel).first.click(timeout=3000)
            return
        except Exception:
            pass

    loose.click(timeout=3000)


def select_gender(page, gender_ui: str) -> None:
    text_map = {"M": "Male", "F": "Female", "Non Binary": "Non-Binary"}
    target_text = text_map.get(gender_ui, gender_ui)

    try:
        page.get_by_role("radio", name=re.compile(rf"^{re.escape(target_text)}$", re.I)).check(timeout=100)
        return
    except Exception:
        pass

    container = page.locator("div").filter(has_text=re.compile(rf"^{re.escape(target_text)}$", re.I)).first
    for sel in ["#radio-group-item", "[role='radio']", "button[role='radio']"]:
        try:
            container.locator(sel).first.click(timeout=3000)
            return
        except Exception:
            pass

    container.click(timeout=3000)

def preflight_validate_mapping_against_dropdown(
    page,
    mapping: Dict[str, str],
    label_regex,
    label_name: str,
    timeout_ms: int = 8000,
):
    """
    Validates that all mapped UI values exist in the WAI dropdown options.
    Raises a clear error if any UI values are missing (typos, UI change, etc.).
    """
    ui_opts = get_combobox_options(page, label_regex, timeout_ms=timeout_ms)
    ui_norm = {norm(x).casefold() for x in ui_opts}

    bad = []
    for raw, ui_val in (mapping or {}).items():
        if norm(ui_val).casefold() not in ui_norm:
            bad.append((raw, ui_val))

    if bad:
        raise RuntimeError(
            f"CONFIG_MAPPING_ERROR: Some {label_name} mapping UI values do not exist in WAI dropdown.\n"
            f"Examples (raw -> ui_value): {bad[:10]}\n"
            f"UI {label_name} options ({len(ui_opts)}): {ui_opts}"
        )
    
def create_new_student(
    page,
    student: Student,
    cfg: Dict[str, Any],
    log_fn=None,
    run_id=None,
    on_form_expanded=None,
):
    page.get_by_role("button", name="New Student").click()
    expect(page.get_by_placeholder("First Name")).to_be_visible(timeout=10000)

    page.get_by_placeholder("First Name").fill(student.first_name)
    page.get_by_placeholder("Last Name").fill(student.last_name)
    page.get_by_placeholder("SSID").fill(student.ssid)
    page.get_by_placeholder("MM/DD/YYYY").fill(student.dob)

    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting gender: {student.gender_ui}")
    select_gender(page, student.gender_ui)

    # WAI Project is now pre-filled / read-only on new student form — skip it.
    # (codegen shows it just Tabs past it)
    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] WAI Project pre-filled (skipping dropdown).")

    # School of Attendance is now a Radix combobox (NOT a native <select>)
    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting school: {student.wai_school}")

    school_label_regex = re.compile(r"School of Attendance", re.I)
    
    # Wait for the school combobox to be ready
    school_combo = page.get_by_label(school_label_regex)
    expect(school_combo).to_be_visible(timeout=15000)
    expect(school_combo).to_be_enabled(timeout=15000)

    # Use the existing robust combobox selector
    select_combobox_option(page, school_label_regex, student.wai_school)

    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting record type: Baseline 2025-26")
    select_radio_by_label_text(page, "Baseline 2025-26")

    click_save_robust(
        page,
        timeout_ms=15000,
        log_fn=log_fn,
        run_id=run_id,
        toast_regex=re.compile(r"baseline.*successfully\s+added", re.I),
    )

    expect(page.get_by_label(re.compile(r"Grade", re.I))).to_be_visible(timeout=10000)

    if on_form_expanded:
        on_form_expanded(page)

    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting grade: {student.grade_ui}")
    select_combobox_option(page, re.compile(r"Grade", re.I), str(student.grade_ui))

    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting disability: {student.disability_ui}")
    select_combobox_option(page, re.compile(r"Disability", re.I), student.disability_ui)

    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting ethnicity: {student.ethnicity_ui}")
    select_radio_by_label_text(page, student.ethnicity_ui)

    if log_fn and run_id:
        _log(log_fn, f"[{run_id}] Selecting race: {student.race_ui}")
    select_combobox_option(page, re.compile(r"Race", re.I), student.race_ui)

    page.keyboard.press("Escape")
    page.wait_for_timeout(100)

    click_save_robust(
        page,
        timeout_ms=15000,
        log_fn=log_fn,
        run_id=run_id,
        toast_regex=re.compile(r"successfully\s+saved", re.I),
    )


def fill_text_input(
    page,
    patterns: List[re.Pattern],
    value: str,
    field_name: str,
    timeout_ms: int = 8000,
    selectors: List[str] | None = None,
) -> None:
    last_err = None
    for selector in selectors or []:
        try:
            locator = page.locator(selector).first
            expect(locator).to_be_visible(timeout=timeout_ms)
            locator.fill("")
            locator.fill(value)
            return
        except Exception as e:
            last_err = e

    for pattern in patterns:
        try:
            locator = page.get_by_label(pattern).first
            expect(locator).to_be_visible(timeout=timeout_ms)
            locator.fill("")
            locator.fill(value)
            return
        except Exception as e:
            last_err = e

        try:
            locator = page.get_by_placeholder(pattern).first
            expect(locator).to_be_visible(timeout=timeout_ms)
            locator.fill("")
            locator.fill(value)
            return
        except Exception as e:
            last_err = e

    raise RuntimeError(f"Could not locate editable field for {field_name}: {last_err}")


def read_text_input_value(
    page,
    selectors: List[str],
    field_name: str,
    timeout_ms: int = 5000,
) -> str:
    last_err = None
    for selector in selectors:
        try:
            matches = page.locator(selector)
            matches.first.wait_for(state="attached", timeout=timeout_ms)
            blank_visible_value = None

            for i in range(matches.count()):
                locator = matches.nth(i)
                try:
                    if not locator.is_visible():
                        continue

                    value = locator.input_value(timeout=timeout_ms).strip()
                    if not value:
                        attr_value = locator.get_attribute("value", timeout=timeout_ms)
                        value = (attr_value or "").strip()

                    if value:
                        return value

                    blank_visible_value = ""
                except Exception as e:
                    last_err = e

            if blank_visible_value is not None:
                return blank_visible_value
        except Exception as e:
            last_err = e

    raise RuntimeError(f"Could not read field value for {field_name}: {last_err}")


def checkbox_is_checked(locator) -> bool:
    try:
        return bool(locator.is_checked(timeout=1000))
    except Exception:
        return bool(locator.evaluate(
            """el => {
                if (typeof el.checked === 'boolean') return el.checked;
                const aria = el.getAttribute('aria-checked');
                if (aria !== null) return aria === 'true';
                const state = el.getAttribute('data-state') || el.closest('[data-state]')?.getAttribute('data-state');
                return state === 'checked';
            }"""
        ))


def open_addresses_tab(page, log_fn=None, run_id=None, timeout_ms: int = 10000) -> None:
    def _l(msg: str):
        if log_fn and run_id:
            _log(log_fn, f"[{run_id}] {msg}")
        elif log_fn:
            _log(log_fn, msg)

    close_transient_overlays(page)
    tab_name = re.compile(r"^Addresses$", re.I)
    candidates = [
        page.get_by_role("tab", name=tab_name).first,
        page.get_by_role("link", name=tab_name).first,
        page.get_by_role("button", name=tab_name).first,
        page.get_by_text("Addresses", exact=True).first,
    ]

    last_err = None
    for attempt in range(1, 4):
        for candidate in candidates:
            try:
                if candidate.count() == 0 or not candidate.is_visible():
                    continue
                _l(f"Opening Addresses tab (attempt {attempt}/3)...")
                candidate.scroll_into_view_if_needed(timeout=3000)
                candidate.click(timeout=5000)
                wait_for_search_loader_to_settle(page)
                page.locator("#studentAddress").first.wait_for(state="visible", timeout=timeout_ms)
                return
            except Exception as e:
                last_err = e
                close_transient_overlays(page)
                page.wait_for_timeout(250)

    raise RuntimeError(f"Could not open Addresses tab: {last_err}")


def overwrite_existing_student_address(
    page,
    student: Student,
    log_fn=None,
    run_id=None,
) -> Dict[str, str]:
    def _l(msg: str):
        if log_fn and run_id:
            _log(log_fn, f"[{run_id}] {msg}")
        elif log_fn:
            _log(log_fn, msg)

    _l("Reading Parent/Guardian Street Address before editing address fields.")
    parent_guardian_address = read_text_input_value(
        page,
        [
            "#parentGuardianAddress",
            "input[aria-describedby='parentGuardianAddress']",
            "input[name='parentGuardianAddress']",
        ],
        "Parent/Guardian Street Address",
    )
    _l(f"Parent/Guardian Street Address value length before edits: {len(parent_guardian_address)}.")

    field_specs = [
        (
            "street_address",
            student.street_address,
            "Street address",
            [
                re.compile(r"Street\s+Address", re.I),
                re.compile(r"Address(?:\s+Line)?\s*1", re.I),
                re.compile(r"^Address$", re.I),
            ],
            [
                "#studentAddress",
                "input[aria-describedby='studentAddress']",
                "input[placeholder='Street Address']",
            ],
        ),
        (
            "city",
            student.city,
            "City",
            [re.compile(r"^City$", re.I)],
            [
                "#studentCity",
                "input[aria-describedby='studentCity']",
                "input[placeholder='City']",
            ],
        ),
        (
            "zip_code",
            student.zip_code,
            "Zip",
            [re.compile(r"Zip(\s*Code)?", re.I), re.compile(r"Postal", re.I)],
            [],
        ),
        (
            "phone_number",
            student.phone_number,
            "Phone Number",
            [re.compile(r"Phone(\s*Number)?", re.I), re.compile(r"Primary\s+Phone", re.I)],
            [],
        ),
        (
            "parent_name",
            student.parent_name,
            "Parent Name",
            [
                re.compile(r"^Name\(s\)$", re.I),
                re.compile(r"Parent\s+Name", re.I),
                re.compile(r"Parent\s*/\s*Guardian", re.I),
                re.compile(r"Guardian\s+Name", re.I),
                re.compile(r"Contact\s+Name", re.I),
            ],
            [
                "#parentGuardianNames",
                "input[aria-describedby='parentGuardianNames']",
                "input[placeholder='Name(s)']",
            ],
        ),
    ]

    for _, value, field_name, patterns, selectors in field_specs:
        _l(f"Updating {field_name}: {value}")
        fill_text_input(page, patterns, value, field_name, selectors=selectors)

    if parent_guardian_address:
        _l("Parent/Guardian Street Address already has a value; skipping Same address as student.")
    else:
        _l("Parent/Guardian Street Address is blank; checking Same address as student.")
        same_address = page.locator("#sameAddress").first
        expect(same_address).to_be_visible(timeout=8000)
        if checkbox_is_checked(same_address):
            _l("Same address as student is already checked; leaving it unchanged.")
        else:
            same_address.check(timeout=3000)

    _l(f"Updating Parent/Guardian Phone: {student.phone_number}")
    fill_text_input(
        page,
        [re.compile(r"Phone(\s*Number)?", re.I)],
        student.phone_number,
        "Parent/Guardian Phone",
        selectors=[
            "#parentGuardianPhone",
            "input[aria-describedby='parentGuardianPhone']",
        ],
    )

    return {
        "street_address": student.street_address,
        "city": student.city,
        "state": student.state,
        "zip_code": student.zip_code,
        "phone_number": student.phone_number,
        "parent_name": student.parent_name,
    }


ARRAY_SERVICE_TARGETS = [
    (
        "Career / Vocational Assessments",
        re.compile(r"^Career\s*/\s*Vocational\s+Assessments$", re.I),
    ),
    (
        "Employment / Post-Secondary Education Planning",
        re.compile(r"^Employment\s*/\s*Post[-\s]*Secondary\s+Education\s+Planning$", re.I),
    ),
    (
        "Career Awareness / Exploration Activities",
        re.compile(r"^Career\s+Awareness\s*/\s*Exploration\s+Activities$", re.I),
    ),
    (
        "Career Preparation / Job Search",
        re.compile(r"^Career\s+Preparation\s*/\s*Job\s+Search$", re.I),
    ),
    (
        "Self-Advocacy / Disability Awareness",
        re.compile(r"^Self[-\s]*Advocacy\s*/\s*Disability\s+Awareness$", re.I),
    ),
    (
        "Youth Leadership",
        re.compile(r"^Youth\s+Leadership$", re.I),
    ),
]

SERVE_LAYOUT_CHANGED_PREFIX = "SERVE_LAYOUT_CHANGED"


def open_array_of_services(page) -> None:
    candidates = [
        page.get_by_role("link", name=re.compile(r"Array\s+of\s+Services", re.I)).first,
        page.get_by_role("tab", name=re.compile(r"Array\s+of\s+Services", re.I)).first,
        page.get_by_text(re.compile(r"^Array\s+of\s+Services$", re.I)).first,
    ]

    last_err = None
    for candidate in candidates:
        try:
            if candidate.count() > 0 and candidate.is_visible():
                candidate.click(timeout=5000)
                break
        except Exception as e:
            last_err = e
    else:
        raise RuntimeError(f"Could not find/click Array of Services tab: {last_err}")

    try:
        page.wait_for_load_state("domcontentloaded", timeout=10000)
    except Exception:
        pass

    expect(page.get_by_text(ARRAY_SERVICE_TARGETS[0][1]).first).to_be_visible(timeout=15000)
    expect(page.get_by_text(ARRAY_SERVICE_TARGETS[1][1]).first).to_be_visible(timeout=15000)


def serve_layout_changed(message: str) -> RuntimeError:
    return RuntimeError(
        f"{SERVE_LAYOUT_CHANGED_PREFIX}: {message} "
        "The WAI Array of Services page structure appears to have changed and the code needs to be updated."
    )


def _fill_array_service_row_by_geometry(page, row_label: str, hours: str) -> Dict[str, Any]:
    result = page.evaluate(
        """
        ({rowLabel, hours}) => {
            const norm = (s) => String(s || '').replace(/\\s+/g, ' ').trim().toLowerCase();
            const display = (s) => String(s || '').replace(/\\s+/g, ' ').trim();
            const needle = norm(rowLabel);
            const visible = (el) => {
                if (!el) return false;
                const style = window.getComputedStyle(el);
                const rect = el.getBoundingClientRect();
                return style.visibility !== 'hidden' && style.display !== 'none' && rect.width > 0 && rect.height > 0;
            };
            const textOf = (el) => norm(el.innerText || el.textContent || '');
            const displayTextOf = (el) => display(el.innerText || el.textContent || '');
            const area = (el) => {
                const rect = el.getBoundingClientRect();
                return rect.width * rect.height;
            };
            const exact = Array.from(document.querySelectorAll('body *'))
                .filter(el => visible(el) && textOf(el) === needle)
                .sort((a, b) => {
                    const ar = a.getBoundingClientRect();
                    const br = b.getBoundingClientRect();
                    return area(a) - area(b) || ar.left - br.left;
                });
            const label = exact[0];

            if (!label) {
                return {
                    ok: false,
                    reason: `Could not find exact Array of Services row label "${rowLabel}".`,
                };
            }

            const labelRect = label.getBoundingClientRect();
            const labelCenterY = labelRect.top + (labelRect.height / 2);

            const headerCandidates = Array.from(document.querySelectorAll('body *'))
                .filter(el => visible(el) && textOf(el) === 'wai')
                .map(el => {
                    const rect = el.getBoundingClientRect();
                    return {
                        el,
                        rect,
                        centerX: rect.left + (rect.width / 2),
                        centerY: rect.top + (rect.height / 2),
                        text: displayTextOf(el),
                    };
                })
                .filter(item => item.centerY < labelCenterY)
                .sort((a, b) => {
                    const aDist = labelCenterY - a.centerY;
                    const bDist = labelCenterY - b.centerY;
                    return aDist - bDist || area(a.el) - area(b.el);
                });

            const header = headerCandidates[0];
            if (!header) {
                return {
                    ok: false,
                    reason: `Could not find exact "WAI" column header above Array of Services row "${rowLabel}".`,
                };
            }

            const inputs = Array.from(document.querySelectorAll('input'))
                .filter(input => {
                    if (!visible(input)) return false;
                    if (String(input.type || '').toLowerCase() === 'checkbox') return false;
                    const rect = input.getBoundingClientRect();
                    const centerY = rect.top + (rect.height / 2);
                    return Math.abs(centerY - labelCenterY) <= 70;
                })
                .sort((a, b) => a.getBoundingClientRect().left - b.getBoundingClientRect().left);

            if (!inputs.length) {
                return {
                    ok: false,
                    reason: `Could not find hour inputs aligned with Array of Services row "${rowLabel}".`,
                };
            }

            const scored = inputs
                .map((input, index) => {
                    const rect = input.getBoundingClientRect();
                    const centerX = rect.left + (rect.width / 2);
                    return {
                        input,
                        index,
                        rect,
                        distance: Math.abs(centerX - header.centerX),
                    };
                })
                .sort((a, b) => a.distance - b.distance);

            const selected = scored[0];
            const field = selected && selected.input;
            if (!field) {
                return {
                    ok: false,
                    reason: `Could not locate WAI hours input for Array of Services row "${rowLabel}".`,
                };
            }

            if (selected.distance > 90) {
                return {
                    ok: false,
                    reason: `Closest input for row "${rowLabel}" is too far from the WAI column header (${Math.round(selected.distance)} px).`,
                };
            }

            field.scrollIntoView({ block: 'center', inline: 'center' });
            const setValue = Object.getOwnPropertyDescriptor(window.HTMLInputElement.prototype, 'value').set;
            const originalValue = display(field.value);
            const numericOriginal = Number(originalValue);
            const shouldFill = originalValue !== '' && Number.isFinite(numericOriginal) && numericOriginal === 0;

            if (shouldFill) {
                field.focus();
                setValue.call(field, '');
                field.dispatchEvent(new Event('input', { bubbles: true }));
                setValue.call(field, hours);
                field.dispatchEvent(new Event('input', { bubbles: true }));
                field.dispatchEvent(new Event('change', { bubbles: true }));
                field.blur();
            }

            const fieldRect = field.getBoundingClientRect();
            return {
                ok: true,
                value: field.value,
                originalValue,
                action: shouldFill ? 'filled' : 'left_existing',
                inputCount: inputs.length,
                selectedIndex: selected.index,
                rowLabel,
                labelText: displayTextOf(label),
                columnHeader: header.text,
                columnHeaderLeft: Math.round(header.rect.left),
                columnHeaderTop: Math.round(header.rect.top),
                distanceFromColumnHeader: Math.round(selected.distance),
                fieldLeft: Math.round(fieldRect.left),
                fieldTop: Math.round(fieldRect.top),
            };
        }
        """,
        {"rowLabel": row_label, "hours": hours},
    )

    if not result or not result.get("ok"):
        raise serve_layout_changed((result or {}).get("reason") or f"Could not resolve row {row_label!r}.")

    return result


def fill_array_service_wai_hours(page, row_label: str, label_regex: re.Pattern, hours: str = "0.5") -> Dict[str, Any]:
    expect(page.get_by_text(label_regex).first).to_be_visible(timeout=8000)
    result = _fill_array_service_row_by_geometry(page, row_label, hours)
    value = norm((result or {}).get("value"))
    action = (result or {}).get("action")
    if action == "filled" and value != hours:
        raise RuntimeError(
            f"Array of Services WAI hours for {row_label!r} expected {hours!r} after fill, got {value!r}; "
            f"locator result={result!r}"
        )
    if not value:
        raise RuntimeError(
            f"Array of Services WAI hours for {row_label!r} is blank after evaluation; locator result={result!r}"
        )
    page.wait_for_timeout(150)
    return result


def fill_array_of_services_dry_run(page, hours: str = "0.5", log_fn=None, run_id=None) -> List[Dict[str, Any]]:
    def _l(msg: str) -> None:
        if log_fn and run_id:
            _log(log_fn, f"[{run_id}] {msg}")
        elif log_fn:
            _log(log_fn, msg)

    open_array_of_services(page)
    results: List[Dict[str, Any]] = []
    for row_label, label_regex in ARRAY_SERVICE_TARGETS:
        _l(f"Setting Array of Services WAI hours: {row_label} = {hours}")
        result = fill_array_service_wai_hours(page, row_label, label_regex, hours=hours)
        results.append(result)
        if result.get("action") == "filled":
            _l(
                f"Array of Services filled: {row_label}; "
                f"{result.get('originalValue')!r} -> {result.get('value')!r}; "
                f"column={result.get('columnHeader')!r}; distance={result.get('distanceFromColumnHeader')}px"
            )
        else:
            _l(
                f"Array of Services left unchanged: {row_label}; "
                f"existing value={result.get('value')!r}; "
                f"column={result.get('columnHeader')!r}; distance={result.get('distanceFromColumnHeader')}px"
            )

    missing = [label for label, _ in ARRAY_SERVICE_TARGETS if label not in {r.get("rowLabel") for r in results}]
    if missing:
        raise RuntimeError(f"Array of Services verification missing rows: {missing}")

    _l("Array of Services verification before save:")
    for result in results:
        _l(
            f"  {result.get('rowLabel')}: WAI={result.get('value')!r} "
            f"({result.get('action')}, was {result.get('originalValue')!r})"
        )

    page.evaluate("window.scrollTo(0, document.body.scrollHeight)")
    page.wait_for_timeout(500)
    return results


# =========================
# Runner
# =========================

def run(
    excel_path: str,
    config_path: str = "config/config.yaml",
    username: str | None = None,
    password: str | None = None,
    log_fn=None,
    override_headless: bool | None = None,
    stop_event=None,
    user_log_fn=None,
) -> None:

    cfg = load_config(config_path)

    did_preflight = False

    def on_form_expanded(page):
        nonlocal did_preflight
        if did_preflight:
            return
        if not cfg.get("ui", {}).get("preflight_validate", True):
            did_preflight = True
            return

        # Validate Race mapping values against the actual WAI Race dropdown options
        preflight_validate_mapping_against_dropdown(
            page,
            cfg.get("race_mapping", {}),
            re.compile(r"Race", re.I),
            "Race",
            timeout_ms=8000,
        )
        did_preflight = True
        
    # --- UI hooks (closures that see run() parameters safely) ---
    ui_hold_default = float(cfg.get("ui", {}).get("min_status_seconds", 1.5))

    def _user(msg: str, min_hold_sec: float = 0.0):
        if user_log_fn:
            user_log_fn(msg)
        else:
            _log(log_fn, msg)
        hold = float(min_hold_sec or 0.0)
        if hold > 0:
            time.sleep(hold)

    def _should_stop() -> bool:
        return bool(stop_event and stop_event.is_set())

    if "search_project" not in cfg["workability"] or "create_project" not in cfg["workability"]:
        raise KeyError("config.yaml workability must include search_project and create_project.")

    ledger_path = cfg["ledger"]["path"]
    ensure_dir("output/screenshots")
    ensure_dir("output/traces")
    ensure_dir(os.path.dirname(ledger_path))

    run_id = str(uuid.uuid4())[:8]

    username_default = cfg.get("credentials", {}).get("username", "")
    if username is None or not str(username).strip():
        username = username_default or input("WAI Username (email): ").strip()

    if password is None:
        password = getpass("WAI Password: ")

    _log(log_fn, f"[{run_id}] Starting run. Excel: {excel_path}")
    _user("🚀 Starting Workabili-Bot3000 run…", 0.5)

    sanitization_log_path = os.path.join("output", "logs", "input_sanitization.log")
    _sanitization_log = make_input_audit_logger(sanitization_log_path, run_id, log_fn=log_fn)

    rows = read_excel_rows(excel_path, cfg=cfg, log_fn=_sanitization_log)
    update_addresses = rows_have_address_columns(rows, cfg)
    if rows_have_partial_address_columns(rows, cfg):
        raise ValueError(
            "INPUT_HEADER_ERROR: Create can update addresses only when all address headers are present: "
            f"{expected_address_headers(cfg)!r}"
        )

    # -------------------------
    # Run summary stats (for UI)
    # -------------------------
    stats = {
        "input_rows": len(rows),
        "created": 0,
        "transfer_requested": 0,
        "transfer_pending": 0,
        "transferred_prior_year": 0,
        "already_owned": 0,
        "address_patched": 0,
        "skipped": 0,
        "errors": 0,
        "stopped": 0,
    }

    def _inc(key: str, n: int = 1):
        stats[key] = int(stats.get(key, 0)) + n



    if not rows:
        _user("⚠️ No rows found in Excel.", 0.5)
        return {
            "run_id": run_id,
            "ledger_path": ledger_path,
            **stats,
        }

    prepared_rows, invalid_rows = prepare_input_rows(
        rows,
        cfg,
        validation_log_fn=_sanitization_log,
        require_address_fields=update_addresses,
    )

    for invalid in invalid_rows:
        _log(
            log_fn,
            f"[{run_id}] {invalid.action}: row {invalid.excel_row_num} "
            f"{invalid.ssid or '(no SSID)'} ({invalid.display_name}) -> {invalid.details}",
        )
        _inc("skipped")
        _user(f"{invalid.action}||display_name={invalid.display_name}", ui_hold_default)
        append_ledger_xlsx(ledger_path, {
            "run_id": run_id,
            "timestamp": datetime.utcnow().isoformat(),
            "ssid": invalid.ssid,
            "student_name": invalid.display_name,
            "action": invalid.action,
            "details": invalid.details,
            "screenshot": "",
            "trace": "",
        })

    if not prepared_rows:
        _user("⚠️ No valid rows found in Excel.", 0.5)
        _log(log_fn, f"[{run_id}] No valid rows found after validation. Ledger: {ledger_path}")
        return {
            "run_id": run_id,
            "ledger_path": ledger_path,
            **stats,
        }

    total = len(prepared_rows)

    resume = bool(cfg.get("run", {}).get("resume", True))
    completed: Set[str] = set()
    last_action_by_ssid: Dict[str, str] = {}
    transfer_requested_ssids: Set[str] = set()

    if resume:
        completed, last_action_by_ssid, transfer_requested_ssids = load_ledger_state_xlsx(ledger_path)

    headless = bool(cfg.get("run", {}).get("headless", False))
    if override_headless is not None:
        headless = bool(override_headless)

    slow_mo_ms = int(cfg.get("run", {}).get("slow_mo_ms", 0))
    trace_on_failure = bool(cfg.get("trace", {}).get("on_failure", True))
    stop_on_error = bool(cfg.get("run", {}).get("stop_on_error", False))

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms if slow_mo_ms > 0 else None)
        context = browser.new_context(viewport={"width": 1600, "height": 1400})
        page = context.new_page()

        try:
            login(page, username, password, cfg, log_fn=log_fn)
        except Exception as e:
            _log(log_fn, f"[{run_id}] FATAL LOGIN ERROR -> {e}")

            # Ensure browser shuts down so UI doesn't hang
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass

            # IMPORTANT: bubble error up to GUI so it shows the BIG modal + shake
            raise

        if _should_stop():
            _user("🛑 Stopped by user.", 0.5)
            context.close()
            browser.close()
            return

        _log(log_fn, f"[{run_id}] Logged in. Opening Student Records...")
        _user("📚 Logged in. Opening Student Records…", 0.5)
        goto_student_records(page, cfg=cfg, log_fn=log_fn, run_id=run_id)

        for i, prepared in enumerate(prepared_rows):
            if _should_stop():
                _user("🛑 Stopped by user.", 0.5)
                _inc("stopped")
                break

            student = prepared.student
            display_name = prepared.display_name

            _user(f"🔎 {i+1} of {total}: Checking {display_name}…", 0.5)

            if resume and student.ssid in completed:
                last_action = last_action_by_ssid.get(student.ssid, "UNKNOWN")
                _log(log_fn, f"[{run_id}] SKIPPED_RESUME: {student.ssid} ({display_name}) (already processed: {last_action})")
                _user(f"SKIPPED_RESUME||display_name={display_name}||last_action={last_action}", ui_hold_default)
                _inc("skipped")
                append_ledger_xlsx(ledger_path, {
                    "run_id": run_id,
                    "timestamp": datetime.utcnow().isoformat(),
                    "ssid": student.ssid,
                    "student_name": display_name,
                    "action": "SKIPPED_RESUME",
                    "details": f"Skipped (already processed in prior ledger: {last_action})",
                    "screenshot": "",
                    "trace": "",
                })
                continue

            trace_path = ""
            screenshot_path = ""

            try:
                if trace_on_failure:
                    context.tracing.start(screenshots=True, snapshots=True, sources=True)

                _log(log_fn, f"[{run_id}] Processing SSID {student.ssid} ({display_name})")

                if _should_stop():
                    _user("🛑 Stopped by user.", 0.5)
                    break

                goto_student_records(page, cfg=cfg, log_fn=log_fn, run_id=run_id)

                _log(log_fn, f"[{run_id}] Searching SSID {student.ssid} under '{cfg['workability']['search_project']}'...")
                search_by_ssid(page, student.ssid, cfg)

                outcome = determine_search_outcome(page, timeout_ms=3000)

                if outcome == "FOUND":
                    _log(log_fn, f"[{run_id}] FOUND (Edit visible). Checking owning org BEFORE Edit...")

                    owning_org = get_found_row_owning_org(page, student.ssid)
                    _log(log_fn, f"[{run_id}] Owning org for {student.ssid}: {owning_org!r}")

                    _log(log_fn, f"[{run_id}] Opening found student record to check for prior-year transfer path...")
                    open_existing_student_edit(page)

                    prior_year_transfer = transfer_prior_year_student(
                        page,
                        cfg,
                        student.wai_school,
                        save=True,
                        log_fn=log_fn,
                        run_id=run_id,
                    )
                    if prior_year_transfer:
                        details = (
                            "Transferred prior-year record. "
                            f"Transfer to={prior_year_transfer['transfer_to_project']}; "
                            f"School={prior_year_transfer['school']}"
                        )
                        _log(log_fn, f"[{run_id}] TRANSFERRED_PRIOR_YEAR: {student.ssid} ({display_name}) -> {details}")
                        _user(f"TRANSFERRED_PRIOR_YEAR||display_name={display_name}", ui_hold_default)
                        _inc("transferred_prior_year")
                        append_ledger_xlsx(ledger_path, {
                            "run_id": run_id,
                            "timestamp": datetime.utcnow().isoformat(),
                            "ssid": student.ssid,
                            "student_name": display_name,
                            "action": "TRANSFERRED_PRIOR_YEAR",
                            "details": details,
                            "screenshot": "",
                            "trace": "",
                        })
                        completed.add(student.ssid)
                        if trace_on_failure:
                            context.tracing.stop()
                        continue

                    if is_already_owned_by_us(owning_org, cfg):
                        _log(log_fn, f"[{run_id}] ALREADY_OWNED: {student.ssid} ({display_name}) -> owned by us, skipping.")
                        _user(f"ALREADY_OWNED||display_name={display_name}", ui_hold_default)

                        screenshot_path = os.path.join("output", "screenshots", f"already_owned_{run_id}_{student.ssid}.png")
                        try:
                            page.screenshot(path=screenshot_path, full_page=False)
                        except Exception:
                            screenshot_path = ""
                        _inc("already_owned")
                        append_ledger_xlsx(ledger_path, {
                            "run_id": run_id,
                            "timestamp": datetime.utcnow().isoformat(),
                            "ssid": student.ssid,
                            "student_name": display_name,
                            "action": "ALREADY_OWNED",
                            "details": f"Student already enrolled in WAI under our org: {owning_org}",
                            "screenshot": screenshot_path,   # ✅ fixed
                            "trace": "",
                        })
                        completed.add(student.ssid)
                        if trace_on_failure:
                            context.tracing.stop()
                        continue

                    _log(log_fn, f"[{run_id}] FOUND (Edit visible, not SBCSS Student). Checking for envelope/transfer status...")
                    _user("📨 Student exists in another org. Checking transfer…", 0.5)

                    action = request_transfer_if_possible(page)

                    details = ""
                    screenshot_path = ""

                    if action == "TRANSFER_REQUESTED":
                        details = "Requested transfer via envelope (Yes)."
                        _user(f"TRANSFER_REQUESTED||display_name={display_name}", ui_hold_default)
                        _inc("transfer_requested")
                    else:  # TRANSFER_PENDING
                        details = "Transfer pending (previously requested); waiting for release."
                        _user(f"TRANSFER_PENDING||display_name={display_name}", ui_hold_default)
                        _inc("transfer_pending")
                        try:
                            screenshot_path = save_transfer_pending_screenshot(page, run_id, student.ssid)
                        except Exception as e:
                            _log(log_fn, f"[{run_id}] WARNING: Failed to save transfer-pending screenshot: {e}")

                    _log(log_fn, f"[{run_id}] {action}: {student.ssid} ({display_name})")

                    append_ledger_xlsx(ledger_path, {
                        "run_id": run_id,
                        "timestamp": datetime.utcnow().isoformat(),
                        "ssid": student.ssid,
                        "student_name": display_name,
                        "action": action,
                        "details": details,
                        "screenshot": screenshot_path,
                        "trace": "",
                    })

                    # You asked to leave “transfers are a completed state” behavior as-is:
                    completed.add(student.ssid)

                else:
                    _log(log_fn, f"[{run_id}] NOT FOUND (0 records). Creating under '{cfg['workability']['create_project']}'...")
                    _user("📝 Not found in WAI. Entering now…", 1.0)

                    create_new_student(
                        page,
                        student,
                        cfg,
                        log_fn=log_fn,
                        run_id=run_id,
                        on_form_expanded=on_form_expanded,
                    )

                    _log(log_fn, f"[{run_id}] CREATED: {student.ssid} ({display_name})")
                    _user(f"CREATED||display_name={display_name}", 2.0)
                    _inc("created")
                    append_ledger_xlsx(ledger_path, {
                        "run_id": run_id,
                        "timestamp": datetime.utcnow().isoformat(),
                        "ssid": student.ssid,
                        "student_name": display_name,
                        "action": "CREATED",
                        "details": f"Created. Grade={student.grade_ui}; Disability={student.disability_ui}; Race={student.race_ui}; School={student.wai_school}",
                        "screenshot": "",
                        "trace": "",
                    })
                    completed.add(student.ssid)

                    if update_addresses:
                        _log(log_fn, f"[{run_id}] CREATED opening Addresses tab for {student.ssid} ({display_name})")
                        open_addresses_tab(page, log_fn=log_fn, run_id=run_id)
                        patch_results = overwrite_existing_student_address(page, student, log_fn=log_fn, run_id=run_id)

                        _log(log_fn, f"[{run_id}] CREATED saving address/contact fields for {student.ssid} ({display_name})")
                        click_save_robust(
                            page,
                            timeout_ms=15000,
                            log_fn=log_fn,
                            run_id=run_id,
                            toast_regex=re.compile(r"(successfully\s+saved|address\s+saved\s+successfully)", re.I),
                        )
                        page.wait_for_timeout(1000)

                        address_screenshot_path = os.path.join(
                            "output",
                            "screenshots",
                            f"address_patched_{run_id}_{student.ssid}.png",
                        )
                        try:
                            page.screenshot(path=address_screenshot_path, full_page=False)
                        except Exception:
                            address_screenshot_path = ""

                        address_details = (
                            f"Street={patch_results['street_address']}; City={patch_results['city']}; "
                            f"State={patch_results['state']}; Zip={patch_results['zip_code']}; "
                            f"Phone={patch_results['phone_number']}; Parent={patch_results['parent_name']}"
                        )
                        _inc("address_patched")
                        _log(log_fn, f"[{run_id}] ADDRESS_PATCHED: {student.ssid} ({display_name}) -> {address_details}")
                        _user(f"ADDRESS_PATCHED||display_name={display_name}", ui_hold_default)
                        append_ledger_xlsx(ledger_path, {
                            "run_id": run_id,
                            "timestamp": datetime.utcnow().isoformat(),
                            "ssid": student.ssid,
                            "student_name": display_name,
                            "action": "ADDRESS_PATCHED",
                            "details": address_details,
                            "screenshot": address_screenshot_path,
                            "trace": "",
                        })

                if trace_on_failure:
                    context.tracing.stop()

            except Exception as exc:
                if trace_on_failure:
                    trace_path = os.path.join("output", "traces", f"trace_{run_id}_{student.ssid}.zip")
                    try:
                        context.tracing.stop(path=trace_path)
                    except Exception:
                        trace_path = ""

                try:
                    screenshot_path = os.path.join("output", "screenshots", f"error_{run_id}_{student.ssid}.png")
                    page.screenshot(path=screenshot_path, full_page=True)
                except Exception:
                    screenshot_path = ""

                try:
                    close_transient_overlays(page)
                except Exception:
                    pass

                _log(log_fn, f"[{run_id}] ERROR: {student.ssid} ({display_name}) -> {exc}")
                _user(f"ERROR||display_name={display_name}", ui_hold_default)
                _inc("errors")
                append_ledger_xlsx(ledger_path, {
                    "run_id": run_id,
                    "timestamp": datetime.utcnow().isoformat(),
                    "ssid": student.ssid,
                    "student_name": display_name,
                    "action": "ERROR",
                    "details": str(exc),
                    "screenshot": screenshot_path,
                    "trace": trace_path,
                })

                if stop_on_error:
                    raise
                continue

        _log(log_fn, f"[{run_id}] Run complete. Ledger: {ledger_path}")
        _user("✅ Run complete. Check the ledger for details.", 0.5)

        context.close()
        browser.close()

        print(f"Run complete. Ledger: {ledger_path}")

        return {
            "run_id": run_id,
            "ledger_path": ledger_path,
            **stats,
        }


def run_serve(
    excel_path: str,
    config_path: str = "config/config.yaml",
    username: str | None = None,
    password: str | None = None,
    log_fn=None,
    override_headless: bool | None = None,
    override_slow_mo_ms: int | None = None,
    stop_event=None,
    user_log_fn=None,
    pause_after_first_fill: bool = False,
    pause_fn=None,
    save_after_fill: bool = False,
    hours: str = "0.5",
) -> Dict[str, Any]:
    cfg = load_config(config_path)

    ui_hold_default = float(cfg.get("ui", {}).get("min_status_seconds", 1.5))

    def _user(msg: str, min_hold_sec: float = 0.0):
        if user_log_fn:
            user_log_fn(msg)
        else:
            _log(log_fn, msg)
        hold = float(min_hold_sec or 0.0)
        if hold > 0:
            time.sleep(hold)

    def _should_stop() -> bool:
        return bool(stop_event and stop_event.is_set())

    if "search_project" not in cfg["workability"]:
        raise KeyError("config.yaml workability must include search_project.")

    ledger_path = cfg["ledger"]["path"]
    ensure_dir("output/screenshots")
    ensure_dir("output/traces")
    ensure_dir(os.path.dirname(ledger_path))

    run_id = str(uuid.uuid4())[:8]

    username_default = cfg.get("credentials", {}).get("username", "")
    if username is None or not str(username).strip():
        username = username_default or input("WAI Username (email): ").strip()

    if password is None:
        password = getpass("WAI Password: ")

    if save_after_fill:
        serve_mode_label = "Marking serves"
        serve_start_message = "Starting Workabili-Bot3000: Marking serves..."
        serve_complete_message = "Marking serves completed.  Check the ledger for details."
        serve_console_complete = f"Marking serves completed. Ledger: {ledger_path}"
    else:
        serve_mode_label = "Serve dry-run"
        serve_start_message = "Starting Workabili-Bot3000 Serve dry-run..."
        serve_complete_message = "Serve dry-run complete. Check the ledger for details."
        serve_console_complete = f"Serve dry-run complete. Ledger: {ledger_path}"
    _log(log_fn, f"[{run_id}] Starting {serve_mode_label}. Excel: {excel_path}")
    _user(serve_start_message, 0.5)

    sanitization_log_path = os.path.join("output", "logs", "input_sanitization.log")
    _sanitization_log = make_input_audit_logger(sanitization_log_path, run_id, log_fn=log_fn)

    rows = read_excel_rows(excel_path, cfg=cfg, log_fn=_sanitization_log)
    # Address filling in serve mode is disabled by request.
    # update_addresses = rows_have_address_columns(rows, cfg)
    # if rows_have_partial_address_columns(rows, cfg):
    #     raise ValueError(
    #         "INPUT_HEADER_ERROR: Serve can update addresses only when all address headers are present: "
    #         f"{expected_address_headers(cfg)!r}"
    #     )
    update_addresses = False

    stats = {
        "input_rows": len(rows),
        "serve_dry_run_filled": 0,
        "serve_saved": 0,
        "serve_skipped_not_owned": 0,
        "serve_not_found": 0,
        "address_dry_run_filled": 0,
        "address_patched": 0,
        "address_not_found": 0,
        "address_skipped_not_owned": 0,
        "created": 0,
        "transfer_requested": 0,
        "transfer_pending": 0,
        "transferred_prior_year": 0,
        "already_owned": 0,
        "skipped": 0,
        "errors": 0,
        "stopped": 0,
    }

    def _inc(key: str, n: int = 1):
        stats[key] = int(stats.get(key, 0)) + n

    if not rows:
        _user("No rows found in Excel.", 0.5)
        return {"run_id": run_id, "ledger_path": ledger_path, **stats}

    prepared_rows, invalid_rows = prepare_input_rows(
        rows,
        cfg,
        validation_log_fn=_sanitization_log,
        require_address_fields=update_addresses,
    )

    for invalid in invalid_rows:
        _log(
            log_fn,
            f"[{run_id}] {invalid.action}: row {invalid.excel_row_num} "
            f"{invalid.ssid or '(no SSID)'} ({invalid.display_name}) -> {invalid.details}",
        )
        _inc("skipped")
        _user(f"{invalid.action}||display_name={invalid.display_name}", ui_hold_default)
        append_ledger_xlsx(ledger_path, {
            "run_id": run_id,
            "timestamp": datetime.utcnow().isoformat(),
            "ssid": invalid.ssid,
            "student_name": invalid.display_name,
            "action": invalid.action,
            "details": invalid.details,
            "screenshot": "",
            "trace": "",
        })

    if not prepared_rows:
        _user("No valid rows found in Excel.", 0.5)
        _log(log_fn, f"[{run_id}] No valid rows found after validation. Ledger: {ledger_path}")
        return {"run_id": run_id, "ledger_path": ledger_path, **stats}

    total = len(prepared_rows)

    headless = bool(cfg.get("run", {}).get("headless", False))
    if override_headless is not None:
        headless = bool(override_headless)

    slow_mo_ms = int(cfg.get("run", {}).get("slow_mo_ms", 0))
    if override_slow_mo_ms is not None:
        slow_mo_ms = int(override_slow_mo_ms)
    trace_on_failure = bool(cfg.get("trace", {}).get("on_failure", True))
    stop_on_error = bool(cfg.get("run", {}).get("stop_on_error", False))

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms if slow_mo_ms > 0 else None)
        context = browser.new_context(viewport={"width": 1600, "height": 1400})
        page = context.new_page()

        try:
            login(page, username, password, cfg, log_fn=log_fn)
        except Exception as e:
            _log(log_fn, f"[{run_id}] FATAL LOGIN ERROR -> {e}")
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass
            raise

        if _should_stop():
            _user("Stopped by user.", 0.5)
            context.close()
            browser.close()
            return {"run_id": run_id, "ledger_path": ledger_path, **stats}

        _log(log_fn, f"[{run_id}] Logged in. Opening Student Records for {serve_mode_label}...")
        _user("Logged in. Opening Student Records...", 0.5)
        goto_student_records(page, cfg=cfg, log_fn=log_fn, run_id=run_id)

        for i, prepared in enumerate(prepared_rows):
            if _should_stop():
                _user("Stopped by user.", 0.5)
                _inc("stopped")
                break

            student = prepared.student
            display_name = prepared.display_name
            trace_path = ""
            screenshot_path = ""

            _user(f"🔎 {i+1} of {total}: Checking {display_name}...", 0.5)

            try:
                if trace_on_failure:
                    context.tracing.start(screenshots=True, snapshots=True, sources=True)

                _log(log_fn, f"[{run_id}] SERVE processing SSID {student.ssid} ({display_name})")

                goto_student_records(page, cfg=cfg, log_fn=log_fn, run_id=run_id)
                _log(log_fn, f"[{run_id}] SERVE searching SSID {student.ssid} under '{cfg['workability']['search_project']}'...")
                search_by_ssid(page, student.ssid, cfg)

                outcome = determine_search_outcome(page, timeout_ms=5000)
                if outcome != "FOUND":
                    details = "Student not found in WAI; Serve dry-run skipped."
                    _log(log_fn, f"[{run_id}] SERVE_NOT_FOUND: {student.ssid} ({display_name})")
                    _user(f"SERVE_NOT_FOUND||display_name={display_name}", ui_hold_default)
                    _inc("serve_not_found")
                    append_ledger_xlsx(ledger_path, {
                        "run_id": run_id,
                        "timestamp": datetime.utcnow().isoformat(),
                        "ssid": student.ssid,
                        "student_name": display_name,
                        "action": "SERVE_NOT_FOUND",
                        "details": details,
                        "screenshot": "",
                        "trace": "",
                    })
                    if trace_on_failure:
                        context.tracing.stop()
                    continue

                owning_org = get_found_row_owning_org(page, student.ssid)
                _log(log_fn, f"[{run_id}] SERVE owning org for {student.ssid}: {owning_org!r}")

                if not is_already_owned_by_us(owning_org, cfg):
                    details = f"Student found in WAI but not owned by configured org; owning org={owning_org!r}."
                    _log(log_fn, f"[{run_id}] SERVE_SKIPPED_NOT_OWNED: {student.ssid} ({display_name}) -> {details}")
                    _user(f"SERVE_SKIPPED_NOT_OWNED||display_name={display_name}", ui_hold_default)
                    _inc("serve_skipped_not_owned")
                    append_ledger_xlsx(ledger_path, {
                        "run_id": run_id,
                        "timestamp": datetime.utcnow().isoformat(),
                        "ssid": student.ssid,
                        "student_name": display_name,
                        "action": "SERVE_SKIPPED_NOT_OWNED",
                        "details": details,
                        "screenshot": "",
                        "trace": "",
                    })
                    if trace_on_failure:
                        context.tracing.stop()
                    continue

                _log(log_fn, f"[{run_id}] SERVE opening Edit for owned student {student.ssid} ({display_name})")
                open_existing_student_edit(page)

                # Address filling in serve mode is disabled by request.
                # if update_addresses:
                #     open_addresses_tab(page, log_fn=log_fn, run_id=run_id)
                #     patch_results = overwrite_existing_student_address(page, student, log_fn=log_fn, run_id=run_id)
                #
                #     address_action = "ADDRESS_DRY_RUN_FILLED"
                #     if save_after_fill:
                #         _log(log_fn, f"[{run_id}] SERVE saving address/contact fields for {student.ssid} ({display_name})")
                #         click_save_robust(
                #             page,
                #             timeout_ms=15000,
                #             log_fn=log_fn,
                #             run_id=run_id,
                #             toast_regex=re.compile(r"(successfully\s+saved|address\s+saved\s+successfully)", re.I),
                #         )
                #         page.wait_for_timeout(1000)
                #         address_action = "ADDRESS_PATCHED"
                #
                #     address_screenshot_prefix = "address_patched" if save_after_fill else "address_dry_run"
                #     address_screenshot_path = os.path.join(
                #         "output",
                #         "screenshots",
                #         f"{address_screenshot_prefix}_{run_id}_{student.ssid}.png",
                #     )
                #     try:
                #         page.screenshot(path=address_screenshot_path, full_page=False)
                #     except Exception:
                #         address_screenshot_path = ""
                #
                #     address_details = (
                #         f"Street={patch_results['street_address']}; City={patch_results['city']}; "
                #         f"State={patch_results['state']}; Zip={patch_results['zip_code']}; "
                #         f"Phone={patch_results['phone_number']}; Parent={patch_results['parent_name']}"
                #     )
                #     if save_after_fill:
                #         _inc("address_patched")
                #     else:
                #         _inc("address_dry_run_filled")
                #
                #     _log(log_fn, f"[{run_id}] {address_action}: {student.ssid} ({display_name}) -> {address_details}")
                #     _user(f"{address_action}||display_name={display_name}", ui_hold_default)
                #     append_ledger_xlsx(ledger_path, {
                #         "run_id": run_id,
                #         "timestamp": datetime.utcnow().isoformat(),
                #         "ssid": student.ssid,
                #         "student_name": display_name,
                #         "action": address_action,
                #         "details": address_details,
                #         "screenshot": address_screenshot_path,
                #         "trace": "",
                #     })

                service_results = fill_array_of_services_dry_run(page, hours=hours, log_fn=log_fn, run_id=run_id)

                action = "SERVE_DRY_RUN_FILLED"
                if save_after_fill:
                    _log(log_fn, f"[{run_id}] SERVE saving Array of Services for {student.ssid} ({display_name})")
                    click_save_robust(page, timeout_ms=15000, log_fn=log_fn, run_id=run_id)
                    page.wait_for_timeout(1000)
                    action = "SERVE_SAVED"

                screenshot_prefix = "serve_saved" if save_after_fill else "serve_dry_run"
                screenshot_path = os.path.join("output", "screenshots", f"{screenshot_prefix}_{run_id}_{student.ssid}.png")
                try:
                    page.screenshot(path=screenshot_path, full_page=False)
                except Exception:
                    screenshot_path = ""

                verified_summary = "; ".join(
                    f"{r.get('rowLabel')}={r.get('value')} ({r.get('action')})"
                    for r in service_results
                )

                if save_after_fill:
                    details = (
                        "0.5's entered into 6 WAI fields.  "
                        f"Rows: {', '.join(label for label, _ in ARRAY_SERVICE_TARGETS)}; "
                        f"Verified: {verified_summary}"
                    )
                    _inc("serve_saved")
                else:
                    details = (
                        "Serve dry-run filled Array of Services WAI hours without saving. "
                        f"Rows: {', '.join(label for label, _ in ARRAY_SERVICE_TARGETS)}; "
                        f"Verified: {verified_summary}"
                    )
                    _inc("serve_dry_run_filled")

                _log(log_fn, f"[{run_id}] {action}: {student.ssid} ({display_name}) -> {details}")
                _user(f"{action}||display_name={display_name}", ui_hold_default)
                append_ledger_xlsx(ledger_path, {
                    "run_id": run_id,
                    "timestamp": datetime.utcnow().isoformat(),
                    "ssid": student.ssid,
                    "student_name": display_name,
                    "action": action,
                    "details": details,
                    "screenshot": screenshot_path,
                    "trace": "",
                })

                if trace_on_failure:
                    context.tracing.stop()

                if pause_after_first_fill:
                    pause_desc = "saved" if save_after_fill else "filled"
                    _log(log_fn, f"[{run_id}] SERVE {serve_mode_label} paused after first {pause_desc} owned student.")
                    if pause_fn:
                        pause_fn(page)
                    else:
                        input(f"{serve_mode_label} {pause_desc} the first owned student. Press Enter to close browser...")
                    break

            except Exception as exc:
                if trace_on_failure:
                    trace_path = os.path.join("output", "traces", f"trace_{run_id}_{student.ssid}.zip")
                    try:
                        context.tracing.stop(path=trace_path)
                    except Exception:
                        trace_path = ""

                try:
                    screenshot_path = os.path.join("output", "screenshots", f"serve_error_{run_id}_{student.ssid}.png")
                    page.screenshot(path=screenshot_path, full_page=True)
                except Exception:
                    screenshot_path = ""

                try:
                    close_transient_overlays(page)
                except Exception:
                    pass

                _log(log_fn, f"[{run_id}] SERVE_ERROR: {student.ssid} ({display_name}) -> {exc}")
                _user(f"SERVE_ERROR||display_name={display_name}", ui_hold_default)
                _inc("errors")
                append_ledger_xlsx(ledger_path, {
                    "run_id": run_id,
                    "timestamp": datetime.utcnow().isoformat(),
                    "ssid": student.ssid,
                    "student_name": display_name,
                    "action": "SERVE_ERROR",
                    "details": str(exc),
                    "screenshot": screenshot_path,
                    "trace": trace_path,
                })

                if stop_on_error:
                    raise
                if SERVE_LAYOUT_CHANGED_PREFIX in str(exc):
                    raise RuntimeError(str(exc))
                continue

        _log(log_fn, f"[{run_id}] {serve_mode_label} complete. Ledger: {ledger_path}")
        _user(serve_complete_message, 0.5)

        context.close()
        browser.close()

        print(serve_console_complete)
        return {"run_id": run_id, "ledger_path": ledger_path, **stats}


def run_address_patch(
    excel_path: str,
    config_path: str = "config/config.yaml",
    username: str | None = None,
    password: str | None = None,
    log_fn=None,
    override_headless: bool | None = None,
    override_slow_mo_ms: int | None = None,
    stop_event=None,
    user_log_fn=None,
    pause_after_first_patch: bool = False,
    pause_fn=None,
    save_after_fill: bool = False,
) -> Dict[str, Any]:
    cfg = load_config(config_path)

    ui_hold_default = float(cfg.get("ui", {}).get("min_status_seconds", 1.5))

    def _user(msg: str, min_hold_sec: float = 0.0):
        if user_log_fn:
            user_log_fn(msg)
        else:
            _log(log_fn, msg)
        hold = float(min_hold_sec or 0.0)
        if hold > 0:
            time.sleep(hold)

    def _should_stop() -> bool:
        return bool(stop_event and stop_event.is_set())

    if "search_project" not in cfg["workability"]:
        raise KeyError("config.yaml workability must include search_project.")

    ledger_path = cfg["ledger"]["path"]
    ensure_dir("output/screenshots")
    ensure_dir("output/traces")
    ensure_dir(os.path.dirname(ledger_path))

    run_id = str(uuid.uuid4())[:8]

    username_default = cfg.get("credentials", {}).get("username", "")
    if username is None or not str(username).strip():
        username = username_default or input("WAI Username (email): ").strip()

    if password is None:
        password = getpass("WAI Password: ")

    patch_mode_label = "Address patch save-run" if save_after_fill else "Address patch dry-run"
    patch_start_message = (
        "Starting Workabili-Bot3000 address patch save-run..."
        if save_after_fill
        else "Starting Workabili-Bot3000 address patch dry-run..."
    )
    patch_complete_message = (
        "Address patch completed. Check the ledger for details."
        if save_after_fill
        else "Address patch dry-run complete. Check the ledger for details."
    )
    patch_console_complete = f"{patch_mode_label} complete. Ledger: {ledger_path}"
    _log(log_fn, f"[{run_id}] Starting {patch_mode_label}. Excel: {excel_path}")
    _user(patch_start_message, 0.5)

    sanitization_log_path = os.path.join("output", "logs", "input_sanitization.log")
    _sanitization_log = make_input_audit_logger(sanitization_log_path, run_id, log_fn=log_fn)

    rows = read_excel_rows(
        excel_path,
        cfg=cfg,
        log_fn=_sanitization_log,
        require_address_columns=True,
    )

    stats = {
        "input_rows": len(rows),
        "address_dry_run_filled": 0,
        "address_patched": 0,
        "address_not_found": 0,
        "address_skipped_not_owned": 0,
        "skipped": 0,
        "errors": 0,
        "stopped": 0,
    }

    def _inc(key: str, n: int = 1):
        stats[key] = int(stats.get(key, 0)) + n

    if not rows:
        _user("No rows found in Excel.", 0.5)
        return {"run_id": run_id, "ledger_path": ledger_path, **stats}

    prepared_rows, invalid_rows = prepare_input_rows(
        rows,
        cfg,
        validation_log_fn=_sanitization_log,
        require_address_fields=True,
    )

    for invalid in invalid_rows:
        _log(
            log_fn,
            f"[{run_id}] {invalid.action}: row {invalid.excel_row_num} "
            f"{invalid.ssid or '(no SSID)'} ({invalid.display_name}) -> {invalid.details}",
        )
        _inc("skipped")
        _user(f"{invalid.action}||display_name={invalid.display_name}", ui_hold_default)
        append_ledger_xlsx(ledger_path, {
            "run_id": run_id,
            "timestamp": datetime.utcnow().isoformat(),
            "ssid": invalid.ssid,
            "student_name": invalid.display_name,
            "action": invalid.action,
            "details": invalid.details,
            "screenshot": "",
            "trace": "",
        })

    if not prepared_rows:
        _user("No valid rows found in Excel.", 0.5)
        _log(log_fn, f"[{run_id}] No valid address-patch rows found. Ledger: {ledger_path}")
        return {"run_id": run_id, "ledger_path": ledger_path, **stats}

    total = len(prepared_rows)

    headless = bool(cfg.get("run", {}).get("headless", False))
    if override_headless is not None:
        headless = bool(override_headless)

    slow_mo_ms = int(cfg.get("run", {}).get("slow_mo_ms", 0))
    if override_slow_mo_ms is not None:
        slow_mo_ms = int(override_slow_mo_ms)
    trace_on_failure = bool(cfg.get("trace", {}).get("on_failure", True))
    stop_on_error = bool(cfg.get("run", {}).get("stop_on_error", False))

    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless, slow_mo=slow_mo_ms if slow_mo_ms > 0 else None)
        context = browser.new_context(viewport={"width": 1600, "height": 1400})
        page = context.new_page()

        try:
            login(page, username, password, cfg, log_fn=log_fn)
        except Exception as e:
            _log(log_fn, f"[{run_id}] FATAL LOGIN ERROR -> {e}")
            try:
                context.close()
            except Exception:
                pass
            try:
                browser.close()
            except Exception:
                pass
            raise

        if _should_stop():
            _user("Stopped by user.", 0.5)
            context.close()
            browser.close()
            return {"run_id": run_id, "ledger_path": ledger_path, **stats}

        _log(log_fn, f"[{run_id}] Logged in. Opening Student Records for {patch_mode_label}...")
        _user("Logged in. Opening Student Records...", 0.5)
        goto_student_records(page, cfg=cfg, log_fn=log_fn, run_id=run_id)

        for i, prepared in enumerate(prepared_rows):
            if _should_stop():
                _user("Stopped by user.", 0.5)
                _inc("stopped")
                break

            student = prepared.student
            display_name = prepared.display_name
            trace_path = ""
            screenshot_path = ""

            _user(f"🔎 {i+1} of {total}: Patching {display_name}...", 0.5)

            try:
                if trace_on_failure:
                    context.tracing.start(screenshots=True, snapshots=True, sources=True)

                _log(log_fn, f"[{run_id}] ADDRESS processing SSID {student.ssid} ({display_name})")

                goto_student_records(page, cfg=cfg, log_fn=log_fn, run_id=run_id)
                _log(log_fn, f"[{run_id}] ADDRESS searching SSID {student.ssid} under '{cfg['workability']['search_project']}'...")
                search_by_ssid(page, student.ssid, cfg)

                outcome = determine_search_outcome(page, timeout_ms=5000)
                if outcome != "FOUND":
                    details = "Student not found in WAI; address patch skipped."
                    _log(log_fn, f"[{run_id}] ADDRESS_NOT_FOUND: {student.ssid} ({display_name})")
                    _user(f"ADDRESS_NOT_FOUND||display_name={display_name}", ui_hold_default)
                    _inc("address_not_found")
                    append_ledger_xlsx(ledger_path, {
                        "run_id": run_id,
                        "timestamp": datetime.utcnow().isoformat(),
                        "ssid": student.ssid,
                        "student_name": display_name,
                        "action": "ADDRESS_NOT_FOUND",
                        "details": details,
                        "screenshot": "",
                        "trace": "",
                    })
                    if trace_on_failure:
                        context.tracing.stop()
                    continue

                owning_org = get_found_row_owning_org(page, student.ssid)
                _log(log_fn, f"[{run_id}] ADDRESS owning org for {student.ssid}: {owning_org!r}")

                if not is_already_owned_by_us(owning_org, cfg):
                    details = f"Student found in WAI but not owned by configured org; owning org={owning_org!r}."
                    _log(log_fn, f"[{run_id}] ADDRESS_SKIPPED_NOT_OWNED: {student.ssid} ({display_name}) -> {details}")
                    _user(f"ADDRESS_SKIPPED_NOT_OWNED||display_name={display_name}", ui_hold_default)
                    _inc("address_skipped_not_owned")
                    append_ledger_xlsx(ledger_path, {
                        "run_id": run_id,
                        "timestamp": datetime.utcnow().isoformat(),
                        "ssid": student.ssid,
                        "student_name": display_name,
                        "action": "ADDRESS_SKIPPED_NOT_OWNED",
                        "details": details,
                        "screenshot": "",
                        "trace": "",
                    })
                    if trace_on_failure:
                        context.tracing.stop()
                    continue

                _log(log_fn, f"[{run_id}] ADDRESS opening Edit for owned student {student.ssid} ({display_name})")
                open_existing_student_edit(page)
                open_addresses_tab(page, log_fn=log_fn, run_id=run_id)
                patch_results = overwrite_existing_student_address(page, student, log_fn=log_fn, run_id=run_id)

                action = "ADDRESS_DRY_RUN_FILLED"
                if save_after_fill:
                    _log(log_fn, f"[{run_id}] ADDRESS saving contact fields for {student.ssid} ({display_name})")
                    click_save_robust(
                        page,
                        timeout_ms=15000,
                        log_fn=log_fn,
                        run_id=run_id,
                        toast_regex=re.compile(r"(successfully\s+saved|address\s+saved\s+successfully)", re.I),
                    )
                    page.wait_for_timeout(1000)
                    action = "ADDRESS_PATCHED"

                screenshot_prefix = "address_patched" if save_after_fill else "address_dry_run"
                screenshot_path = os.path.join("output", "screenshots", f"{screenshot_prefix}_{run_id}_{student.ssid}.png")
                try:
                    page.screenshot(path=screenshot_path, full_page=False)
                except Exception:
                    screenshot_path = ""

                details = (
                    f"Street={patch_results['street_address']}; City={patch_results['city']}; "
                    f"State={patch_results['state']}; Zip={patch_results['zip_code']}; "
                    f"Phone={patch_results['phone_number']}; Parent={patch_results['parent_name']}"
                )

                if save_after_fill:
                    _inc("address_patched")
                else:
                    _inc("address_dry_run_filled")

                _log(log_fn, f"[{run_id}] {action}: {student.ssid} ({display_name}) -> {details}")
                _user(f"{action}||display_name={display_name}", ui_hold_default)
                append_ledger_xlsx(ledger_path, {
                    "run_id": run_id,
                    "timestamp": datetime.utcnow().isoformat(),
                    "ssid": student.ssid,
                    "student_name": display_name,
                    "action": action,
                    "details": details,
                    "screenshot": screenshot_path,
                    "trace": "",
                })

                if trace_on_failure:
                    context.tracing.stop()

                if pause_after_first_patch:
                    patch_desc = "saved" if save_after_fill else "filled"
                    _log(log_fn, f"[{run_id}] ADDRESS {patch_mode_label} paused after first {patch_desc} owned student.")
                    if pause_fn:
                        pause_fn(page)
                    else:
                        input(f"{patch_mode_label} {patch_desc} the first owned student. Press Enter to close browser...")
                    break

            except Exception as exc:
                if trace_on_failure:
                    trace_path = os.path.join("output", "traces", f"trace_{run_id}_{student.ssid}.zip")
                    try:
                        context.tracing.stop(path=trace_path)
                    except Exception:
                        trace_path = ""

                try:
                    screenshot_path = os.path.join("output", "screenshots", f"address_error_{run_id}_{student.ssid}.png")
                    page.screenshot(path=screenshot_path, full_page=True)
                except Exception:
                    screenshot_path = ""

                try:
                    close_transient_overlays(page)
                except Exception:
                    pass

                _log(log_fn, f"[{run_id}] ADDRESS_ERROR: {student.ssid} ({display_name}) -> {exc}")
                _user(f"ADDRESS_ERROR||display_name={display_name}", ui_hold_default)
                _inc("errors")
                append_ledger_xlsx(ledger_path, {
                    "run_id": run_id,
                    "timestamp": datetime.utcnow().isoformat(),
                    "ssid": student.ssid,
                    "student_name": display_name,
                    "action": "ADDRESS_ERROR",
                    "details": str(exc),
                    "screenshot": screenshot_path,
                    "trace": trace_path,
                })

                if stop_on_error:
                    raise
                continue

        _log(log_fn, f"[{run_id}] {patch_mode_label} complete. Ledger: {ledger_path}")
        _user(patch_complete_message, 0.5)

        context.close()
        browser.close()

        print(patch_console_complete)
        return {"run_id": run_id, "ledger_path": ledger_path, **stats}


if __name__ == "__main__":
    excel = input("Path to input Excel file: ").strip('"').strip()
    run(excel_path=excel, config_path="config/config.yaml")
